"""Opening Stock: quantity file + previous-year Closing Stock product amounts."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from app.engines.financials_engine.config.product_rule_book import (
    map_pivots_to_closing_stock_categories,
)
from app.engines.financials_engine.engine.audit import validated_opening_to_pivot
from app.engines.financials_engine.engine.opening_stock import (
    build_opening_measures_for_layout,
    validate_opening_stock,
)
from app.engines.financials_engine.parsers.opening_stock_loader import (
    load_previous_year_product_sheets,
)


def _prev_year_category_workbook_bytes() -> bytes:
    """Mirrors Eximp Dia/Eme Closing Stock layout (product rows, not one sheet per SKU)."""
    wb = Workbook()
    dia = wb.active
    dia.title = 'Dia'
    dia.append(['DETAILS OF JEWELS CLOSING STOCK - DIAMONDS'])
    dia.append(
        [
            'Particulars',
            'Opening stock',
            None,
            'Purchases',
            None,
            'Closing stock',
            None,
        ]
    )
    dia.append([1, '2  (Qty)', '3  (Amt.)', '4  (Qty)', '5  (Amt.)', '15 (Qty)', '16  (Amt.)'])
    dia.append(['Diamonds - Beads'])
    dia.append(['Di. Beads', 263.03, 100.0, 0, 0, 177.86, 234713.15])
    dia.append(['TOTAL', 263.03, 100.0, 0, 0, 177.86, 234713.15])
    dia.append(['Polki', 5.5, 50.0, 0, 0, 5.5, 200.0])

    eme = wb.create_sheet('Eme')
    eme.append(['Particulars', 'Opening stock', None, 'Closing stock', None])
    eme.append([1, 'Qty', 'Amt.', 'Qty', 'Amt.'])
    eme.append(['JEM 100', 10, 20, 8, 999.5])

    bs = wb.create_sheet('Balance Sheet')
    bs.append(['Particulars', 'Closing stock', None])
    bs.append(['Something', 1, 999999])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _prev_year_product_sheet_workbook_bytes() -> bytes:
    """Dedicated product sheet named after the product."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Chakri'
    ws.append(['Particulars', 'Opening stock', None, 'Closing stock', None])
    ws.append([1, 'Qty', 'Amt.', 'Qty', 'Amt.'])
    ws.append(['Receipts', 0, 0, 0, 0])
    ws.append(['Issues', 0, 0, 0, 0])
    ws.append(['Closing Balance', 12.5, 0, 12.5, 45678.9])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestOpeningStockProductSheetMapping:
    def test_indexes_products_from_category_sheets_skips_totals(self):
        index = load_previous_year_product_sheets(
            _prev_year_category_workbook_bytes(),
            'prev.xlsx',
        )
        assert 'di. beads' in index
        assert index['di. beads']['closingStockAmount'] == 234713.15
        assert index['di. beads']['sheetName'] == 'Dia'
        assert index['polki']['closingStockAmount'] == 200.0
        assert index['jem 100']['closingStockAmount'] == 999.5
        assert 'total' not in index
        assert 'something' not in index

    def test_indexes_dedicated_product_sheet_by_sheet_name(self):
        index = load_previous_year_product_sheets(
            _prev_year_product_sheet_workbook_bytes(),
            'prev.xlsx',
        )
        assert 'chakri' in index
        assert index['chakri']['closingStockAmount'] == 45678.9
        assert index['chakri']['sheetName'] == 'Chakri'

    def test_maps_qty_from_quantity_file_and_amount_from_closing_stock(self):
        sheets = load_previous_year_product_sheets(
            _prev_year_category_workbook_bytes(),
            'prev.xlsx',
        )
        result = validate_opening_stock(
            quantity_rows=[
                {'product': 'Di. beads', 'openingBalance': 263.03},
                {'product': 'Polki', 'openingBalance': 5.5},
                {'product': 'Emeralds JEM 100', 'openingBalance': 10},
            ],
            previous_year_sheets=sheets,
        )
        report = result['report']
        assert report['matchedCount'] == 3
        validated = {r['product']: r for r in result['validatedOpening']}
        assert validated['Di. beads']['openingQty'] == 263.03
        assert validated['Di. beads']['openingAmt'] == 234713.15
        assert validated['Polki']['openingAmt'] == 200.0
        assert validated['Emeralds JEM 100']['openingAmt'] == 999.5

    def test_unmatched_when_product_missing(self):
        sheets = load_previous_year_product_sheets(
            _prev_year_category_workbook_bytes(),
            'prev.xlsx',
        )
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Missing Product', 'openingBalance': 3}],
            previous_year_sheets=sheets,
        )
        assert result['report']['matchedCount'] == 0
        assert result['report']['previousYearMappingRequiredCount'] == 1
        row = result['validatedOpening'][0]
        assert row['openingQty'] == 3
        assert row['openingAmt'] is None
        assert row['reason'] == 'Previous Year Mapping Required'

    def test_maps_validated_opening_into_closing_stock_layout_by_name(self):
        opening_pivot = validated_opening_to_pivot(
            [
                {
                    'product': 'Di. beads',
                    'openingQty': 263.03,
                    'openingAmt': 234713.15,
                    'status': 'matched',
                },
                {
                    'product': 'Emeralds JEM 100',
                    'openingQty': 10,
                    'openingAmt': 999.5,
                    'status': 'matched',
                },
            ]
        )
        mapped = map_pivots_to_closing_stock_categories(
            sales_pivot=[],
            purchases_pivot=[],
            opening_pivot=opening_pivot,
            rule_book={
                'Diamond': {
                    'Diamonds - Beads': ['Di. Beads'],
                    'Diamonds': [],
                },
                'Emerald': ['JEM 100'],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        beads = next(
            row
            for row in mapped['layoutByCategory']['Diamond']
            if row.get('kind') == 'product' and row.get('label') == 'Di. Beads'
        )
        jem = next(
            row
            for row in mapped['layoutByCategory']['Emerald']
            if row.get('kind') == 'product' and row.get('label') == 'JEM 100'
        )
        assert beads['openingQty'] == 263.03
        assert beads['openingAmt'] == 234713
        assert jem['openingQty'] == 10
        assert jem['openingAmt'] == 1000
        assert mapped['productsWithOpeningData'] == 2

    def test_build_opening_measures_for_layout_uses_name_keys_not_rule_book(self):
        by_display, unmapped = build_opening_measures_for_layout(
            [
                {'product': 'Emeralds JEM 100', 'sumOfQuantity': 10, 'sumOfGross': 999.5},
            ],
            ['JEM 100', 'Unrelated Product'],
        )
        assert unmapped == []
        assert by_display['JEM 100']['sumOfQuantity'] == 10
        assert by_display['JEM 100']['sumOfGross'] == 999.5
        assert 'Unrelated Product' not in by_display

    def test_quantity_typo_sythetic_jsy_300_maps_to_layout(self):
        opening_pivot = validated_opening_to_pivot(
            [
                {
                    'product': 'Sythetic JSY 300',
                    'openingQty': 42.5,
                    'openingAmt': None,
                    'status': 'unmatched',
                }
            ]
        )
        mapped = map_pivots_to_closing_stock_categories(
            sales_pivot=[],
            purchases_pivot=[],
            opening_pivot=opening_pivot,
            rule_book={
                'Diamond': [],
                'Emerald': [],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [
                        'Synthetic JSY 50',
                        'Synthetic JSY 150',
                        'Synthetic JSY 300',
                        'Synthetic JSY 100',
                    ],
                },
            },
        )
        row = next(
            item
            for item in mapped['layoutByCategory']['Precious and Semi Precious']
            if item.get('kind') == 'product' and item.get('label') == 'Synthetic JSY 300'
        )
        assert row['openingQty'] == 42.5
        sibling = next(
            item
            for item in mapped['layoutByCategory']['Precious and Semi Precious']
            if item.get('kind') == 'product' and item.get('label') == 'Synthetic JSY 50'
        )
        assert sibling['openingQty'] is None


def _fallback_workbook_bytes() -> bytes:
    """Prev-year FP1 / FP 1 rows; qty file uses Rule Book name Flat polki FP 1."""
    wb = Workbook()
    dia = wb.active
    dia.title = 'Dia'
    dia.append(
        [
            'Particulars',
            'Opening stock',
            None,
            'Purchases',
            None,
            'Closing stock',
            None,
        ]
    )
    dia.append([1, 'Qty', 'Amt.', 'Qty', 'Amt.', 'Qty', 'Amt.'])
    dia.append(['Diamonds - Flat polki'])
    dia.append(['FP1', 0, 0, 0, 0, 5.0, 100.0])
    dia.append(['FP 1', 0, 0, 0, 0, 5.0, 200.0])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _rosecut_workbook_bytes() -> bytes:
    """Dia Rosecut section with RC1 / RC 2 / RC 10 previous-year product names."""
    wb = Workbook()
    dia = wb.active
    dia.title = 'Dia'
    dia.append(
        [
            'Particulars',
            'Opening stock',
            None,
            'Purchases',
            None,
            'Closing stock',
            None,
        ]
    )
    dia.append([1, 'Qty', 'Amt.', 'Qty', 'Amt.', 'Qty', 'Amt.'])
    dia.append(['Diamonds Rosecut diamonds'])
    dia.append(['RC1', 0, 0, 0, 0, 12.5, 500.0])
    dia.append(['RC 2', 0, 0, 0, 0, 8.0, 300.0])
    dia.append(['RC 10', 0, 0, 0, 0, 99.0, 9999.0])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _diamond_rosecut_rule_book() -> dict:
    return {
        'Diamond': {
            'Diamonds Rosecut diamonds': ['Di. RC 1', 'Di. RC 2', 'Di. RC 10'],
        },
        'Emerald': [],
        'Pearls': [],
        'Rubie': [],
        'Precious and Semi Precious': {
            'Precious Stones': [],
            'Semi Precious': [],
            'Synthetic Stones': [],
        },
    }


def _chakri_variant_workbook_bytes() -> bytes:
    """Uncut - diamonds: Chakri a + Chakri b sum to current Chakri Opening Balance."""
    wb = Workbook()
    dia = wb.active
    dia.title = 'Dia'
    dia.append(
        [
            'Particulars',
            'Opening stock',
            None,
            'Purchases',
            None,
            'Closing stock',
            None,
        ]
    )
    dia.append([1, 'Qty', 'Amt.', 'Qty', 'Amt.', 'Qty', 'Amt.'])
    dia.append(['Uncut - diamonds'])
    dia.append(['Chakri a', 0, 0, 0, 0, 100.0, 1000.0])
    dia.append(['Chakri b', 0, 0, 0, 0, 77.26, 2000.0])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _chakri_orphan_workbook_bytes() -> bytes:
    """Uncut - diamonds: Chakri qty from Old Product A + B + C."""
    wb = Workbook()
    dia = wb.active
    dia.title = 'Dia'
    dia.append(
        [
            'Particulars',
            'Opening stock',
            None,
            'Purchases',
            None,
            'Closing stock',
            None,
        ]
    )
    dia.append([1, 'Qty', 'Amt.', 'Qty', 'Amt.', 'Qty', 'Amt.'])
    dia.append(['Uncut - diamonds'])
    dia.append(['Old Product A', 0, 0, 0, 0, 40.0, 1000.0])
    dia.append(['Old Product B', 0, 0, 0, 0, 35.0, 2000.0])
    dia.append(['Old Product C', 0, 0, 0, 0, 25.0, 3000.0])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestOpeningStockSubcategoryFallback:
    def test_chakri_sums_orphan_previous_products_when_qty_reconciles(self):
        from app.engines.financials_engine.parsers.opening_stock_loader import (
            load_previous_year_opening_stock,
        )

        payload = load_previous_year_opening_stock(_chakri_orphan_workbook_bytes(), 'prev.xlsx')
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Chakri', 'openingBalance': 100.0}],
            previous_year_sheets=payload['productIndex'],
            subcategory_products=payload['subcategoryProducts'],
            sheet_products=payload['sheetProducts'],
            rule_book={
                'Diamond': {
                    'Uncut - diamonds': ['Chakri', 'Polki'],
                    'Diamonds - Beads': [],
                },
                'Emerald': [],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        report = result['report']
        assert report['fallbackMatchedCount'] == 1
        row = result['validatedOpening'][0]
        assert row['openingQty'] == 100.0
        assert row['openingAmt'] == 6000.0
        assert row['status'] == 'matched_fallback'
        assert set(row['previousYearProducts']) == {'Old Product A', 'Old Product B', 'Old Product C'}

    def test_chakri_matches_chakri_a_and_chakri_b_variants(self):
        from app.engines.financials_engine.parsers.opening_stock_loader import (
            load_previous_year_opening_stock,
        )

        payload = load_previous_year_opening_stock(_chakri_variant_workbook_bytes(), 'prev.xlsx')
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Chakri', 'openingBalance': 177.26}],
            previous_year_sheets=payload['productIndex'],
            subcategory_products=payload['subcategoryProducts'],
            sheet_products=payload['sheetProducts'],
            rule_book={
                'Diamond': {'Uncut - diamonds': ['Chakri', 'Polki']},
                'Emerald': [],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        row = result['validatedOpening'][0]
        assert row['status'] == 'matched_fallback'
        assert row['openingQty'] == 177.26
        assert row['openingAmt'] == 3000.0
        assert set(row['previousYearProducts']) == {'Chakri a', 'Chakri b'}

    def test_fallback_sums_multiple_previous_products_when_qty_matches(self):
        from app.engines.financials_engine.parsers.opening_stock_loader import (
            load_previous_year_opening_stock,
        )

        payload = load_previous_year_opening_stock(_fallback_workbook_bytes(), 'prev.xlsx')
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Flat polki FP 1', 'openingBalance': 10.0}],
            previous_year_sheets=payload['productIndex'],
            subcategory_products=payload['subcategoryProducts'],
            sheet_products=payload['sheetProducts'],
        )
        report = result['report']
        assert report['exactMatchedCount'] == 0
        assert report['fallbackMatchedCount'] == 1
        assert report['matchedCount'] == 1
        row = result['validatedOpening'][0]
        assert row['openingQty'] == 10.0
        assert row['openingAmt'] == 300.0
        assert row['status'] == 'matched_fallback'
        assert row['ruleBookProduct'] == 'Flat polki FP 1'
        assert set(row['previousYearProducts']) == {'FP1', 'FP 1'}

    def test_fallback_quantity_mismatch(self):
        from app.engines.financials_engine.parsers.opening_stock_loader import (
            load_previous_year_opening_stock,
        )

        payload = load_previous_year_opening_stock(_fallback_workbook_bytes(), 'prev.xlsx')
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Flat polki FP 1', 'openingBalance': 9.0}],
            previous_year_sheets=payload['productIndex'],
            subcategory_products=payload['subcategoryProducts'],
            sheet_products=payload['sheetProducts'],
        )
        assert result['report']['quantityMismatchCount'] == 1
        assert result['report']['fallbackMatchedCount'] == 0
        row = result['validatedOpening'][0]
        assert row['status'] == 'quantity_mismatch'
        assert row['reason'] == 'Quantity Mismatch'
        assert row['openingAmt'] is None

    def test_exact_match_untouched_when_primary_resolves(self):
        from app.engines.financials_engine.parsers.opening_stock_loader import (
            load_previous_year_opening_stock,
        )

        payload = load_previous_year_opening_stock(
            _prev_year_category_workbook_bytes(),
            'prev.xlsx',
        )
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Di. beads', 'openingBalance': 263.03}],
            previous_year_sheets=payload['productIndex'],
            subcategory_products=payload['subcategoryProducts'],
            sheet_products=payload['sheetProducts'],
        )
        assert result['report']['exactMatchedCount'] == 1
        assert result['report']['fallbackMatchedCount'] == 0
        assert result['validatedOpening'][0]['status'] == 'matched'

    def test_fallback_resolves_into_closing_stock_layout(self):
        from app.engines.financials_engine.parsers.opening_stock_loader import (
            load_previous_year_opening_stock,
        )

        payload = load_previous_year_opening_stock(_fallback_workbook_bytes(), 'prev.xlsx')
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Flat polki FP 1', 'openingBalance': 10.0}],
            previous_year_sheets=payload['productIndex'],
            subcategory_products=payload['subcategoryProducts'],
            sheet_products=payload['sheetProducts'],
        )
        opening_pivot = validated_opening_to_pivot(result['validatedOpening'])
        mapped = map_pivots_to_closing_stock_categories(
            opening_pivot=opening_pivot,
            rule_book={
                'Diamond': {
                    'Diamonds - Flat polki': ['Flat polki FP 1'],
                },
                'Emerald': [],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        product = next(
            row
            for row in mapped['layoutByCategory']['Diamond']
            if row.get('kind') == 'product' and row.get('label') == 'Flat polki FP 1'
        )
        assert product['openingQty'] == 10.0
        assert product['openingAmt'] == 300

    def test_fallback_opening_written_to_layout_by_rule_book_product(self):
        """Fallback row with non-matching qty-file label still fills the Rule Book row."""
        opening_pivot = validated_opening_to_pivot(
            [
                {
                    'product': 'qty-file-label-not-on-layout',
                    'ruleBookProduct': 'Chakri',
                    'category': 'Diamond',
                    'subcategory': 'Uncut - diamonds',
                    'openingQty': 100.0,
                    'openingAmt': 6000.0,
                    'status': 'matched_fallback',
                }
            ]
        )
        mapped = map_pivots_to_closing_stock_categories(
            opening_pivot=opening_pivot,
            rule_book={
                'Diamond': {
                    'Uncut - diamonds': ['Chakri', 'Polki'],
                },
                'Emerald': [],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        chakri = next(
            item
            for item in mapped['layoutByCategory']['Diamond']
            if item.get('kind') == 'product' and item.get('label') == 'Chakri'
        )
        assert chakri['openingQty'] == 100.0
        assert chakri['openingAmt'] == 6000

    def test_chakri_orphan_fallback_end_to_end_on_layout(self):
        from app.engines.financials_engine.parsers.opening_stock_loader import (
            load_previous_year_opening_stock,
        )

        payload = load_previous_year_opening_stock(_chakri_orphan_workbook_bytes(), 'prev.xlsx')
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Chakri', 'openingBalance': 100.0}],
            previous_year_sheets=payload['productIndex'],
            subcategory_products=payload['subcategoryProducts'],
            sheet_products=payload['sheetProducts'],
            rule_book={
                'Diamond': {'Uncut - diamonds': ['Chakri', 'Polki']},
                'Emerald': [],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        opening_pivot = validated_opening_to_pivot(result['validatedOpening'])
        mapped = map_pivots_to_closing_stock_categories(
            opening_pivot=opening_pivot,
            rule_book={
                'Diamond': {'Uncut - diamonds': ['Chakri', 'Polki']},
                'Emerald': [],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        chakri = next(
            item
            for item in mapped['layoutByCategory']['Diamond']
            if item.get('kind') == 'product' and item.get('label') == 'Chakri'
        )
        assert chakri['openingQty'] == 100.0
        assert chakri['openingAmt'] == 6000


class TestOpeningStockRosecutFallback:
    def test_di_rc_1_matches_previous_rc1_token(self):
        from app.engines.financials_engine.parsers.opening_stock_loader import (
            load_previous_year_opening_stock,
        )

        payload = load_previous_year_opening_stock(_rosecut_workbook_bytes(), 'prev.xlsx')
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Di. RC 1', 'openingBalance': 12.5}],
            previous_year_sheets=payload['productIndex'],
            subcategory_products=payload['subcategoryProducts'],
            sheet_products=payload['sheetProducts'],
            rule_book=_diamond_rosecut_rule_book(),
        )
        row = result['validatedOpening'][0]
        assert row['status'] == 'matched_fallback'
        assert row['openingQty'] == 12.5
        assert row['openingAmt'] == 500.0
        assert row['previousYearProducts'] == ['RC1']
        assert row['subcategory'] == 'Diamonds Rosecut diamonds'

    def test_di_rc_1_does_not_match_rc_10(self):
        from app.engines.financials_engine.parsers.opening_stock_loader import (
            load_previous_year_opening_stock,
        )

        payload = load_previous_year_opening_stock(_rosecut_workbook_bytes(), 'prev.xlsx')
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Di. RC 1', 'openingBalance': 99.0}],
            previous_year_sheets=payload['productIndex'],
            subcategory_products=payload['subcategoryProducts'],
            sheet_products=payload['sheetProducts'],
            rule_book=_diamond_rosecut_rule_book(),
        )
        row = result['validatedOpening'][0]
        assert row['status'] == 'quantity_mismatch'
        assert row['previousClosingQty'] == 12.5

    def test_di_rc_2_matches_previous_rc_2_token(self):
        from app.engines.financials_engine.parsers.opening_stock_loader import (
            load_previous_year_opening_stock,
        )

        payload = load_previous_year_opening_stock(_rosecut_workbook_bytes(), 'prev.xlsx')
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Di. RC 2', 'openingBalance': 8.0}],
            previous_year_sheets=payload['productIndex'],
            subcategory_products=payload['subcategoryProducts'],
            sheet_products=payload['sheetProducts'],
            rule_book=_diamond_rosecut_rule_book(),
        )
        row = result['validatedOpening'][0]
        assert row['status'] == 'matched_fallback'
        assert row['openingAmt'] == 300.0
        assert row['previousYearProducts'] == ['RC 2']

    def test_rosecut_fallback_written_to_closing_stock_layout(self):
        from app.engines.financials_engine.parsers.opening_stock_loader import (
            load_previous_year_opening_stock,
        )

        payload = load_previous_year_opening_stock(_rosecut_workbook_bytes(), 'prev.xlsx')
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Di. RC 1', 'openingBalance': 12.5}],
            previous_year_sheets=payload['productIndex'],
            subcategory_products=payload['subcategoryProducts'],
            sheet_products=payload['sheetProducts'],
            rule_book=_diamond_rosecut_rule_book(),
        )
        opening_pivot = validated_opening_to_pivot(result['validatedOpening'])
        mapped = map_pivots_to_closing_stock_categories(
            opening_pivot=opening_pivot,
            rule_book=_diamond_rosecut_rule_book(),
        )
        rc1 = next(
            item
            for item in mapped['layoutByCategory']['Diamond']
            if item.get('kind') == 'product' and item.get('label') == 'Di. RC 1'
        )
        assert rc1['openingQty'] == 12.5
        assert rc1['openingAmt'] == 500


class TestOpeningStockSubcategoryFallbackContinued:
    def test_previous_year_mapping_required_when_no_subcategory_rows(self):
        result = validate_opening_stock(
            quantity_rows=[{'product': 'Chakri', 'openingBalance': 5.0}],
            previous_year_sheets={},
            subcategory_products={},
            sheet_products={},
            rule_book={
                'Diamond': {'Uncut - diamonds': ['Chakri', 'Polki']},
                'Emerald': [],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        assert result['report']['previousYearMappingRequiredCount'] == 1
        row = result['validatedOpening'][0]
        assert row['reason'] == 'Previous Year Mapping Required'
        assert row['openingAmt'] is None
