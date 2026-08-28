"""Closing Stock blank template structure tests."""

from io import BytesIO

from openpyxl import load_workbook

from app.engines.financials_engine.engine.closing_stock_template import (
    CLOSING_STOCK_CATEGORIES,
    LEAF_COLUMNS,
    build_closing_stock_template_bytes,
    build_pivots_workbook_bytes,
    closing_stock_report_title,
)


class TestClosingStockTemplate:
    def test_five_category_sheets_and_titles(self):
        raw = build_closing_stock_template_bytes(
            products_by_category={
                'Diamond': ['Gold Ring'],
                'Emerald': ['Gold Chain'],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': [],
            },
            layout_by_category={
                'Diamond': [
                    {'kind': 'subcategory', 'label': 'Diamonds'},
                    {'kind': 'product', 'label': 'Gold Ring'},
                    {'kind': 'subcategory_total', 'label': 'TOTAL'},
                    {'kind': 'grand_total', 'label': 'GRAND TOTAL'},
                ],
                'Emerald': [
                    {'kind': 'product', 'label': 'Gold Chain'},
                    {'kind': 'grand_total', 'label': 'GRAND TOTAL'},
                ],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': [],
            },
            company_name='Sample Jewellers',
            address='Hyderabad',
            financial_year='AY 2025-26',
        )
        wb = load_workbook(BytesIO(raw))
        assert wb.sheetnames == list(CLOSING_STOCK_CATEGORIES)

        for category in CLOSING_STOCK_CATEGORIES:
            ws = wb[category]
            assert ws['A1'].value == 'Sample Jewellers'
            assert ws['A2'].value == 'Hyderabad'
            assert 'AY 2025-26' in str(ws['A3'].value)
            assert ws['A4'].value == closing_stock_report_title(category)
            assert 'DIAMONDS' not in str(ws['A4'].value)
            assert ws['A6'].value == 'Particulars / Product'

        assert wb['Diamond'].cell(row=11, column=1).value == 'Gold Ring'
        assert wb['Emerald'].cell(row=10, column=1).value == 'Gold Chain'
        assert wb['Diamond']['A4'].value == 'DETAILS OF JEWELS CLOSING STOCK - DIAMOND'
        assert wb['Emerald']['A4'].value == 'DETAILS OF JEWELS CLOSING STOCK - EMERALD'
        assert (
            wb['Precious and Semi Precious']['A4'].value
            == 'DETAILS OF JEWELS CLOSING STOCK - PRECIOUS AND SEMI PRECIOUS'
        )

    def test_multi_level_headers_receipts_and_issues(self):
        raw = build_closing_stock_template_bytes(
            products_by_category={
                'Diamond': ['Ring'],
                'Emerald': [],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': [],
            }
        )
        wb = load_workbook(BytesIO(raw))
        ws = wb['Diamond']

        assert ws.cell(row=6, column=2).value == 'Opening Stock'
        assert ws.cell(row=6, column=6).value == 'Receipts'
        assert ws.cell(row=7, column=6).value == 'Internal Stock Transfer'
        assert ws.cell(row=7, column=8).value == 'Jubilee Hills'
        assert ws.cell(row=7, column=10).value == 'Kokapet'
        assert ws.cell(row=7, column=12).value == 'Total'

        assert ws.cell(row=6, column=15).value == 'Issues'
        assert ws.cell(row=7, column=15).value == 'Internal Stock Transfer'
        assert ws.cell(row=7, column=17).value == 'Banjara Hills'
        assert ws.cell(row=7, column=19).value == 'Kokapet'
        assert ws.cell(row=7, column=21).value == 'Total'

        assert ws.cell(row=8, column=2).value == 'Qty'
        assert ws.cell(row=8, column=3).value == 'Amt.'
        assert ws.cell(row=9, column=2).value == '1'
        assert ws.cell(row=9, column=1 + len(LEAF_COLUMNS)).value == str(len(LEAF_COLUMNS))

    def test_receipts_and_issues_category_order(self):
        receipt_l2 = [path[1] for path, _ in LEAF_COLUMNS if path[0] == 'Receipts']
        receipt_ordered = []
        for name in receipt_l2:
            if name and name not in receipt_ordered:
                receipt_ordered.append(name)
        assert receipt_ordered == [
            'Internal Stock Transfer',
            'Jubilee Hills',
            'Kokapet',
            'Total',
        ]

        issues_l2 = [path[1] for path, _ in LEAF_COLUMNS if path[0] == 'Issues']
        issues_ordered = []
        for name in issues_l2:
            if name and name not in issues_ordered:
                issues_ordered.append(name)
        assert issues_ordered == [
            'Internal Stock Transfer',
            'Banjara Hills',
            'Kokapet',
            'Total',
        ]

    def test_pivots_workbook_two_sheets(self):
        raw = build_pivots_workbook_bytes(
            sales_pivot=[{'product': 'A', 'sumOfQuantity': 1, 'sumOfGross': 10}],
            purchases_pivot=[{'product': 'B', 'sumOfQuantity': 2, 'sumOfGross': 20}],
        )
        wb = load_workbook(BytesIO(raw))
        assert wb.sheetnames == ['Sales Pivot', 'Purchases Pivot']
        assert wb['Sales Pivot']['A2'].value == 'A'
        assert wb['Purchases Pivot']['A2'].value == 'B'
