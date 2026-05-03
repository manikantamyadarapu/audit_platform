"""Large-workbook behaviour: counts, performance payload, optional row cap."""

from io import BytesIO

import pytest
from openpyxl import Workbook

from app.config.settings import get_settings
from app.processors.gross_weight_processor import GrossWeightProcessor
from app.processors.pan_processor import PanProcessor


def _gw_bytes_many_tabular_rows(n_data: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'SNo'
    ws['B1'] = 'Manual Gross wt.'
    ws['C1'] = 'Auto Gross Wt.'
    ws['D1'] = 'Difference in Gross wt.'
    for i in range(n_data):
        r = 2 + i
        ws.cell(row=r, column=1, value=f'Voucher No: T{i}')
        ws.cell(row=r, column=2, value=1.0)
        ws.cell(row=r, column=3, value=1.0)
        ws.cell(row=r, column=4, value=0.0)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_gross_weight_tabular_large_sheet_has_performance_and_row_stats():
    proc = GrossWeightProcessor()
    out = proc.process(_gw_bytes_many_tabular_rows(500), original_filename='scale.xlsx')
    assert out['success'] is True
    assert out['totalRows'] == 500
    assert out['errorRows'] == 0
    assert 'performance' in out
    assert 'rowStats' in out
    assert out['rowStats']['tabularParsedRows'] == 500
    assert out['rowStats']['tabularRawRowsScanned'] == 500
    assert out['performance']['rowsPerSecond'] > 0
    assert out['summary']['layoutMode'] == 'tabular'


def test_excel_max_rows_truncates_and_surfaces_flag(monkeypatch: pytest.MonkeyPatch) -> None:
    get_settings.cache_clear()
    monkeypatch.setenv('EXCEL_MAX_ROWS', '15')
    get_settings.cache_clear()
    proc = GrossWeightProcessor()
    out = proc.process(_gw_bytes_many_tabular_rows(40), original_filename='cap.xlsx')
    get_settings.cache_clear()
    monkeypatch.delenv('EXCEL_MAX_ROWS', raising=False)
    get_settings.cache_clear()

    assert out['rowStats']['scanCapTruncated'] is True
    assert out['totalRows'] < 40
    assert out['rowStats']['tabularRawRowsScanned'] == out['totalRows']


def test_pan_streaming_returns_performance(monkeypatch: pytest.MonkeyPatch) -> None:
    get_settings.cache_clear()
    monkeypatch.setenv('EXCEL_PAN_HEADER_PROBE_ROWS', '10')
    get_settings.cache_clear()
    wb = Workbook()
    ws = wb.active
    ws.append(['Total Value', 'PAN', 'PAN1', 'Add. proof', 'Add. Proof 2'])
    ws.append([1000, 'ABCDE1234F', '', 'ok', ''])
    buf = BytesIO()
    wb.save(buf)
    proc = PanProcessor()
    out = proc.process(buf.getvalue())
    get_settings.cache_clear()
    monkeypatch.delenv('EXCEL_PAN_HEADER_PROBE_ROWS', raising=False)
    get_settings.cache_clear()

    assert out['success'] is True
    assert 'performance' in out
    assert 'rowStats' in out
    assert out['rowStats'].get('engine') in ('openpyxl_read_only', 'pandas')
