from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app


client = TestClient(app)


def test_export_invalid_pan_rows_returns_excel_file() -> None:
    payload = {
        'records': [
            {
                'rowNumber': 77,
                'date': '03-04-2025',
                'voucherNo': 'JH/B/2526/1',
                'party': 'Kundan Crafts Private Limited',
                'totalValue': 18790000,
                'pan': '',
                'pan1': 'pending',
                'addProof': '',
                'addProof2': '',
                'issues': ['MISSING_PAN_ABOVE_2L', 'MISSING_ADDRESS_PROOF_ABOVE_50K'],
                'messages': [
                    'incorrect pan format',
                    'addressing missing',
                ],
                'Message': 'incorrect pan format; addressing missing',
            }
        ]
    }

    response = client.post('/api/process/pan/export-invalid', json=payload)

    assert response.status_code == 200
    assert response.headers['content-type'].startswith(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    assert 'attachment; filename=' in response.headers['content-disposition']

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook['Invalid PAN Rows']

    headers = [cell.value for cell in sheet[1]]
    assert headers == [
        'rowNumber',
        'date',
        'voucherNo',
        'party',
        'totalValue',
        'pan',
        'pan1',
        'addProof',
        'addProof2',
        'issues',
        'messages',
    ]
    assert sheet.cell(row=2, column=1).value == 77
    assert (
        sheet.cell(row=2, column=10).value
        == 'MISSING_PAN_ABOVE_2L, MISSING_ADDRESS_PROOF_ABOVE_50K'
    )
    assert sheet.cell(row=2, column=11).value == (
        'incorrect pan format, '
        'addressing missing'
    )


def test_export_invalid_pan_rows_with_empty_records_returns_400() -> None:
    response = client.post('/api/process/pan/export-invalid', json={'records': []})

    assert response.status_code == 400
    body = response.json()
    assert body['success'] is False
    assert body['detail'] == 'No invalid records found to export'
