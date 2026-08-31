"""Closing Stock product Rule Book mapping tests."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.engines.financials_engine.config.product_rule_book import (
    compute_rule_book_fingerprint,
    count_rule_book_products,
    load_closing_stock_product_rule_book,
    map_pivots_to_closing_stock_categories,
    map_product_names_to_categories,
)
from app.engines.financials_engine.engine.closing_stock_template import (
    CLOSING_STOCK_CATEGORIES,
    PURCHASES_AMT_LEAF_IDX,
    PURCHASES_QTY_LEAF_IDX,
    SALES_AMT_LEAF_IDX,
    SALES_QTY_LEAF_IDX,
    build_closing_stock_template_bytes,
    leaf_index_for_measure,
)


SAMPLE_RULE_BOOK = {
    'Diamond': {
        'Diamonds - Beads': ['Product A'],
        'Diamonds': ['Product B'],
    },
    'Emerald': ['Product C'],
    'Pearls': ['Product D'],
    'Rubie': ['Product E'],
    'Precious and Semi Precious': {
        'Precious Stones': ['Product F'],
        'Semi Precious': ['Product G'],
        'Synthetic Stones': ['Product H'],
    },
}


def _leaf_col(idx: int) -> int:
    return 2 + idx


def _product_row(layout: list[dict], label: str) -> dict:
    for row in layout:
        if row.get('kind') == 'product' and row.get('label') == label:
            return row
    raise AssertionError(f'Product row not found: {label}')


class TestClosingStockProductRuleBook:
    def test_semantic_column_indices(self):
        assert leaf_index_for_measure('purchasesQty') == 2
        assert leaf_index_for_measure('purchasesAmt') == 3
        assert leaf_index_for_measure('salesQty') == 21
        assert leaf_index_for_measure('salesAmt') == 22
        assert PURCHASES_QTY_LEAF_IDX == 2
        assert SALES_QTY_LEAF_IDX == 21
        assert SALES_AMT_LEAF_IDX == 22

    def test_maps_all_rule_book_products_regardless_of_pivot_input(self):
        mapped = map_product_names_to_categories(
            ['Product A', 'Unknown'],
            rule_book=SAMPLE_RULE_BOOK,
        )
        assert mapped['Diamond'] == ['Product A', 'Product B']
        assert mapped['Emerald'] == ['Product C']
        assert mapped['Pearls'] == ['Product D']
        assert mapped['Rubie'] == ['Product E']
        assert mapped['Precious and Semi Precious'] == [
            'Product F',
            'Product G',
            'Product H',
        ]

    def test_blank_products_always_kept_in_layout(self):
        """Products with no Sales/Purchases values must still appear with blank measures."""
        result = map_pivots_to_closing_stock_categories(
            sales_pivot=[
                {'product': 'Product A', 'sumOfQuantity': 2, 'sumOfGross': 100},
            ],
            purchases_pivot=[],
            rule_book=SAMPLE_RULE_BOOK,
        )
        assert result['productsDisplayed'] == 8
        blank_labels = []
        for layout in result['layoutByCategory'].values():
            for row in layout:
                if row.get('kind') != 'product':
                    continue
                if all(
                    row.get(k) is None
                    for k in ('salesQty', 'salesAmt', 'purchasesQty', 'purchasesAmt')
                ):
                    blank_labels.append(row['label'])
        assert 'Product B' in blank_labels
        assert 'Product C' in blank_labels
        assert 'Product D' in blank_labels
        assert 'Product E' in blank_labels
        assert 'Product F' in blank_labels
        assert 'Product G' in blank_labels
        assert 'Product H' in blank_labels
        assert 'Product A' not in blank_labels

    def test_empty_pivots_still_include_all_rule_book_products(self):
        result = map_pivots_to_closing_stock_categories(
            sales_pivot=[],
            purchases_pivot=[],
            rule_book=SAMPLE_RULE_BOOK,
        )
        assert result['productsDisplayed'] == 8
        assert result['productsWithSalesData'] == 0
        assert result['productsWithPurchaseData'] == 0
        assert result['unmappedProducts'] == []

        product_h = _product_row(
            result['layoutByCategory']['Precious and Semi Precious'],
            'Product H',
        )
        assert product_h['salesQty'] is None
        assert product_h['purchasesQty'] is None

    def test_attaches_sales_and_purchases_to_rule_book_products(self):
        result = map_pivots_to_closing_stock_categories(
            sales_pivot=[
                {'product': 'Product A', 'sumOfQuantity': 2, 'sumOfGross': 100},
                {'product': 'Product C', 'sumOfQuantity': 1, 'sumOfGross': 50},
                {'product': 'Product F', 'sumOfQuantity': 4, 'sumOfGross': 40},
                {'product': 'Orphan', 'sumOfQuantity': 9, 'sumOfGross': 9},
            ],
            purchases_pivot=[
                {'product': 'Product B', 'sumOfQuantity': 3, 'sumOfGross': 30},
                {'product': 'Product A', 'sumOfQuantity': 1, 'sumOfGross': 10},
                {'product': 'Product G', 'sumOfQuantity': 2, 'sumOfGross': 20},
            ],
            rule_book=SAMPLE_RULE_BOOK,
        )

        assert result['productsDisplayed'] == 8
        assert result['productsWithSalesData'] == 3
        assert result['productsWithPurchaseData'] == 3
        assert result['unmappedProducts'] == ['Orphan']

        product_a = _product_row(result['layoutByCategory']['Diamond'], 'Product A')
        assert product_a['salesQty'] == 2
        assert product_a['salesAmt'] == 100
        assert product_a['purchasesQty'] == 1
        assert product_a['purchasesAmt'] == 10

        recon = result['reconciliation']
        assert recon['mappedOutputMatch'] is True
        assert recon['salesQtyMatch'] is False
        assert recon['salesAmtMatch'] is False
        assert recon['purchasesQtyMatch'] is True
        assert recon['purchasesAmtMatch'] is True
        assert recon['outputSalesQty'] == 7
        assert recon['outputSalesAmt'] == 190
        assert recon['outputPurchasesQty'] == 6
        assert recon['outputPurchasesAmt'] == 60

        product_b = _product_row(result['layoutByCategory']['Diamond'], 'Product B')
        assert product_b['salesQty'] is None
        assert product_b['purchasesQty'] == 3

        product_d = _product_row(result['layoutByCategory']['Pearls'], 'Product D')
        assert product_d['salesQty'] is None
        assert product_d['purchasesQty'] is None

        product_h = _product_row(
            result['layoutByCategory']['Precious and Semi Precious'],
            'Product H',
        )
        assert product_h['salesQty'] is None

    def test_case_insensitive_whitespace_match(self):
        result = map_pivots_to_closing_stock_categories(
            sales_pivot=[
                {'product': ' jem 100 ', 'sumOfQuantity': 5, 'sumOfGross': 500},
            ],
            purchases_pivot=[],
            rule_book={
                'Diamond': {'Diamonds': []},
                'Emerald': ['JEM 100'],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        product = _product_row(result['layoutByCategory']['Emerald'], 'JEM 100')
        assert product['salesQty'] == 5
        assert product['salesAmt'] == 500

    def test_rename_mapping_displays_rule_book_name_and_keeps_values(self):
        """
        Pivot may use short/old/spacing variants; output label is always Rule Book name.
        """
        result = map_pivots_to_closing_stock_categories(
            sales_pivot=[
                {'product': 'JPS 1000', 'sumOfQuantity': 3, 'sumOfGross': 30},
                {'product': 'Flatpolki FP 1', 'sumOfQuantity': 1, 'sumOfGross': 10},
                {'product': 'Synthetic JSY 100', 'sumOfQuantity': 2, 'sumOfGross': 20},
                {'product': 'JSY 150', 'sumOfQuantity': 4, 'sumOfGross': 40},
            ],
            purchases_pivot=[
                {'product': 'Pearls JPS 100', 'sumOfQuantity': 5, 'sumOfGross': 50},
            ],
            rule_book={
                'Diamond': {
                    'Diamonds - Flat polki': ['Flat polki FP 1', 'Flat polki FP 2'],
                },
                'Emerald': [],
                'Pearls': ['Pearls JPS 100', 'Pearls JPS 1000'],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': ['Synthetic JSY 100', 'Synthetic JSY 150'],
                },
            },
        )

        pearls = result['productsByCategory']['Pearls']
        assert pearls == ['Pearls JPS 100', 'Pearls JPS 1000']
        assert 'JPS 1000' not in pearls

        pearl_1000 = _product_row(result['layoutByCategory']['Pearls'], 'Pearls JPS 1000')
        assert pearl_1000['salesQty'] == 3
        assert pearl_1000['salesAmt'] == 30

        pearl_100 = _product_row(result['layoutByCategory']['Pearls'], 'Pearls JPS 100')
        assert pearl_100['purchasesQty'] == 5

        flat = _product_row(result['layoutByCategory']['Diamond'], 'Flat polki FP 1')
        assert flat['salesQty'] == 1
        assert 'FP 1' not in [
            row['label']
            for row in result['layoutByCategory']['Diamond']
            if row.get('kind') == 'product'
        ]
        assert 'Flatpolki FP 1' not in [
            row['label']
            for row in result['layoutByCategory']['Diamond']
            if row.get('kind') == 'product'
        ]

        syn_100 = _product_row(
            result['layoutByCategory']['Precious and Semi Precious'],
            'Synthetic JSY 100',
        )
        assert syn_100['salesQty'] == 2
        syn_150 = _product_row(
            result['layoutByCategory']['Precious and Semi Precious'],
            'Synthetic JSY 150',
        )
        assert syn_150['salesQty'] == 4
        assert result['unmappedProducts'] == []

    def test_core_sku_does_not_confuse_jru_100_with_jru_1000(self):
        result = map_pivots_to_closing_stock_categories(
            sales_pivot=[
                {'product': 'JRU 100', 'sumOfQuantity': 1, 'sumOfGross': 10},
                {'product': 'JRU 1000', 'sumOfQuantity': 2, 'sumOfGross': 20},
            ],
            purchases_pivot=[],
            rule_book={
                'Diamond': {'Diamonds': []},
                'Emerald': [],
                'Pearls': [],
                'Rubie': ['Rubies JRU 100', 'Rubies JRU 1000'],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        r100 = _product_row(result['layoutByCategory']['Rubie'], 'Rubies JRU 100')
        r1000 = _product_row(result['layoutByCategory']['Rubie'], 'Rubies JRU 1000')
        assert r100['salesQty'] == 1
        assert r1000['salesQty'] == 2
        assert result['productsByCategory']['Rubie'] == ['Rubies JRU 100', 'Rubies JRU 1000']

    def test_unrelated_orphan_pivot_stays_unmapped(self):
        result = map_pivots_to_closing_stock_categories(
            sales_pivot=[
                {'product': 'Emeralds JEM 100', 'sumOfQuantity': 2, 'sumOfGross': 20},
                {'product': 'Gold Ornaments 22K', 'sumOfQuantity': 9, 'sumOfGross': 9},
            ],
            purchases_pivot=[],
            rule_book={
                'Diamond': {'Diamonds': []},
                'Emerald': ['Emeralds JEM 100'],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        product = _product_row(result['layoutByCategory']['Emerald'], 'Emeralds JEM 100')
        assert product['salesQty'] == 2
        assert result['unmappedProducts'] == ['Gold Ornaments 22K']

    def test_workbook_writes_sales_and_purchases_columns(self):
        mapped = map_pivots_to_closing_stock_categories(
            sales_pivot=[
                {'product': 'Product A', 'sumOfQuantity': 2, 'sumOfGross': 100},
                {'product': 'Product C', 'sumOfQuantity': 1, 'sumOfGross': 50},
            ],
            purchases_pivot=[
                {'product': 'Product A', 'sumOfQuantity': 1, 'sumOfGross': 10},
                {'product': 'Product B', 'sumOfQuantity': 3, 'sumOfGross': 30},
            ],
            rule_book=SAMPLE_RULE_BOOK,
        )
        raw = build_closing_stock_template_bytes(
            products_by_category=mapped['productsByCategory'],
            layout_by_category=mapped['layoutByCategory'],
        )
        wb = load_workbook(BytesIO(raw))
        diamond = wb['Diamond']

        product_a_row = 11
        assert diamond.cell(row=product_a_row, column=1).value == 'Product A'
        assert diamond.cell(row=product_a_row, column=_leaf_col(PURCHASES_QTY_LEAF_IDX)).value == 1
        assert diamond.cell(row=product_a_row, column=_leaf_col(PURCHASES_AMT_LEAF_IDX)).value == 10
        assert diamond.cell(row=product_a_row, column=_leaf_col(SALES_QTY_LEAF_IDX)).value == 2
        assert diamond.cell(row=product_a_row, column=_leaf_col(SALES_AMT_LEAF_IDX)).value == 100
        assert diamond.cell(row=product_a_row, column=_leaf_col(0)).value is None
        # Sales must not land in Issues Total columns (indices 19–20).
        assert diamond.cell(row=product_a_row, column=_leaf_col(19)).value is None
        assert diamond.cell(row=product_a_row, column=_leaf_col(20)).value is None
        assert diamond.cell(row=8, column=_leaf_col(SALES_QTY_LEAF_IDX)).value == 'Qty'
        assert diamond.cell(row=6, column=_leaf_col(SALES_QTY_LEAF_IDX)).value == 'Sales'

        product_b_row = 14
        assert diamond.cell(row=product_b_row, column=_leaf_col(PURCHASES_QTY_LEAF_IDX)).value == 3
        assert diamond.cell(row=product_b_row, column=_leaf_col(SALES_QTY_LEAF_IDX)).value is None

        subtotal_row = 12
        # TOTAL uses ROUND(SUM(unrounded)), written as a value — not SUM of rounded cells.
        assert diamond.cell(row=subtotal_row, column=_leaf_col(PURCHASES_QTY_LEAF_IDX)).value == 1
        assert diamond.cell(row=subtotal_row, column=_leaf_col(PURCHASES_AMT_LEAF_IDX)).value == 10
        assert diamond.cell(row=subtotal_row, column=_leaf_col(SALES_QTY_LEAF_IDX)).value == 2
        assert diamond.cell(row=subtotal_row, column=_leaf_col(SALES_AMT_LEAF_IDX)).value == 100

    def test_dual_pivot_names_sum_onto_one_rule_book_product(self):
        """Short and long pivot names for the same SKU must both claim and SUM."""
        result = map_pivots_to_closing_stock_categories(
            sales_pivot=[
                {'product': 'JPS 1000', 'sumOfQuantity': 3, 'sumOfGross': 30},
                {'product': 'Pearls JPS 1000', 'sumOfQuantity': 2, 'sumOfGross': 20},
            ],
            purchases_pivot=[],
            rule_book={
                'Diamond': {'Diamonds': []},
                'Emerald': [],
                'Pearls': ['Pearls JPS 1000'],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )
        pearl = _product_row(result['layoutByCategory']['Pearls'], 'Pearls JPS 1000')
        assert pearl['salesQty'] == 5
        assert pearl['salesAmt'] == 50
        recon = result['reconciliation']
        assert recon['mappedOutputMatch'] is True
        assert recon['pivotOutputMatch'] is True
        assert recon['salesQtyMatch'] is True

    def test_product_rounds_amount_total_rounds_unrounded_amount_sum(self):
        """
        Qty is never rounded.
        Product Amount = ROUND(each).
        TOTAL Amount = ROUND(SUM(unrounded)), not SUM(rounded product amounts).
        """
        sales_pivot = [
            {'product': 'Product X', 'sumOfQuantity': 100.49, 'sumOfGross': 100.49},
            {'product': 'Product Y', 'sumOfQuantity': 200.49, 'sumOfGross': 200.49},
        ]
        result = map_pivots_to_closing_stock_categories(
            sales_pivot=sales_pivot,
            purchases_pivot=[],
            rule_book={
                'Diamond': {'Diamonds': []},
                'Emerald': ['Product X', 'Product Y'],
                'Pearls': [],
                'Rubie': [],
                'Precious and Semi Precious': {
                    'Precious Stones': [],
                    'Semi Precious': [],
                    'Synthetic Stones': [],
                },
            },
        )

        assert sales_pivot[0]['sumOfQuantity'] == 100.49
        assert sales_pivot[1]['sumOfGross'] == 200.49

        x = _product_row(result['layoutByCategory']['Emerald'], 'Product X')
        y = _product_row(result['layoutByCategory']['Emerald'], 'Product Y')
        assert x['salesQty'] == 100.49
        assert y['salesQty'] == 200.49
        assert x['salesAmt'] == 100
        assert y['salesAmt'] == 200

        grand = next(
            row
            for row in result['layoutByCategory']['Emerald']
            if row.get('kind') == 'grand_total'
        )
        assert grand['salesQty'] == 300.98
        assert grand['salesAmt'] == 301
        assert (x['salesAmt'] + y['salesAmt']) == 300
        assert grand['salesAmt'] != (x['salesAmt'] + y['salesAmt'])

    def test_reconciliation_with_unmapped_pivot_reduces_output_totals(self):
        result = map_pivots_to_closing_stock_categories(
            sales_pivot=[
                {'product': 'Product A', 'sumOfQuantity': 2, 'sumOfGross': 100},
                {'product': 'Orphan', 'sumOfQuantity': 9, 'sumOfGross': 9},
            ],
            purchases_pivot=[],
            rule_book=SAMPLE_RULE_BOOK,
        )
        recon = result['reconciliation']
        assert recon['salesPivotQty'] == 11
        assert recon['outputSalesQty'] == 2
        assert recon['mappedOutputMatch'] is True
        assert recon['pivotOutputMatch'] is False
        assert result['unmappedProducts'] == ['Orphan']

    def test_production_rule_book_counts_and_display(self):
        book = load_closing_stock_product_rule_book()
        counts = count_rule_book_products(book)
        total = sum(counts.values())
        mapped = map_pivots_to_closing_stock_categories(rule_book=book)
        assert mapped['productsDisplayed'] == total
        assert mapped['ruleBookProductTotal'] == total
        assert mapped['ruleBookFingerprint'] == compute_rule_book_fingerprint(book)
        for category in CLOSING_STOCK_CATEGORIES:
            assert len(mapped['productsByCategory'][category]) == counts[category]

    def test_renamed_product_uses_new_rule_book_name_not_old(self):
        old_name = 'OLD PRODUCT NAME'
        new_name = 'NEW PRODUCT NAME'
        rule_v1 = {
            'Diamond': {'Diamonds': [old_name]},
            'Emerald': [],
            'Pearls': [],
            'Rubie': [],
            'Precious and Semi Precious': {
                'Precious Stones': [],
                'Semi Precious': [],
                'Synthetic Stones': [],
            },
        }
        rule_v2 = {
            **rule_v1,
            'Diamond': {'Diamonds': [new_name]},
        }
        fp_v1 = compute_rule_book_fingerprint(rule_v1)
        fp_v2 = compute_rule_book_fingerprint(rule_v2)
        assert fp_v1 != fp_v2

        result = map_pivots_to_closing_stock_categories(
            sales_pivot=[{'product': new_name, 'sumOfQuantity': 5, 'sumOfGross': 50}],
            purchases_pivot=[],
            rule_book=rule_v2,
        )
        products = result['productsByCategory']['Diamond']
        assert new_name in products
        assert old_name not in products
        layout_labels = [
            row['label']
            for row in result['layoutByCategory']['Diamond']
            if row.get('kind') == 'product'
        ]
        assert new_name in layout_labels
        assert old_name not in layout_labels
        product_row = _product_row(result['layoutByCategory']['Diamond'], new_name)
        assert product_row['salesQty'] == 5
        assert product_row['salesAmt'] == 50

    def test_disk_reload_picks_up_renamed_product(self, tmp_path):
        """Editing the JSON file must change output without code changes."""
        import json
        from pathlib import Path

        from app.engines.financials_engine.config import product_rule_book as prb

        rule_path = tmp_path / 'closing_stock_product_rule_book.json'
        book_v1 = {
            'Diamond': {'Diamonds': ['LEGACY SKU']},
            'Emerald': ['JEM KEEP'],
            'Pearls': [],
            'Rubie': [],
            'Precious and Semi Precious': {
                'Precious Stones': [],
                'Semi Precious': [],
                'Synthetic Stones': [],
            },
        }
        book_v2 = {
            **book_v1,
            'Diamond': {'Diamonds': ['RENAMED SKU']},
        }
        rule_path.write_text(json.dumps(book_v1), encoding='utf-8')

        original_path = prb._RULE_BOOK_PATH
        try:
            prb._RULE_BOOK_PATH = Path(rule_path)
            first = map_pivots_to_closing_stock_categories(
                sales_pivot=[{'product': 'LEGACY SKU', 'sumOfQuantity': 1, 'sumOfGross': 10}],
                purchases_pivot=[],
            )
            assert first['productsByCategory']['Diamond'] == ['LEGACY SKU']
            assert 'RENAMED SKU' not in first['productsByCategory']['Diamond']
            fp1 = first['ruleBookFingerprint']

            rule_path.write_text(json.dumps(book_v2), encoding='utf-8')
            second = map_pivots_to_closing_stock_categories(
                sales_pivot=[
                    {'product': 'renamed sku', 'sumOfQuantity': 2, 'sumOfGross': 20},
                    {'product': 'LEGACY SKU', 'sumOfQuantity': 9, 'sumOfGross': 9},
                ],
                purchases_pivot=[],
            )
            assert second['ruleBookFingerprint'] != fp1
            assert second['productsByCategory']['Diamond'] == ['RENAMED SKU']
            assert 'LEGACY SKU' not in second['productsByCategory']['Diamond']
            labels = [
                row['label']
                for row in second['layoutByCategory']['Diamond']
                if row.get('kind') == 'product'
            ]
            assert labels == ['RENAMED SKU']
            renamed = _product_row(second['layoutByCategory']['Diamond'], 'RENAMED SKU')
            assert renamed['salesQty'] == 2
            assert renamed['salesAmt'] == 20
            assert 'LEGACY SKU' in second['unmappedProducts']
        finally:
            prb._RULE_BOOK_PATH = original_path
