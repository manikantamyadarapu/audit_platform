from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

client = TestClient(app)


def test_export_invalid_gross_weight_rows_returns_excel_file() -> None:
    payload = {
        'records': [
            {
                'voucherNo': 'JH/2526/ 1',
                'manualGross': 11.12,
                'autoGross': 13.15,
                'difference': 2.03,
                'status': 'invalid',
                'issues': ['Manual Gross and Auto Gross mismatch', 'Difference must be exactly 0.00'],
                'voucherRow': 2,
                'dataRow': 2,
            }
        ]
    }

    response = client.post('/api/process/gross-weight/export-invalid', json=payload)

    assert response.status_code == 200
    assert response.headers['content-type'].startswith(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    assert 'attachment; filename=' in response.headers['content-disposition']

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook['Invalid gross weight rows']

    headers = [cell.value for cell in sheet[1]]
    assert headers == [
        'voucherNo',
        'manualGross',
        'autoGross',
        'difference',
        'status',
        'issues',
    ]
    assert sheet.cell(row=2, column=1).value == 'JH/2526/ 1'
    issues_cell = sheet.cell(row=2, column=6).value
    assert 'Manual Gross and Auto Gross mismatch' in issues_cell
    assert 'Difference must be exactly 0.00' in issues_cell


def test_export_invalid_gross_weight_rows_with_empty_records_returns_400() -> None:
    response = client.post('/api/process/gross-weight/export-invalid', json={'records': []})

    assert response.status_code == 400
    body = response.json()
    assert body['success'] is False
    assert body['detail'] == 'No invalid records found to export'
