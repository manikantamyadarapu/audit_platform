"""Unit tests for semi-structured gross weight processor."""

from io import BytesIO

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment

from app.processors.gross_weight_processor import GrossWeightProcessor
from app.validators.gross_weight_validator import (
    display_float_two_dp,
    to_decimal_two_dp,
    validate_triplet,
)


def _bytes_workbook(ws_builder) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws_builder(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_valid_voucher_basic():
    def fill(ws):
        ws['A4'] = 'Voucher No: JH/25-26 1'
        ws['B5'] = 5.2
        ws['C5'] = 5.2
        ws['D5'] = 0.0

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['success'] is True
    assert out['module'] == 'gross_weight'
    assert out.get('layoutEngine') == 'gross-weight-v2'
    assert out['fileType'] == 'gross_weight'
    assert out['totalRows'] == 1
    assert out['errorRows'] == 0
    assert out['summary']['total'] == 1
    assert out['summary']['valid'] == 1
    assert out['summary']['invalid'] == 0
    assert out['records'][0]['status'] == 'valid'
    assert out['records'][0]['voucherNo'] == 'JH/25-26 1'
    assert out['records'][0]['manualGross'] == 5.2
    assert out['records'][0]['autoGross'] == 5.2
    assert out['records'][0]['difference'] == 0.0
    assert out['records'][0]['issues'] == []


def test_half_up_quantize_and_display_matches_production_rules():
    assert float(to_decimal_two_dp(433.815)) == 433.82
    assert float(to_decimal_two_dp(433.825)) == 433.83
    assert float(to_decimal_two_dp(0.004)) == 0.0
    assert float(to_decimal_two_dp(0.005)) == 0.01
    assert float(to_decimal_two_dp(-0.005)) == -0.01
    assert display_float_two_dp(433.815) == 433.82
    assert display_float_two_dp(433.825) == 433.83


def test_validate_triplet_diff_only_violation_when_manual_equals_auto_but_diff_not_zero():
    issues, mm, dv, dov = validate_triplet(100.0, 100.0, 5.0)
    assert issues
    assert mm is False
    assert dv is True
    assert dov is True


def test_equal_weights_non_zero_difference_invalid_only_diff_message():
    """Equal manual/auto but non-zero difference cell: invalid on difference rule only (#3 removed)."""

    def fill(ws):
        ws['A1'] = 'SNo'
        ws['B1'] = 'Manual Gross wt.'
        ws['C1'] = 'Auto Gross Wt.'
        ws['D1'] = 'Difference in Gross wt.'
        ws['A2'] = 'Voucher No: JH/2526/ 4954'
        ws['B2'] = 140.55
        ws['C2'] = 140.55
        ws['D2'] = 127.0

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['errorRows'] == 1
    issues = out['records'][0]['issues']
    assert any('Difference must be exactly 0.00' in i for i in issues)
    assert not any('Difference does not match' in i for i in issues)
    assert not any('Manual Gross and Auto Gross mismatch' in i for i in issues)


def test_invalid_manual_auto_and_difference():
    def fill(ws):
        ws['A1'] = 'Voucher No: JH/25-26 77'
        ws['B2'] = 8.2
        ws['C2'] = 8.5
        ws['D2'] = 0.3

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['totalRows'] == 1
    assert out['errorRows'] == 1
    assert out['summary']['invalid'] == 1
    issues = out['records'][0]['issues']
    assert 'Manual Gross and Auto Gross mismatch' in issues
    assert any('Difference must be exactly 0.00' in i for i in issues)
    assert out['summary']['mismatchCount'] == 1
    assert out['summary']['differenceViolations'] == 1
    assert out['summary']['diffOnlyViolations'] == 0


def test_blank_spacer_row_between_voucher_and_values():
    def fill(ws):
        ws['A2'] = 'Voucher No: SPACER-1'
        ws['A3'] = None
        ws['B4'] = '1,000.50'
        ws['C4'] = '1000.50'
        ws['D4'] = 0

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['totalRows'] == 1
    assert out['errorRows'] == 0
    assert out['records'][0]['manualGross'] == 1000.5


def test_merged_voucher_label_row():
    def fill(ws):
        ws.merge_cells('A4:D4')
        top = ws['A4']
        top.value = 'Voucher No: MERGED-1'
        top.alignment = Alignment(horizontal='left')
        ws['B5'] = 2.5
        ws['C5'] = 2.5
        ws['D5'] = 0.0

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['totalRows'] == 1
    assert out['records'][0]['voucherNo'] == 'MERGED-1'
    assert out['records'][0]['status'] == 'valid'


def test_missing_data_row():
    def fill(ws):
        ws['A1'] = 'Voucher No: NO-DATA'
        ws['A2'] = 'noise only'

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['totalRows'] == 1
    assert out['errorRows'] == 1
    assert 'Could not locate values row after voucher label' in out['records'][0]['issues']


def test_multiple_vouchers():
    def fill(ws):
        ws['A1'] = 'Voucher No: A'
        ws['B2'] = 1
        ws['C2'] = 1
        ws['D2'] = 0
        ws['A3'] = 'Voucher No: B'
        ws['B4'] = 2
        ws['C4'] = 3
        ws['D4'] = 1

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['totalRows'] == 2
    assert out['summary']['valid'] == 1
    assert out['summary']['invalid'] == 1


def test_empty_workbook():
    def fill(ws):
        ws['A1'] = None

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['totalRows'] == 0
    assert out['records'] == []


def test_tabular_export_columns_like_user_sheet():
    """Headers: SNo, Manual Gross wt., Auto Gross Wt., Difference in Gross wt. — same row per voucher."""

    def fill(ws):
        ws['A1'] = 'SNo'
        ws['B1'] = 'Manual Gross wt.'
        ws['C1'] = 'Auto Gross Wt.'
        ws['D1'] = 'Difference in Gross wt.'
        ws['A2'] = 'Voucher No: JH/2526/ 10206'
        ws['B2'] = 11.12
        ws['C2'] = 13.15
        ws['D2'] = 2.03
        ws['A3'] = 'Voucher No: JH/2526/ 1084'
        ws['B3'] = 3.08
        ws['C3'] = 3.26
        ws['D3'] = 0.18

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out.get('layoutEngine') == 'gross-weight-v2'
    assert out['totalRows'] == 2
    assert out['errorRows'] == 2
    assert out['summary']['invalid'] == 2
    assert out['records'][0]['voucherNo'] == 'JH/2526/ 10206'
    assert out['records'][0]['manualGross'] == 11.12
    assert out['records'][0]['autoGross'] == 13.15
    assert out['records'][0]['difference'] == 2.03
    assert 'Manual Gross and Auto Gross mismatch' in out['records'][0]['issues']
    assert any('Difference must be exactly 0.00' in i for i in out['records'][0]['issues'])


def test_tabular_signed_zero_difference_valid():
    """Difference cell may be -0.00 from Excel; still valid when manual == auto."""

    def fill(ws):
        ws['A1'] = 'SNo'
        ws['B1'] = 'Manual Gross wt.'
        ws['C1'] = 'Auto Gross Wt.'
        ws['D1'] = 'Difference in Gross wt.'
        ws['A2'] = 'Voucher No: JH/2526/ 1'
        ws['B2'] = 5.2
        ws['C2'] = 5.2
        ws['D2'] = -0.0

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['records'][0]['status'] == 'valid'
    assert out['records'][0]['issues'] == []


def test_tabular_skips_repeated_header_row_mid_sheet():
    """Repeated column headers mid-sheet are omitted (not counted invalid), like PAN skips non-rows."""

    def fill(ws):
        ws['A1'] = 'SNo'
        ws['B1'] = 'Manual Gross wt.'
        ws['C1'] = 'Auto Gross Wt.'
        ws['D1'] = 'Difference in Gross wt.'
        ws['A2'] = 'Voucher No: JH/2526/ 1'
        ws['B2'] = 1.0
        ws['C2'] = 1.0
        ws['D2'] = 0.0
        ws['A3'] = 'SNo'
        ws['B3'] = 'Manual Gross wt.'
        ws['C3'] = 'Auto Gross Wt.'
        ws['D3'] = 'Difference in Gross wt.'
        ws['A4'] = 'Voucher No: JH/2526/ 2'
        ws['B4'] = 2.0
        ws['C4'] = 2.0
        ws['D4'] = 0.0

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['totalRows'] == 2
    assert {r['voucherNo'] for r in out['records']} == {'JH/2526/ 1', 'JH/2526/ 2'}


def test_tabular_skips_grand_total_guard_row():
    def fill(ws):
        ws['A1'] = 'SNo'
        ws['B1'] = 'Manual Gross wt.'
        ws['C1'] = 'Auto Gross Wt.'
        ws['D1'] = 'Difference in Gross wt.'
        ws['A2'] = 'Voucher No: JH/2526/ 1'
        ws['B2'] = 1.0
        ws['C2'] = 1.0
        ws['D2'] = 0.0
        ws['A3'] = 'Grand Total'
        ws['B3'] = 999.0
        ws['C3'] = 999.0
        ws['D3'] = 0.0

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['totalRows'] == 1
    assert out['records'][0]['voucherNo'] == 'JH/2526/ 1'


def test_tabular_all_valid_rows():
    def fill(ws):
        ws['A1'] = 'SNo'
        ws['B1'] = 'Manual Gross wt.'
        ws['C1'] = 'Auto Gross Wt.'
        ws['D1'] = 'Difference in Gross wt.'
        ws['A2'] = 'Voucher No: JH/2526/ 1'
        ws['B2'] = 5.2
        ws['C2'] = 5.2
        ws['D2'] = 0.0

    proc = GrossWeightProcessor()
    out = proc.process(_bytes_workbook(fill), original_filename='gw.xlsx')
    assert out['totalRows'] == 1
    assert out['errorRows'] == 0
    assert out['records'][0]['status'] == 'valid'
