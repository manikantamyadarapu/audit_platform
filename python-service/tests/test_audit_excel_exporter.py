"""Tests for reusable multi-sheet audit Excel exporter + Cash Ledger total report."""

from io import BytesIO

from openpyxl import load_workbook

from app.utils.audit_excel_exporter import (
    EMPTY_SHEET_MESSAGE,
    build_multi_sheet_audit_workbook,
    sanitize_sheet_name,
)
from app.utils.excel_exporter import (
    CASH_LEDGER_EXPORT_COLUMNS,
    CASH_LEDGER_EXPORT_HEADER_MAP,
    export_cash_ledger_total_error_report,
)


def test_sanitize_sheet_name_unique_and_safe() -> None:
    used: set[str] = set()
    first = sanitize_sheet_name('Cash Payments >= ₹10,000', used=used)
    second = sanitize_sheet_name('Cash Payments >= ₹10,000', used=used)
    assert first == 'Cash Payments >= ₹10,000'
    assert second != first
    assert '/' not in sanitize_sheet_name('A/B')


def test_build_multi_sheet_creates_empty_placeholder() -> None:
    payload = build_multi_sheet_audit_workbook(
        {
            'Negative Cash': [],
            'Cash Payments >= ₹10,000': [
                {
                    'rowNumber': 10,
                    'date': '01-04-2025',
                    'voucher_no': 'V1',
                    'branch': 'HQ',
                    'contra_account': 'Rent',
                    'debit': '',
                    'credit': 15000,
                    'balance': '1 Dr',
                    'Message': 'Cash Payments>=Rs. 10,000/-',
                }
            ],
            'Cash Receipts >= ₹2,00,000': [],
        },
        columns=CASH_LEDGER_EXPORT_COLUMNS,
        header_map=CASH_LEDGER_EXPORT_HEADER_MAP,
    )
    wb = load_workbook(BytesIO(payload))
    assert wb.sheetnames == [
        'Negative Cash',
        'Cash Payments >= ₹10,000',
        'Cash Receipts >= ₹2,00,000',
    ]
    assert wb['Negative Cash']['A1'].value == EMPTY_SHEET_MESSAGE
    assert wb['Cash Receipts >= ₹2,00,000']['A1'].value == EMPTY_SHEET_MESSAGE
    pay = wb['Cash Payments >= ₹10,000']
    assert pay['A1'].value == 'Row No'
    assert pay['I1'].value == 'Message'
    assert pay.max_row == 2


def test_export_cash_ledger_total_error_report_groups_by_rule() -> None:
    records = [
        {
            'rowNumber': 1,
            'date': '01-04-2025',
            'voucher_no': 'V1',
            'branch': 'HQ',
            'contra_account': 'X',
            'debit': '',
            'credit': '',
            'balance': '-10 Dr',
            'Message': 'Negative Cash',
            'issues': ['NEGATIVE_CASH_BALANCE'],
        },
        {
            'rowNumber': 2,
            'date': '02-04-2025',
            'voucher_no': 'V2',
            'branch': 'HQ',
            'contra_account': 'Rent',
            'debit': '',
            'credit': 20000,
            'balance': '1 Dr',
            'Message': 'Cash Payments>=Rs. 10,000/-',
            'issues': ['CASH_PAYMENT_GT_10000'],
        },
    ]
    payload = export_cash_ledger_total_error_report(records)
    wb = load_workbook(BytesIO(payload))
    assert len(wb.sheetnames) == 3
    assert wb['Negative Cash'].max_row == 2
    assert wb['Cash Payments >= ₹10,000'].max_row == 2
    assert wb['Cash Receipts >= ₹2,00,000']['A1'].value == EMPTY_SHEET_MESSAGE
    headers = [cell.value for cell in wb['Negative Cash'][1]]
    assert 'Issue Code' not in headers
    assert 'Severity' not in headers
    assert 'issues' not in headers


def test_export_cash_ledger_total_error_report_allows_all_empty() -> None:
    payload = export_cash_ledger_total_error_report([])
    wb = load_workbook(BytesIO(payload))
    assert len(wb.sheetnames) == 3
    for name in wb.sheetnames:
        assert wb[name]['A1'].value == EMPTY_SHEET_MESSAGE
