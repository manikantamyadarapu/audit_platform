from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

client = TestClient(app)


def test_export_gross_weight_invalid_rows() -> None:
    payload = {
        'records': [
            {
                'rowNumber': 5,
                'manualGrossWeight': 10.5,
                'autoGrossWeight': 11.0,
                'difference': 0.5,
                'issues': ['GROSS_WEIGHT_MISMATCH'],
                'messages': ['Manual gross weight does not match auto gross weight.'],
            }
        ],
        'summary': {'mismatchCount': 1},
        'processingStatistics': {'totalRows': 10},
        'executionTiming': {'validationMs': 12.5},
    }
    response = client.post('/api/process/gross-weight/export-invalid', json=payload)
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook['Invalid Gross Weight Rows']
    headers = [cell.value for cell in sheet[1]]
    assert headers[:6] == [
        'Row Number',
        'Voucher No',
        'Manual Gross Weight',
        'Auto Gross Weight',
        'Difference',
        'Message',
    ]
    assert 'SNo' not in headers
    assert 'Message' in headers
    assert 'Summary' in workbook.sheetnames
    assert 'Issue Breakdown' in workbook.sheetnames
    assert 'Issue Grouping' in workbook.sheetnames
    assert 'Processing Statistics' in workbook.sheetnames
    assert 'Execution Timing' in workbook.sheetnames


def test_export_sales_invalid_rows() -> None:
    payload = {
        'records': [
            {
                'rowNumber': 2,
                'voucherNo': 'V1',
                'partyName': 'Acme Ltd',
                'salesAccount': 'Sales A',
                'product': 'Prod',
                'unitRate': 155,
                'issues': ['INVALID_PRODUCT_MAPPING'],
                'messages': ['Product does not belong to the uploaded sales account in the master sales verification sheet.'],
            }
        ],
        'summary': {'invalidProductMappings': 1},
        'processingStatistics': {'totalRows': 7},
        'executionTiming': {'validationMs': 15.0},
    }
    response = client.post('/api/process/sales/export-invalid', json=payload)
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook['Invalid Sales Rows']
    headers = [cell.value for cell in sheet[1]]
    assert 'rowNumber' in headers
    assert 'salesAccount' in headers
    assert 'product' in headers
    assert 'unitRate' in headers
    assert 'messages' in headers
    summary_sheet = workbook['Summary']
    summary_metrics = [cell.value for cell in summary_sheet['A'] if cell.value]
    assert 'reportTitle' in summary_metrics
    assert 'summary.invalidProductMappings' in summary_metrics
