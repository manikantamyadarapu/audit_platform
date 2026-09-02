"""Blank Closing Stock working-paper Excel template (no calculations yet)."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.indian_number_format import apply_indian_number_format

# Category sheets in workbook order — reused later for per-category calculations.
CLOSING_STOCK_CATEGORIES: tuple[str, ...] = (
    'Diamond',
    'Emerald',
    'Pearls',
    'Rubie',
    'Precious and Semi Precious',
)

# Sheets whose Rule Book section is subcategory → products.
CATEGORIES_WITH_SUBCATEGORIES: frozenset[str] = frozenset(
    {
        'Diamond',
        'Precious and Semi Precious',
    }
)


def closing_stock_report_title(category: str) -> str:
    """DETAILS OF JEWELS CLOSING STOCK - {CATEGORY}."""
    return f'DETAILS OF JEWELS CLOSING STOCK - {str(category).strip().upper()}'


def subcategory_total_label(category: str, subcategory: str) -> str:
    """Label for a subcategory TOTAL row."""
    if category == 'Precious and Semi Precious':
        return f'TOTAL - {str(subcategory).strip().upper()}'
    return 'TOTAL'


# Semantic measure keys — resolved from LEAF_COLUMNS after definition (never hardcode indices).
CLOSING_STOCK_MEASURE_PATHS: dict[str, tuple[str | None, str | None, str]] = {
    'openingQty': ('Opening Stock', None, 'Qty'),
    'openingAmt': ('Opening Stock', None, 'Amt.'),
    'purchasesQty': ('Purchases', None, 'Qty'),
    'purchasesAmt': ('Purchases', None, 'Amt.'),
    'receiptsIstQty': ('Receipts', 'Internal Stock Transfer', 'Qty'),
    'receiptsIstAmt': ('Receipts', 'Internal Stock Transfer', 'Amt.'),
    'receiptsJubileeQty': ('Receipts', 'Jubilee Hills', 'Qty'),
    'receiptsJubileeAmt': ('Receipts', 'Jubilee Hills', 'Amt.'),
    'receiptsKokapetQty': ('Receipts', 'Kokapet', 'Qty'),
    'receiptsKokapetAmt': ('Receipts', 'Kokapet', 'Amt.'),
    'receiptsTotalQty': ('Receipts', 'Total', 'Qty'),
    'receiptsTotalAmt': ('Receipts', 'Total', 'Amt.'),
    'issuesIstQty': ('Issues', 'Internal Stock Transfer', 'Qty'),
    'issuesIstAmt': ('Issues', 'Internal Stock Transfer', 'Amt.'),
    'issuesBanjaraQty': ('Issues', 'Banjara Hills', 'Qty'),
    'issuesBanjaraAmt': ('Issues', 'Banjara Hills', 'Amt.'),
    'issuesKokapetQty': ('Issues', 'Kokapet', 'Qty'),
    'issuesKokapetAmt': ('Issues', 'Kokapet', 'Amt.'),
    'issuesTotalQty': ('Issues', 'Total', 'Qty'),
    'issuesTotalAmt': ('Issues', 'Total', 'Amt.'),
    'salesQty': ('Sales', None, 'Qty'),
    'salesAmt': ('Sales', None, 'Amt.'),
}

LEAF_COLUMNS: tuple[tuple[tuple[str | None, str | None, str], str], ...] = (
    (('Opening Stock', None, 'Qty'), '1'),
    (('Opening Stock', None, 'Amt.'), '2'),
    (('Purchases', None, 'Qty'), '3'),
    (('Purchases', None, 'Amt.'), '4'),
    (('Receipts', 'Internal Stock Transfer', 'Qty'), '5'),
    (('Receipts', 'Internal Stock Transfer', 'Amt.'), '6'),
    (('Receipts', 'Jubilee Hills', 'Qty'), '7'),
    (('Receipts', 'Jubilee Hills', 'Amt.'), '8'),
    (('Receipts', 'Kokapet', 'Qty'), '9'),
    (('Receipts', 'Kokapet', 'Amt.'), '10'),
    (('Receipts', 'Total', 'Qty'), '11'),
    (('Receipts', 'Total', 'Amt.'), '12'),
    (('Average Rate', None, 'Amt.'), '13'),
    (('Issues', 'Internal Stock Transfer', 'Qty'), '14'),
    (('Issues', 'Internal Stock Transfer', 'Amt.'), '15'),
    (('Issues', 'Banjara Hills', 'Qty'), '16'),
    (('Issues', 'Banjara Hills', 'Amt.'), '17'),
    (('Issues', 'Kokapet', 'Qty'), '18'),
    (('Issues', 'Kokapet', 'Amt.'), '19'),
    (('Issues', 'Total', 'Qty'), '20'),
    (('Issues', 'Total', 'Amt.'), '21'),
    (('Sales', None, 'Qty'), '22'),
    (('Sales', None, 'Amt.'), '23'),
    (('Closing Stock', None, 'Qty'), '24'),
    (('Closing Stock', None, 'Amt.'), '25'),
    (('Gross Profit', None, 'Amt.'), '26'),
    (('GP AY 2025-26', None, 'Qty'), '27'),
    (('GP AY 2025-26', None, 'Amt.'), '28'),
    (('Deviation', None, 'Qty'), '29'),
    (('Deviation', None, 'Amt.'), '30'),
    (('Deviation', None, '%'), '31'),
)


def _leaf_index_for_path(path: tuple[str | None, str | None, str]) -> int:
    for idx, (leaf_path, _num) in enumerate(LEAF_COLUMNS):
        if leaf_path == path:
            return idx
    raise KeyError(f'Closing Stock column not found for path {path!r}')


def leaf_index_for_measure(measure_key: str) -> int:
    """Return 0-based LEAF_COLUMNS index for a semantic measure key."""
    return _leaf_index_for_path(CLOSING_STOCK_MEASURE_PATHS[measure_key])


PURCHASES_QTY_LEAF_IDX = leaf_index_for_measure('purchasesQty')
PURCHASES_AMT_LEAF_IDX = leaf_index_for_measure('purchasesAmt')
SALES_QTY_LEAF_IDX = leaf_index_for_measure('salesQty')
SALES_AMT_LEAF_IDX = leaf_index_for_measure('salesAmt')

# Back-compat aliases (tests / callers may still reference these names).
REPORT_TITLE = closing_stock_report_title('Diamond')
SHEET_NAME = CLOSING_STOCK_CATEGORIES[0]

THIN = Border(
    left=Side(style='thin', color='94A3B8'),
    right=Side(style='thin', color='94A3B8'),
    top=Side(style='thin', color='94A3B8'),
    bottom=Side(style='thin', color='94A3B8'),
)

HEADER_FILL = PatternFill('solid', fgColor='0F766E')
SUBHEADER_FILL = PatternFill('solid', fgColor='115E59')
LEAF_FILL = PatternFill('solid', fgColor='134E4A')
NUMBER_FILL = PatternFill('solid', fgColor='ECFDF5')
SUBCATEGORY_FILL = PatternFill('solid', fgColor='CCFBF1')
TOTAL_FILL = PatternFill('solid', fgColor='FEF3C7')
GRAND_TOTAL_FILL = PatternFill('solid', fgColor='FDE68A')
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='0F172A')
META_FONT = Font(name='Calibri', size=11, color='334155')
HEADER_FONT = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
LEAF_FONT = Font(name='Calibri', size=8, bold=True, color='FFFFFF')
NUMBER_FONT = Font(name='Calibri', size=8, bold=True, color='0F766E')
BODY_FONT = Font(name='Calibri', size=10, color='0F172A')
SUBCATEGORY_FONT = Font(name='Calibri', size=10, bold=True, color='0F766E')
TOTAL_FONT = Font(name='Calibri', size=10, bold=True, color='92400E')
GRAND_TOTAL_FONT = Font(name='Calibri', size=10, bold=True, color='78350F')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _merge_same_run(ws, row: int, start_col: int, end_col: int, value: str, fill, font) -> None:
    if end_col < start_col:
        return
    cell = ws.cell(row=row, column=start_col, value=value)
    cell.fill = fill
    cell.font = font
    cell.alignment = CENTER
    cell.border = THIN
    if end_col > start_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.border = THIN
        c.alignment = CENTER


def _write_grouped_row(
    ws,
    row: int,
    values: Sequence[str | None],
    fill,
    font,
    *,
    start_col: int = 2,
) -> None:
    """Merge consecutive equal non-empty labels; blank labels stay as individual cells."""
    n = len(values)
    i = 0
    while i < n:
        label = values[i] or ''
        if not label:
            cell = ws.cell(row=row, column=start_col + i, value='')
            cell.fill = fill
            cell.border = THIN
            cell.alignment = CENTER
            i += 1
            continue
        j = i
        while j + 1 < n and (values[j + 1] or '') == label:
            j += 1
        _merge_same_run(ws, row, start_col + i, start_col + j, label, fill, font)
        i = j + 1


def _is_summable_leaf(path: tuple[str | None, str | None, str]) -> bool:
    """Numeric columns that TOTAL/GRAND TOTAL may SUM. Skip Average Rate and %."""
    level1, _level2, leaf = path
    if leaf == '%':
        return False
    if level1 == 'Average Rate':
        return False
    return True


def _leaf_excel_column(leaf_idx: int) -> int:
    return 2 + leaf_idx


def _write_product_measures(
    ws: Worksheet,
    row: int,
    entry: Mapping[str, Any],
) -> None:
    """Write wired Closing Stock measures for one product row; unwired columns stay blank."""
    for measure_key in CLOSING_STOCK_MEASURE_PATHS:
        value = entry.get(measure_key)
        if value is None:
            continue
        leaf_idx = leaf_index_for_measure(measure_key)
        cell = ws.cell(row=row, column=_leaf_excel_column(leaf_idx), value=value)
        apply_indian_number_format(cell)
        cell.alignment = CENTER


# Back-compat alias for older call sites / tests.
_write_product_sales_purchases = _write_product_measures


def _write_sum_formulas(
    ws: Worksheet,
    row: int,
    *,
    sum_rows: Sequence[int],
    last_col: int,
    fill,
    entry: Mapping[str, Any] | None = None,
) -> None:
    """
    Write TOTAL / GRAND TOTAL cells.

    Sales/Purchases Qty/Amt use precomputed ROUND(SUM(unrounded)) from ``entry``
    when present — never Excel SUM of already-rounded product cells.
    Other summable columns keep SUM formulas for future measures.
    """
    measure_leaf_indexes = {
        leaf_index_for_measure(key): key for key in CLOSING_STOCK_MEASURE_PATHS
    }

    if not sum_rows and not entry:
        for col in range(2, last_col + 1):
            cell = ws.cell(row=row, column=col, value=None)
            cell.border = THIN
            cell.fill = fill
            cell.alignment = CENTER
        return

    first = min(sum_rows) if sum_rows else None
    last = max(sum_rows) if sum_rows else None
    for idx, (path, _num) in enumerate(LEAF_COLUMNS):
        col = 2 + idx
        cell = ws.cell(row=row, column=col)
        cell.border = THIN
        cell.fill = fill
        cell.alignment = CENTER

        measure_key = measure_leaf_indexes.get(idx)
        if measure_key and entry is not None and entry.get(measure_key) is not None:
            cell.value = entry.get(measure_key)
            cell.font = TOTAL_FONT
            apply_indian_number_format(cell)
            continue

        if _is_summable_leaf(path) and first is not None and last is not None and first <= last:
            letter = get_column_letter(col)
            cell.value = f'=SUM({letter}{first}:{letter}{last})'
            cell.font = TOTAL_FONT
            apply_indian_number_format(cell)
        else:
            cell.value = None


def _write_title_and_headers(
    ws: Worksheet,
    *,
    category: str,
    company_name: str,
    address: str,
    financial_year: str,
) -> tuple[int, int]:
    """Write title block + multi-level headers. Returns (data_start_row, last_col)."""
    leaf_count = len(LEAF_COLUMNS)
    last_col = 1 + leaf_count

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws['A1'] = company_name or 'Company Name'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = CENTER

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws['A2'] = address or 'Address'
    ws['A2'].font = META_FONT
    ws['A2'].alignment = CENTER

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    ws['A3'] = f'Financial Year: {financial_year}'
    ws['A3'].font = META_FONT
    ws['A3'].alignment = CENTER

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
    ws['A4'] = closing_stock_report_title(category)
    ws['A4'].font = Font(name='Calibri', size=12, bold=True, color='0F766E')
    ws['A4'].alignment = CENTER

    header_row_l1 = 6
    header_row_l2 = 7
    header_row_leaf = 8
    header_row_num = 9

    ws.merge_cells(start_row=header_row_l1, start_column=1, end_row=header_row_num, end_column=1)
    particulars = ws.cell(row=header_row_l1, column=1, value='Particulars / Product')
    particulars.fill = HEADER_FILL
    particulars.font = HEADER_FONT
    particulars.alignment = CENTER
    particulars.border = THIN
    for r in range(header_row_l1, header_row_num + 1):
        cell = ws.cell(row=r, column=1)
        cell.fill = HEADER_FILL
        cell.border = THIN

    level1 = [path[0] for path, _ in LEAF_COLUMNS]
    level2 = [path[1] or '' for path, _ in LEAF_COLUMNS]
    leaves = [path[2] for path, _ in LEAF_COLUMNS]
    numbers = [num for _, num in LEAF_COLUMNS]

    _write_grouped_row(ws, header_row_l1, level1, HEADER_FILL, HEADER_FONT)
    _write_grouped_row(ws, header_row_l2, level2, SUBHEADER_FILL, HEADER_FONT)

    for idx, leaf in enumerate(leaves):
        cell = ws.cell(row=header_row_leaf, column=2 + idx, value=leaf)
        cell.fill = LEAF_FILL
        cell.font = LEAF_FONT
        cell.alignment = CENTER
        cell.border = THIN

    for idx, num in enumerate(numbers):
        cell = ws.cell(row=header_row_num, column=2 + idx, value=num)
        cell.fill = NUMBER_FILL
        cell.font = NUMBER_FONT
        cell.alignment = CENTER
        cell.border = THIN

    ws.column_dimensions['A'].width = 36
    for col in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 9

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[4].height = 20
    for r in (header_row_l1, header_row_l2, header_row_leaf, header_row_num):
        ws.row_dimensions[r].height = 28

    ws.freeze_panes = 'B10'
    ws.print_title_rows = '1:9'
    return header_row_num + 1, last_col


def _normalize_layout_rows(layout_rows: Sequence[dict[str, Any]] | None, products: Sequence[str]) -> list[dict[str, Any]]:
    if layout_rows:
        return [dict(row) for row in layout_rows]
    product_list = [str(p).strip() for p in products if str(p).strip()]
    rows: list[dict[str, Any]] = [{'kind': 'product', 'label': p} for p in product_list]
    if rows:
        rows.append({'kind': 'grand_total', 'label': 'GRAND TOTAL'})
    elif not rows:
        rows = [{'kind': 'product', 'label': ''}]
    return rows


def _write_closing_stock_sheet(
    ws: Worksheet,
    *,
    category: str,
    products: Sequence[str],
    layout_rows: Sequence[dict[str, Any]] | None = None,
    company_name: str,
    address: str,
    financial_year: str,
) -> None:
    """Apply the shared Closing Stock layout to one worksheet."""
    data_start, last_col = _write_title_and_headers(
        ws,
        category=category,
        company_name=company_name,
        address=address,
        financial_year=financial_year,
    )

    rows = _normalize_layout_rows(layout_rows, products)
    product_rows_all: list[int] = []
    current_subcategory_product_rows: list[int] = []
    row = data_start

    for entry in rows:
        kind = str(entry.get('kind') or 'product')
        label = str(entry.get('label') or '')

        if kind == 'subcategory':
            current_subcategory_product_rows = []
            name_cell = ws.cell(row=row, column=1, value=label)
            name_cell.font = SUBCATEGORY_FONT
            name_cell.alignment = LEFT
            name_cell.border = THIN
            name_cell.fill = SUBCATEGORY_FILL
            for col in range(2, last_col + 1):
                cell = ws.cell(row=row, column=col, value=None)
                cell.border = THIN
                cell.fill = SUBCATEGORY_FILL
            row += 1
            continue

        if kind == 'product':
            # Always keep the product row (blank cells when measures are None).
            name_cell = ws.cell(row=row, column=1, value=label)
            name_cell.font = BODY_FONT
            name_cell.alignment = LEFT
            name_cell.border = THIN
            for col in range(2, last_col + 1):
                cell = ws.cell(row=row, column=col, value=None)
                cell.border = THIN
                cell.alignment = CENTER
            _write_product_measures(ws, row, entry)
            if label.strip():
                product_rows_all.append(row)
                current_subcategory_product_rows.append(row)
            row += 1
            continue

        if kind == 'subcategory_total':
            name_cell = ws.cell(row=row, column=1, value=label or 'TOTAL')
            name_cell.font = TOTAL_FONT
            name_cell.alignment = LEFT
            name_cell.border = THIN
            name_cell.fill = TOTAL_FILL
            _write_sum_formulas(
                ws,
                row,
                sum_rows=current_subcategory_product_rows,
                last_col=last_col,
                fill=TOTAL_FILL,
                entry=entry,
            )
            current_subcategory_product_rows = []
            row += 1
            continue

        if kind == 'grand_total':
            name_cell = ws.cell(row=row, column=1, value=label or 'GRAND TOTAL')
            name_cell.font = GRAND_TOTAL_FONT
            name_cell.alignment = LEFT
            name_cell.border = THIN
            name_cell.fill = GRAND_TOTAL_FILL
            _write_sum_formulas(
                ws,
                row,
                sum_rows=product_rows_all,
                last_col=last_col,
                fill=GRAND_TOTAL_FILL,
                entry=entry,
            )
            row += 1
            continue

        # Unknown kind → treat as product label
        name_cell = ws.cell(row=row, column=1, value=label)
        name_cell.font = BODY_FONT
        name_cell.alignment = LEFT
        name_cell.border = THIN
        for col in range(2, last_col + 1):
            cell = ws.cell(row=row, column=col, value=None)
            cell.border = THIN
            cell.alignment = CENTER
        row += 1


def build_closing_stock_template_bytes(
    *,
    products: Sequence[str] | None = None,
    products_by_category: dict[str, Sequence[str]] | None = None,
    layout_by_category: dict[str, Sequence[dict[str, Any]]] | None = None,
    company_name: str = '',
    address: str = '',
    financial_year: str = 'AY 2025-26',
    categories: Sequence[str] | None = None,
) -> bytes:
    """
    Build a blank Closing Stock workbook with one sheet per jewel category.

    Prefer ``layout_by_category`` (subcategory / TOTAL / GRAND TOTAL structure from the Rule Book).
    ``products_by_category`` is used when layout is not provided.
    """
    sheet_categories = tuple(categories) if categories else CLOSING_STOCK_CATEGORIES
    if not sheet_categories:
        raise ValueError('At least one Closing Stock category is required')

    fallback_products = [str(p).strip() for p in (products or []) if str(p).strip()]
    by_category = products_by_category or {}
    layouts = layout_by_category or {}
    wb = Workbook()

    for index, category in enumerate(sheet_categories):
        # Excel sheet titles max 31 chars; our longest name fits.
        sheet_title = category[:31]
        if index == 0:
            ws = wb.active
            ws.title = sheet_title
        else:
            ws = wb.create_sheet(title=sheet_title)

        layout = list(layouts.get(category) or [])
        if category in by_category:
            sheet_products = [
                str(p).strip() for p in (by_category.get(category) or []) if str(p).strip()
            ]
        elif by_category or layouts:
            sheet_products = []
        else:
            sheet_products = fallback_products

        _write_closing_stock_sheet(
            ws,
            category=category,
            products=sheet_products,
            layout_rows=layout or None,
            company_name=company_name,
            address=address,
            financial_year=financial_year,
        )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_pivots_workbook_bytes(
    *,
    sales_pivot: Sequence[dict[str, Any]],
    purchases_pivot: Sequence[dict[str, Any]],
) -> bytes:
    """Excel workbook with Sales Pivot and Purchases Pivot sheets."""
    from openpyxl.styles import Font as OxFont

    wb = Workbook()
    header = ('Product', 'Sum of Quantity', 'Sum of Gross')

    def _write_sheet(title: str, rows: Sequence[dict[str, Any]], *, first: bool) -> None:
        ws = wb.active if first else wb.create_sheet(title)
        if first:
            ws.title = title
        for col, label in enumerate(header, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = OxFont(bold=True, color='FFFFFF')
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN
        for r_idx, row in enumerate(rows, start=2):
            ws.cell(row=r_idx, column=1, value=row.get('product')).border = THIN
            qty = ws.cell(row=r_idx, column=2, value=row.get('sumOfQuantity'))
            qty.border = THIN
            apply_indian_number_format(qty)
            gross = ws.cell(row=r_idx, column=3, value=row.get('sumOfGross'))
            gross.border = THIN
            apply_indian_number_format(gross)
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.freeze_panes = 'A2'

    _write_sheet('Sales Pivot', sales_pivot, first=True)
    _write_sheet('Purchases Pivot', purchases_pivot, first=False)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
