"""Extract Closing Balance from a dedicated product sheet or category tab."""

from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from app.engines.financials_engine.engine.opening_stock import (
    norm_opening_product_name,
    product_sheet_lookup_keys,
)
from app.engines.financials_engine.parsers.workbook_loader import parse_numeric_value
from app.utils.header_cleaner import normalize_header
from app.utils.logger import get_logger
from app.utils.sheet_validation_error import SheetValidationError

HEADER_SCAN_LIMIT = 40

QTY_PRODUCT_ALIASES = frozenset(
    {'product', 'particulars_product', 'particulars', 'item', 'item_name'}
)
QTY_OPENING_ALIASES = frozenset(
    {
        'opening_balance',
        'opening_qty',
        'opening_quantity',
        'opening_stock_qty',
        'op_bal',
        'opening',
    }
)

PARTICULARS_ALIASES = frozenset(
    {
        'particulars',
        'particulars_product',
        'product',
        'narration',
        'description',
        'item',
    }
)
CLOSING_STOCK_HEADER_ALIASES = frozenset(
    {
        'closing_stock',
        'closing_balance',
        'closing',
        'cls_stock',
        'closing_stk',
    }
)
CLOSING_BALANCE_LABEL_ALIASES = frozenset(
    {
        'closing_balance',
        'closing_bal',
        'cls_balance',
        'closing_stock',
        'closing_stock_balance',
        'closing',
    }
)
AMOUNT_HEADER_ALIASES = frozenset(
    {
        'amount',
        'amt',
        'value',
        'closing_stock_amount',
        'closing_amount',
        'closing_balance_amount',
        'closing_bal_amount',
        'closing_value',
        'closing_amt',
        'cls_amt',
        'cls_amount',
    }
)
SKIP_ROW_LABELS = frozenset(
    {
        'total',
        'grand_total',
        'sub_total',
        'subtotal',
        'particulars',
    }
)
NON_STOCK_SHEET_HINTS = frozenset(
    {
        'balance sheet',
        'cash flow',
        'profit',
        'itr',
        'deferred tax',
        'depreciation',
        't shape',
        'trading',
        'abstract',
        'salary',
        'fixed assets',
        'adv frm',
        's crs',
        'bs ann',
        'pl ann',
        'sheet2',
        'bs',
        'pl',
        'cashflow',
        'dep co',
        'dep it',
        'def tax',
    }
)
# Category tabs that contain many products as rows (Eximp layout).
CLOSING_STOCK_CATEGORY_TABS = frozenset(
    {'dia', 'eme', 'prls', 'rubi', 'prec', 'diamond', 'emerald', 'pearls', 'rubie'}
)


def _display_product_name(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ''
    return str(value).replace('\n', ' ').replace('\r', ' ').strip()


def _labels_from_row(row: pd.Series | list[Any]) -> dict[int, str]:
    cells = row.tolist() if hasattr(row, 'tolist') else list(row)
    mapping: dict[int, str] = {}
    for idx, cell in enumerate(cells):
        label = normalize_header(cell)
        if label:
            mapping[idx] = label
    return mapping


def _find_header(
    raw: pd.DataFrame,
    *,
    product_aliases: frozenset[str],
    required_aliases: dict[str, frozenset[str]],
    source_label: str,
) -> tuple[int, dict[str, int]]:
    scan = min(HEADER_SCAN_LIMIT, len(raw.index))
    best: tuple[int, dict[str, int], int] | None = None

    for idx in range(scan):
        labels = _labels_from_row(raw.iloc[idx])
        if not labels:
            continue
        col_map: dict[str, int] = {}
        for col_idx, label in labels.items():
            if label in product_aliases and 'product' not in col_map:
                col_map['product'] = col_idx
            for key, aliases in required_aliases.items():
                if label in aliases and key not in col_map:
                    col_map[key] = col_idx
        score = len(col_map)
        needed = 1 + len(required_aliases)
        if score >= needed and 'product' in col_map:
            return idx, col_map
        if best is None or score > best[2]:
            best = (idx, col_map, score)

    missing = []
    if best is None or 'product' not in best[1]:
        missing.append('Product / Particulars')
    for key in required_aliases:
        if best is None or key not in best[1]:
            missing.append(key.replace('_', ' ').title())
    raise SheetValidationError(
        f'{source_label}: could not find required headers ({", ".join(missing)}). '
        f'Scan the first {HEADER_SCAN_LIMIT} rows for column names.',
        code='MISSING_COLUMNS',
        source=source_label,
        missingColumns=missing,
    )


def load_opening_quantity_workbook(
    file_bytes: bytes,
    file_name: str,
) -> list[dict[str, Any]]:
    """Current Year Opening Quantity File — Product + Opening Balance required."""
    try:
        raw = pd.read_excel(BytesIO(file_bytes), header=None, dtype=object)
    except Exception as exc:
        raise SheetValidationError(
            f'Opening Quantity file unreadable ({file_name}): {exc}',
            code='UNREADABLE_FILE',
            fileName=file_name,
            source='Opening Quantity',
        ) from exc

    if raw.empty:
        raise SheetValidationError(
            f'Opening Quantity file is empty ({file_name})',
            code='EMPTY_FILE',
            fileName=file_name,
            source='Opening Quantity',
        )

    header_idx, col_map = _find_header(
        raw,
        product_aliases=QTY_PRODUCT_ALIASES,
        required_aliases={'opening_balance': QTY_OPENING_ALIASES},
        source_label='Opening Quantity file',
    )

    optional_aliases = {
        'sku': frozenset({'sku', 'item_code', 'code'}),
        'receipts': frozenset({'receipts', 'receipt'}),
        'issues': frozenset({'issues', 'issue'}),
        'closing_balance': frozenset(
            {'closing_balance', 'closing_qty', 'closing_quantity', 'closing'}
        ),
    }
    labels = _labels_from_row(raw.iloc[header_idx])
    for key, aliases in optional_aliases.items():
        if key in col_map:
            continue
        for col_idx, label in labels.items():
            if label in aliases:
                col_map[key] = col_idx
                break

    rows: list[dict[str, Any]] = []
    for ridx in range(header_idx + 1, len(raw.index)):
        series = raw.iloc[ridx]
        product = _display_product_name(series.iloc[col_map['product']])
        if not product:
            continue
        entry: dict[str, Any] = {
            'product': product,
            'openingBalance': parse_numeric_value(series.iloc[col_map['opening_balance']]),
        }
        if 'sku' in col_map:
            entry['sku'] = _display_product_name(series.iloc[col_map['sku']])
        if 'receipts' in col_map:
            entry['receipts'] = parse_numeric_value(series.iloc[col_map['receipts']])
        if 'issues' in col_map:
            entry['issues'] = parse_numeric_value(series.iloc[col_map['issues']])
        if 'closing_balance' in col_map:
            entry['closingBalance'] = parse_numeric_value(series.iloc[col_map['closing_balance']])
        rows.append(entry)
    return rows


def _is_non_stock_sheet(sheet_name: str) -> bool:
    low = str(sheet_name or '').strip().casefold()
    if not low:
        return True
    for hint in NON_STOCK_SHEET_HINTS:
        if hint == low or hint in low:
            return True
    return False


def _is_category_closing_tab(sheet_name: str) -> bool:
    return str(sheet_name or '').strip().casefold() in CLOSING_STOCK_CATEGORY_TABS


def _is_skip_product_label(product: str) -> bool:
    key = normalize_header(product)
    if not key:
        return True
    if key in SKIP_ROW_LABELS:
        return True
    if key.startswith('total') or key.startswith('grand_total'):
        return True
    if key.isdigit():
        return True
    junk_bits = (
        'jewellers',
        'hyderabad',
        'basheerbagh',
        'cin_',
        'details_of_jewels',
        'ay_',
    )
    if any(bit in key for bit in junk_bits):
        return True
    return False


def _is_closing_balance_label(text: Any) -> bool:
    label = normalize_header(text)
    if not label:
        return False
    if label in CLOSING_BALANCE_LABEL_ALIASES:
        return True
    return label.startswith('closing_balance') or label.startswith('closing_stock')


def _detect_closing_stock_columns(raw: pd.DataFrame) -> tuple[int, int, int] | None:
    scan = min(HEADER_SCAN_LIMIT, len(raw.index))
    for idx in range(scan):
        labels = _labels_from_row(raw.iloc[idx])
        if not labels:
            continue
        particulars_col: int | None = None
        closing_col: int | None = None
        for col_idx, label in labels.items():
            if particulars_col is None and label in PARTICULARS_ALIASES:
                particulars_col = col_idx
            if closing_col is None and label in CLOSING_STOCK_HEADER_ALIASES:
                closing_col = col_idx
        if particulars_col is None or closing_col is None:
            continue

        amount_col = closing_col + 1
        if idx + 1 < len(raw.index):
            sub = list(raw.iloc[idx + 1].tolist())
            for offset in (1, 0, 2):
                cidx = closing_col + offset
                if cidx >= len(sub):
                    continue
                sub_label = normalize_header(sub[cidx])
                if 'amt' in sub_label or sub_label.endswith('amount'):
                    amount_col = cidx
                    break
        return idx, particulars_col, amount_col
    return None


def _extract_closing_balance_from_product_sheet(
    raw: pd.DataFrame,
    *,
    sheet_name: str,
) -> dict[str, Any]:
    """
    Dedicated product sheet: locate Closing Balance row and read Qty + Amount.
    """
    if raw is None or raw.empty:
        return {'closingStockAmount': None, 'closingStockQty': None, 'found': False}

    detected = _detect_closing_stock_columns(raw)
    header_idx: int | None = None
    amount_col: int | None = None
    qty_col: int | None = None
    particulars_col: int | None = None

    if detected is not None:
        header_idx, particulars_col, amount_col = detected
        qty_col = amount_col - 1 if amount_col > particulars_col else None
    else:
        scan = min(HEADER_SCAN_LIMIT, len(raw.index))
        for idx in range(scan):
            labels = _labels_from_row(raw.iloc[idx])
            for col_idx, label in labels.items():
                if label in AMOUNT_HEADER_ALIASES and amount_col is None:
                    amount_col = col_idx
                if label in PARTICULARS_ALIASES and particulars_col is None:
                    particulars_col = col_idx
            if amount_col is not None:
                header_idx = idx
                break

    scan_start = (header_idx + 1) if header_idx is not None else 0
    for ridx in range(scan_start, len(raw.index)):
        cells = list(raw.iloc[ridx].tolist())
        label_col: int | None = None
        for cidx, cell in enumerate(cells):
            if not _is_closing_balance_label(cell):
                continue
            if particulars_col is not None and cidx != particulars_col:
                continue
            label_col = cidx
            break
        if label_col is None:
            continue

        amount: float | None = None
        qty: float | None = None
        if amount_col is not None and amount_col < len(cells):
            amount = parse_numeric_value(cells[amount_col])
        if qty_col is not None and qty_col < len(cells):
            qty = parse_numeric_value(cells[qty_col])

        if amount is None:
            for cidx in range(len(cells) - 1, -1, -1):
                if cidx == label_col or cidx == qty_col:
                    continue
                if particulars_col is not None and cidx == particulars_col:
                    continue
                value = parse_numeric_value(cells[cidx])
                if value is not None:
                    amount = value
                    break

        if amount is None and qty is None:
            continue

        return {
            'product': sheet_name,
            'sheetName': sheet_name,
            'closingStockAmount': amount,
            'closingStockQty': qty,
            'found': amount is not None,
            'excelRow': ridx + 1,
            'reason': None if amount is not None else 'closing_balance_amount_missing',
        }

    return {'closingStockAmount': None, 'closingStockQty': None, 'found': False}


def _extract_products_from_category_sheet(
    raw: pd.DataFrame,
    *,
    sheet_name: str,
) -> dict[str, dict[str, Any]]:
    """Category tab (Dia/Eme/…): one row per product with Closing stock Qty/Amt."""
    detected = _detect_closing_stock_columns(raw)
    if detected is None:
        return {}

    header_idx, particulars_col, amount_col = detected
    qty_col = amount_col - 1 if amount_col > particulars_col else None

    products: dict[str, dict[str, Any]] = {}
    current_subcategory: str | None = None

    for ridx in range(header_idx + 1, len(raw.index)):
        cells = list(raw.iloc[ridx].tolist())
        if particulars_col >= len(cells):
            continue
        product = _display_product_name(cells[particulars_col])
        if not product or _is_skip_product_label(product):
            continue

        amount = parse_numeric_value(cells[amount_col]) if amount_col < len(cells) else None
        qty = None
        if qty_col is not None and 0 <= qty_col < len(cells):
            qty = parse_numeric_value(cells[qty_col])

        if amount is None and qty is None:
            current_subcategory = norm_opening_product_name(product) or None
            continue

        key = norm_opening_product_name(product)
        if not key:
            continue

        products[key] = {
            'product': product,
            'sheetName': sheet_name,
            'subcategory': current_subcategory,
            'closingStockAmount': amount,
            'closingStockQty': qty,
            'found': amount is not None,
            'excelRow': ridx + 1,
            'reason': None if amount is not None else 'closing_balance_amount_missing',
        }

    return products


def _register_product_keys(index: dict[str, dict[str, Any]], entry: dict[str, Any]) -> None:
    product = str(entry.get('product') or entry.get('sheetName') or '')
    for key in product_sheet_lookup_keys(product):
        if key not in index:
            index[key] = entry
    sheet_name = str(entry.get('sheetName') or '')
    if sheet_name:
        for key in product_sheet_lookup_keys(sheet_name):
            if key not in index:
                index[key] = entry


def load_previous_year_product_index(
    file_bytes: bytes,
    file_name: str,
    *,
    log: Any | None = None,
) -> dict[str, dict[str, Any]]:
    """
    Index previous-year Closing Balance by product.

    1. Dedicated product sheets — sheet name matches product; read Closing Balance row.
    2. Category tabs (Dia/Eme/…) — product rows with Closing stock Qty/Amt.
    """
    logger = log or get_logger()
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise SheetValidationError(
            f'Previous Year Closing Stock file unreadable ({file_name}): {exc}',
            code='UNREADABLE_FILE',
            fileName=file_name,
            source='Previous Year Closing Stock',
        ) from exc

    sheet_names = list(workbook.sheetnames)
    if not sheet_names:
        workbook.close()
        raise SheetValidationError(
            f'Previous Year Closing Stock file has no sheets ({file_name})',
            code='EMPTY_FILE',
            fileName=file_name,
            source='Previous Year Closing Stock',
        )

    index: dict[str, dict[str, Any]] = {}
    product_sheet_hits: list[str] = []
    category_sheet_hits: list[str] = []

    try:
        for sheet_name in sheet_names:
            display_name = str(sheet_name or '').strip()
            if _is_non_stock_sheet(display_name):
                continue

            try:
                ws = workbook[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
            except Exception as exc:
                logger.warning(
                    'Previous Year sheet unreadable: sheet={} error={}',
                    display_name,
                    exc,
                )
                continue

            if not rows:
                continue

            raw = pd.DataFrame(rows)

            if _is_category_closing_tab(display_name):
                extracted = _extract_products_from_category_sheet(raw, sheet_name=display_name)
                if extracted:
                    category_sheet_hits.append(display_name)
                    for entry in extracted.values():
                        _register_product_keys(index, entry)
                continue

            product_entry = _extract_closing_balance_from_product_sheet(
                raw,
                sheet_name=display_name,
            )
            if product_entry.get('found'):
                product_sheet_hits.append(display_name)
                _register_product_keys(index, product_entry)
                continue

            # Unknown layout — try category-style parse as last resort.
            extracted = _extract_products_from_category_sheet(raw, sheet_name=display_name)
            if extracted:
                category_sheet_hits.append(display_name)
                for entry in extracted.values():
                    _register_product_keys(index, entry)
    finally:
        workbook.close()

    unique_products = {
        str(v.get('product') or '') for v in index.values() if v.get('product')
    }
    logger.info(
        'Previous Year indexed: file={} product_sheets={} category_sheets={} products={}',
        file_name,
        product_sheet_hits,
        category_sheet_hits,
        len(unique_products),
    )
    if not index:
        logger.warning(
            'Previous Year Closing Stock contained no Closing Balance amounts ({})',
            file_name,
        )
    return index


def _build_subcategory_indexes(
    product_index: dict[str, dict[str, Any]],
) -> tuple[dict[tuple[str, str | None], list[dict[str, Any]]], dict[str, list[dict[str, Any]]]]:
    """Group unique previous-year product rows by sheet tab and subcategory section."""
    subcategory_products: dict[tuple[str, str | None], list[dict[str, Any]]] = {}
    sheet_products: dict[str, list[dict[str, Any]]] = {}
    seen_subcat: dict[tuple[str, str | None], set[str]] = {}
    seen_sheet: dict[str, set[str]] = {}

    for entry in product_index.values():
        product = str(entry.get('product') or '').strip()
        product_key = norm_opening_product_name(product)
        if not product_key:
            continue

        sheet_name = str(entry.get('sheetName') or '').strip()
        if sheet_name:
            sheet_seen = seen_sheet.setdefault(sheet_name, set())
            if product_key not in sheet_seen:
                sheet_seen.add(product_key)
                sheet_products.setdefault(sheet_name, []).append(dict(entry))

        subcategory = entry.get('subcategory')
        if sheet_name and subcategory:
            sub_key = (sheet_name, str(subcategory))
            sub_seen = seen_subcat.setdefault(sub_key, set())
            if product_key not in sub_seen:
                sub_seen.add(product_key)
                subcategory_products.setdefault(sub_key, []).append(dict(entry))

    return subcategory_products, sheet_products


def load_previous_year_opening_stock(
    file_bytes: bytes,
    file_name: str,
    *,
    log: Any | None = None,
) -> dict[str, Any]:
    """Product index plus subcategory/sheet groupings for fallback mapping."""
    product_index = load_previous_year_product_index(file_bytes, file_name, log=log)
    subcategory_products, sheet_products = _build_subcategory_indexes(product_index)
    return {
        'productIndex': product_index,
        'subcategoryProducts': subcategory_products,
        'sheetProducts': sheet_products,
    }


def load_previous_year_product_sheets(
    file_bytes: bytes,
    file_name: str,
    *,
    log: Any | None = None,
) -> dict[str, dict[str, Any]]:
    return load_previous_year_product_index(file_bytes, file_name, log=log)


def load_previous_year_closing_workbook(
    file_bytes: bytes,
    file_name: str,
) -> list[dict[str, Any]]:
    index = load_previous_year_product_index(file_bytes, file_name)
    seen: set[str] = set()
    rows: list[dict[str, Any]] = []
    for entry in index.values():
        product = str(entry.get('product') or '')
        key = norm_opening_product_name(product)
        if not key or key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                'product': entry.get('product'),
                'sheetName': entry.get('sheetName'),
                'closingStockQty': entry.get('closingStockQty'),
                'closingStockAmount': entry.get('closingStockAmount'),
                'found': entry.get('found'),
            }
        )
    return rows
