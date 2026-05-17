from __future__ import annotations

from collections.abc import Callable, Iterator, Sequence
from contextlib import contextmanager
from dataclasses import dataclass
from io import BytesIO
from time import perf_counter
from typing import Any

import duckdb
import polars as pl
import pyarrow as pa
from openpyxl import load_workbook

from app.utils.header_cleaner import normalize_header
from app.utils.logger import get_logger

_SUBTOTAL_PATTERN = r'(^\s*sub\s*total\b|\bgrand\s*total\b|^\s*total\s*$)'


@dataclass(slots=True)
class LoadedValidationSheet:
    dataframe: pl.DataFrame
    header_row_index: int
    header_detection_ms: float
    load_ms: float


class VectorizedValidationEngine:
    def __init__(self, processor_name: str) -> None:
        self.processor_name = processor_name
        self._log = get_logger()

    def load_sheet(
        self,
        file_bytes: bytes,
        *,
        row_matches: Callable[[set[str]], bool] | None = None,
        scan_limit: int = 60,
    ) -> LoadedValidationSheet:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            worksheet = workbook.active

            header_scan_start = perf_counter()
            header_row_index: int | None = None
            headers: list[str] = []
            positions: list[int] = []
            columns: dict[str, list[Any]] = {}
            source_excel_row_numbers: list[int] = []

            for physical_row, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                row_tuple = tuple(row)
                if header_row_index is None:
                    if row_matches is not None and row_matches(
                        self._header_labels_from_values(row_tuple)
                    ):
                        header_row_index = physical_row - 1
                        headers, positions = self._normalize_headers(row_tuple)
                        columns = {header: [] for header in headers}
                    continue
                if not headers:
                    continue
                self._append_row(columns, row_tuple, headers, positions)
                source_excel_row_numbers.append(physical_row)

            if header_row_index is None:
                header_row_index = 0

            header_detection_ms = (perf_counter() - header_scan_start) * 1000

            load_start = perf_counter()
            if headers and source_excel_row_numbers:
                excel_rows = pl.Series('source_excel_row_number', source_excel_row_numbers)
                dataframe = pl.DataFrame(columns, strict=False).with_columns(
                    excel_rows.alias('source_excel_row_number'),
                    excel_rows.alias('__excel_row_number__'),
                )
            elif headers:
                dataframe = pl.DataFrame(columns, strict=False).with_columns(
                    pl.lit(None).cast(pl.Int64).alias('source_excel_row_number'),
                    pl.lit(None).cast(pl.Int64).alias('__excel_row_number__'),
                )
            else:
                dataframe = pl.DataFrame(
                    schema={
                        'source_excel_row_number': pl.Int64,
                        '__excel_row_number__': pl.Int64,
                    }
                )

            load_ms = (perf_counter() - load_start) * 1000
            return LoadedValidationSheet(
                dataframe=dataframe,
                header_row_index=header_row_index,
                header_detection_ms=header_detection_ms,
                load_ms=load_ms,
            )
        finally:
            workbook.close()

    @contextmanager
    def duckdb_connection(
        self,
        dataframe: pl.DataFrame,
        *,
        table_name: str = 'source_rows',
    ) -> Iterator[duckdb.DuckDBPyConnection]:
        connection = duckdb.connect(database=':memory:')
        arrow_table: pa.Table = dataframe.to_arrow()
        connection.register(table_name, arrow_table)
        try:
            yield connection
        finally:
            try:
                connection.unregister(table_name)
            except duckdb.Error:
                pass
            connection.close()

    def fetch_frame(self, connection: duckdb.DuckDBPyConnection, sql: str) -> pl.DataFrame:
        return pl.from_arrow(connection.execute(sql).to_arrow_table())

    def log_benchmark(
        self,
        *,
        row_count: int,
        header_row_index: int,
        header_detection_ms: float,
        load_ms: float,
        validation_ms: float,
        extraction_ms: float,
        total_ms: float,
    ) -> None:
        self._log.info(
            (
                f'[{self.processor_name}] vectorized validation benchmark '
                f'rows={row_count} header_row={header_row_index + 1} '
                f'header_detection_ms={header_detection_ms:.2f} '
                f'load_ms={load_ms:.2f} validation_ms={validation_ms:.2f} '
                f'extraction_ms={extraction_ms:.2f} total_ms={total_ms:.2f}'
            )
        )

    @staticmethod
    def user_columns(dataframe: pl.DataFrame) -> list[str]:
        return [
            column
            for column in dataframe.columns
            if not column.startswith('__') and column != 'source_excel_row_number'
        ]

    @staticmethod
    def quote_identifier(column: str) -> str:
        return f'"{column.replace(chr(34), chr(34) * 2)}"'

    @staticmethod
    def _quote_literal(value: str) -> str:
        return "'" + value.replace("'", "''") + "'"

    @classmethod
    def blankable_text_sql(cls, column: str, *, empty_tokens: Sequence[str]) -> str:
        ident = cls.quote_identifier(column)
        trimmed = f'TRIM(CAST({ident} AS VARCHAR))'
        empty_tokens_sql = ', '.join(cls._quote_literal(token) for token in sorted(empty_tokens))
        return (
            'CASE '
            f'WHEN {ident} IS NULL THEN NULL '
            f"WHEN {trimmed} = '' THEN NULL "
            f'WHEN LOWER({trimmed}) IN ({empty_tokens_sql}) THEN NULL '
            f'ELSE {trimmed} END'
        )

    @classmethod
    def header_normalized_sql(cls, column: str) -> str:
        ident = cls.quote_identifier(column)
        return (
            "TRIM(REGEXP_REPLACE(LOWER(TRIM(COALESCE(CAST("
            f"{ident} AS VARCHAR), ''))), '[^a-z0-9]+', '_', 'g'), '_')"
        )

    @classmethod
    def subtotal_sql(cls, columns: Sequence[str]) -> str:
        relevant = [column for column in columns if column in {'voucher_no', 'party', 'narration', 'description'}]
        if not relevant:
            return 'FALSE'
        checks = [
            (
                "COALESCE("
                f"REGEXP_MATCHES(LOWER(TRIM(CAST(COALESCE({cls.quote_identifier(column)}, '') AS VARCHAR))), "
                f"'{_SUBTOTAL_PATTERN}'), FALSE)"
            )
            for column in relevant
        ]
        return '(' + ' OR '.join(checks) + ')'

    @classmethod
    def repeated_header_sql(cls, columns: Sequence[str]) -> str:
        checks: list[str] = []
        if 'total_value' in columns:
            checks.append(f"{cls.header_normalized_sql('total_value')} = 'total_value'")
        if 'pan' in columns and 'pan1' in columns:
            pan = cls.quote_identifier('pan')
            pan1 = cls.quote_identifier('pan1')
            checks.append(
                f"(UPPER(TRIM(CAST(COALESCE({pan}, '') AS VARCHAR))) = 'PAN' "
                f"AND UPPER(REPLACE(TRIM(CAST(COALESCE({pan1}, '') AS VARCHAR)), ' ', '')) = 'PAN1')"
            )
        for column, accepted in (
            ('manual_gross_weight', {'manual_gross_weight', 'manual_gross_wt'}),
            ('manual_gross_wt', {'manual_gross_weight', 'manual_gross_wt'}),
            ('auto_gross_weight', {'auto_gross_weight', 'auto_gross_wt'}),
            ('auto_gross_wt', {'auto_gross_weight', 'auto_gross_wt'}),
            ('difference', {'difference'}),
            ('voucher_no', {'voucher_no'}),
        ):
            if column not in columns:
                continue
            accepted_sql = ', '.join(cls._quote_literal(value) for value in sorted(accepted))
            checks.append(f"{cls.header_normalized_sql(column)} IN ({accepted_sql})")
        if not checks:
            return 'FALSE'
        return '(' + ' OR '.join(checks) + ')'

    @classmethod
    def blank_row_sql(cls, columns: Sequence[str], *, empty_tokens: Sequence[str]) -> str:
        if not columns:
            return 'FALSE'
        checks = [f"{cls.blankable_text_sql(column, empty_tokens=empty_tokens)} IS NULL" for column in columns]
        return '(' + ' AND '.join(checks) + ')'

    @classmethod
    def missing_voucher_sql(cls, columns: Sequence[str], *, empty_tokens: Sequence[str]) -> str:
        if 'voucher_no' not in columns:
            return 'FALSE'
        return f"({cls.blankable_text_sql('voucher_no', empty_tokens=empty_tokens)} IS NULL)"

    @classmethod
    def shared_skip_sql(
        cls,
        columns: Sequence[str],
        *,
        empty_tokens: Sequence[str],
        check_missing_voucher: bool = True,
    ) -> str:
        predicates = [
            cls.blank_row_sql(columns, empty_tokens=empty_tokens),
            cls.repeated_header_sql(columns),
            cls.subtotal_sql(columns),
        ]
        if check_missing_voucher:
            predicates.append(cls.missing_voucher_sql(columns, empty_tokens=empty_tokens))
        return '(' + ' OR '.join(f'COALESCE({predicate}, FALSE)' for predicate in predicates) + ')'

    @classmethod
    def amount_sql(cls, column: str, *, empty_tokens: Sequence[str]) -> str:
        text_expr = cls.blankable_text_sql(column, empty_tokens=empty_tokens)
        cleaned = f"REGEXP_REPLACE({text_expr}, ',', '', 'g')"
        extracted = f"REGEXP_EXTRACT({cleaned}, '(-?[0-9]+(?:\\.[0-9]+)?)', 1)"
        return (
            'CASE '
            f'WHEN {text_expr} IS NULL THEN NULL '
            f"WHEN {extracted} = '' THEN NULL "
            f'ELSE TRY_CAST({extracted} AS DOUBLE) END'
        )

    @classmethod
    def decimal_sql(cls, column: str, *, empty_tokens: Sequence[str], scale: int = 6) -> str:
        text_expr = cls.blankable_text_sql(column, empty_tokens=empty_tokens)
        cleaned = f"REPLACE({text_expr}, ',', '')"
        return (
            'CASE '
            f'WHEN {text_expr} IS NULL THEN NULL '
            f'ELSE TRY_CAST({cleaned} AS DECIMAL(18, {scale})) END'
        )

    @staticmethod
    def _preview_rows(row_iter: Iterator[tuple[Any, ...]], limit: int) -> list[tuple[Any, ...]]:
        preview: list[tuple[Any, ...]] = []
        for idx, row in enumerate(row_iter):
            preview.append(tuple(row))
            if idx + 1 >= max(1, limit):
                break
        return preview

    @staticmethod
    def _header_labels_from_values(row: Sequence[Any]) -> set[str]:
        labels: set[str] = set()
        for value in row:
            label = normalize_header(value)
            if label:
                labels.add(label)
        return labels

    @staticmethod
    def _normalize_headers(row: Sequence[Any]) -> tuple[list[str], list[int]]:
        headers: list[str] = []
        positions: list[int] = []
        seen: dict[str, int] = {}
        for idx, raw_header in enumerate(row):
            label = normalize_header(raw_header)
            if not label:
                continue
            seen[label] = seen.get(label, 0) + 1
            unique_label = label if seen[label] == 1 else f'{label}_{seen[label]}'
            headers.append(unique_label)
            positions.append(idx)
        return headers, positions

    @staticmethod
    def _append_row(
        columns: dict[str, list[Any]],
        row: Sequence[Any],
        headers: Sequence[str],
        positions: Sequence[int],
    ) -> None:
        for header, position in zip(headers, positions, strict=True):
            value = row[position] if position < len(row) else None
            columns[header].append(value)
