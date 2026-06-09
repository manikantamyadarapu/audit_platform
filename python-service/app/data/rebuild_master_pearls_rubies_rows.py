"""Maintenance: replace Jewels Pearls / Rubies product rows in master_sales_rules.xlsx.

Run from python-service: python -m app.data.rebuild_master_pearls_rubies_rows
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

_WORKBOOK = Path(__file__).resolve().parent / 'master_sales_rules.xlsx'

from app.sales_engine.config.loader import load_gemstone_product_catalog

_catalog = load_gemstone_product_catalog().get('accounts') or {}
_pe = _catalog.get('JEWELS SALES ACCOUNT - PEARLS') or {}
_ru = _catalog.get('JEWELS SALES ACCOUNT - RUBIES') or {}
_JPS_NUMBERS: list[int] = list(_pe.get('pearls_jps') or [])
_JRU_NUMBERS: list[int] = list(_ru.get('rubies_jru') or [])
_RUBY_TAIL: tuple[str, ...] = tuple(_ru.get('tail_products') or ())


def _pad_row(values: tuple[object, ...], width: int) -> tuple[object, ...]:
    lst = list(values)
    while len(lst) < width:
        lst.append(None)
    return tuple(lst[:width])


def _next_account_row_index(rows: list[tuple[object, ...]], start_idx: int) -> int | None:
    for j in range(start_idx + 1, len(rows)):
        r = rows[j]
        if r[0] is not None and str(r[0]).strip():
            return j
    return None


def main() -> None:
    wb = load_workbook(_WORKBOOK)
    ws = wb.active
    max_col = ws.max_column or 4
    rows: list[tuple[object, ...]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col):
        rows.append(_pad_row(tuple(c.value for c in row), max_col))

    pearl_start: int | None = None
    for i, r in enumerate(rows):
        cell0 = r[0]
        if cell0 is None:
            continue
        s = str(cell0).strip()
        if 'jewels' in s.lower() and 'pearl' in s.lower():
            pearl_start = i
            break
    if pearl_start is None:
        raise RuntimeError('Could not find Jewels sales account - Pearls row')

    pearl_end = _next_account_row_index(rows, pearl_start)
    if pearl_end is None:
        raise RuntimeError('Could not find row after Pearls block')

    ruby_start = pearl_end
    ruby_end = _next_account_row_index(rows, ruby_start)
    if ruby_end is None:
        ruby_end = len(rows)

    pearl_children = [_pad_row((None, f'Pearls JPS {n}', None, None), max_col) for n in sorted(_JPS_NUMBERS)]
    ruby_children = [_pad_row((None, f'Rubies JRU {n}', None, None), max_col) for n in sorted(_JRU_NUMBERS)] + [
        _pad_row((None, name, None, None), max_col) for name in _RUBY_TAIL
    ]

    rebuilt = (
        rows[: pearl_start + 1]
        + pearl_children
        + rows[ruby_start : ruby_start + 1]
        + ruby_children
        + rows[ruby_end:]
    )

    ws.delete_rows(1, ws.max_row)
    for r_idx, rvals in enumerate(rebuilt, start=1):
        for c_idx, val in enumerate(rvals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(_WORKBOOK)


if __name__ == '__main__':
    main()
