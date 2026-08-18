"""One-off maintenance: expand Jewels Emeralds rows in master_sales_rules.xlsx.

Run from repo root: python -m app.data.rebuild_master_emerald_rows
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

_WORKBOOK = Path(__file__).resolve().parent / 'master_sales_rules.xlsx'

from app.engines.sales_engine.config.loader import load_gemstone_product_catalog

_em = (load_gemstone_product_catalog().get('accounts') or {}).get('JEWELS SALES ACCOUNT - EMERALDS') or {}
_JEM_NUMBERS: list[int] = list(_em.get('emeralds_jem') or [])
_TAIL_PRODUCTS: tuple[str, ...] = tuple(_em.get('tail_products') or ())


def _pad_row(values: tuple[object, ...], width: int) -> tuple[object, ...]:
    lst = list(values)
    while len(lst) < width:
        lst.append(None)
    return tuple(lst[:width])


def main() -> None:
    wb = load_workbook(_WORKBOOK)
    ws = wb.active
    max_col = ws.max_column or 4
    rows: list[tuple[object, ...]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col):
        rows.append(_pad_row(tuple(c.value for c in row), max_col))

    emerald_start: int | None = None
    emerald_end: int | None = None
    for i, r in enumerate(rows):
        cell0 = r[0]
        if cell0 is None:
            continue
        s = str(cell0).strip()
        if 'jewels' in s.lower() and 'emerald' in s.lower():
            emerald_start = i
            break
    if emerald_start is None:
        raise RuntimeError('Could not find Jewels sales account - Emeralds row')

    for j in range(emerald_start + 1, len(rows)):
        r = rows[j]
        if r[0] is not None and str(r[0]).strip():
            emerald_end = j
            break
    if emerald_end is None:
        raise RuntimeError('Could not find row after Emeralds block')

    child_products: list[str] = [f'Emeralds JEM {n}' for n in _JEM_NUMBERS] + list(_TAIL_PRODUCTS)
    new_children = [_pad_row((None, name, None, None), max_col) for name in child_products]

    rebuilt = rows[: emerald_start + 1] + new_children + rows[emerald_end:]
    ws.delete_rows(1, ws.max_row)
    for r_idx, rvals in enumerate(rebuilt, start=1):
        for c_idx, val in enumerate(rvals, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(_WORKBOOK)


if __name__ == '__main__':
    main()
