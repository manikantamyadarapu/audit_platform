"""Enterprise sales audit: Sales Account category vs Product category (streaming openpyxl)."""

from __future__ import annotations

import time
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from app.config.settings import get_settings
from app.processors.base import BaseProcessor
from app.utils.excel_reader import effective_excel_max_row
from app.utils.header_cleaner import normalize_header
from app.utils.logger import get_logger
from app.utils.product_classifier import (
    classify_product_cached,
    expected_category_from_sales_account,
)
from app.utils.response_builder import build_processing_response

log = get_logger('sales-audit')

PROBE_MAX_ROW = 300
MAX_COL = 55
ISSUE_MSG = 'Product category does not match Sales Account'

_TOTAL_SKIP_SUBSTR = (
    'grand total',
    'sub total',
    'subtotal',
    'page total',
    'net total',
)


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
    return str(value).strip()


def _norm_cell_header(value: Any) -> str:
    return normalize_header(_cell_str(value))


def _row_header_layout(row: tuple[Any, ...] | None) -> tuple[int, int, int | None] | None:
    if not row:
        return None
    col_sa: int | None = None
    col_prod: int | None = None
    col_voucher: int | None = None
    for c in range(min(len(row), MAX_COL)):
        raw = row[c]
        if _is_blank(raw):
            continue
        n = _norm_cell_header(raw)
        if n == 'sales_account' and col_sa is None:
            col_sa = c + 1
        elif n == 'product' and col_prod is None:
            col_prod = c + 1
        elif n in ('voucher_no', 'voucher', 'voucher_number', 'voucher_num') and col_voucher is None:
            col_voucher = c + 1
    if col_sa is not None and col_prod is not None:
        return col_sa, col_prod, col_voucher
    return None


def _tuple_at(row: tuple[Any, ...] | None, col_1based: int) -> Any:
    if not row:
        return None
    i = col_1based - 1
    if i < 0 or i >= len(row):
        return None
    return row[i]


def _looks_like_repeated_header(sa_raw: Any, prod_raw: Any) -> bool:
    return _norm_cell_header(sa_raw) == 'sales_account' and _norm_cell_header(prod_raw) == 'product'


def _looks_like_total_line(sa_l: str) -> bool:
    if not sa_l:
        return False
    return any(needle in sa_l for needle in _TOTAL_SKIP_SUBSTR)


def _discover_sheet_layout(
    wb: Workbook,
) -> tuple[str, int, int, int, int, int | None, bool]:
    """
    Returns (sheet_title, header_row_1based, eff_max, col_sa, col_prod, col_voucher, truncated).
    ``wb`` must be an open read_only workbook (same instance used for row iteration).
    """
    settings = get_settings()
    cap = settings.excel_max_rows
    for sheet_title in wb.sheetnames:
        sheet = wb[sheet_title]
        mr = sheet.max_row or 0
        if mr == 0:
            continue
        probe_n = min(PROBE_MAX_ROW, mr)
        for row_idx, row in enumerate(
            sheet.iter_rows(
                min_row=1,
                max_row=probe_n,
                min_col=1,
                max_col=MAX_COL,
                values_only=True,
            ),
            start=1,
        ):
            layout = _row_header_layout(row)
            if layout is not None:
                col_sa, col_prod, col_voucher = layout
                eff_max, truncated = effective_excel_max_row(mr, cap)
                return sheet_title, row_idx, eff_max, col_sa, col_prod, col_voucher, truncated
    raise KeyError('Missing required columns: sales_account, product')


class SalesAuditProcessor(BaseProcessor):
    def process(self, file_bytes: bytes, **kwargs: Any) -> dict[str, Any]:
        filename = (kwargs.get('original_filename') or '').lower()
        if filename.endswith('.xls'):
            raise ValueError(
                'Sales audit fast path requires .xlsx or .xlsm; convert .xls or resave as .xlsx',
            )

        t_wall = time.perf_counter()
        t_read = time.perf_counter()

        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True, keep_links=False)
        try:
            sheet_title, header_row, eff_max, col_sa, col_prod, col_voucher, truncated = _discover_sheet_layout(wb)
        except Exception:
            wb.close()
            raise

        read_ms = (time.perf_counter() - t_read) * 1000.0

        records: list[dict[str, Any]] = []
        breakdown: dict[str, int] = {'14k': 0, '18k': 0, '22k': 0, '24k': 0, 'jadau': 0, 'unknown': 0}
        blank_skipped = 0
        skipped_header = 0
        skipped_total = 0
        fuzzy_matches = 0
        unknown_products = 0

        t_parse = time.perf_counter()
        try:
            ws = wb[sheet_title]
            for row_idx, row in enumerate(
                ws.iter_rows(
                    min_row=header_row + 1,
                    max_row=eff_max,
                    min_col=1,
                    max_col=MAX_COL,
                    values_only=True,
                ),
                start=header_row + 1,
            ):
                sa_raw = _tuple_at(row, col_sa)
                prod_raw = _tuple_at(row, col_prod)
                if _is_blank(sa_raw) and _is_blank(prod_raw):
                    blank_skipped += 1
                    continue
                if _looks_like_repeated_header(sa_raw, prod_raw):
                    skipped_header += 1
                    continue

                sa_disp = _cell_str(sa_raw)
                prod_disp = _cell_str(prod_raw)
                sa_l = sa_disp.lower()

                if _looks_like_total_line(sa_l):
                    skipped_total += 1
                    continue

                vraw = _tuple_at(row, col_voucher) if col_voucher else None
                v_disp = _cell_str(vraw) if not _is_blank(vraw) else ''
                voucher_no = v_disp or f'Row {row_idx}'

                expected = expected_category_from_sales_account(sa_disp)

                # Heavy classification only when the ledger row has a karat/jadau sales account.
                if expected is None:
                    predicted = None
                    used_fuzzy = False
                else:
                    predicted, used_fuzzy = classify_product_cached(prod_disp)
                    if used_fuzzy:
                        fuzzy_matches += 1

                if expected is None:
                    breakdown['unknown'] += 1
                    records.append(
                        {
                            'voucherNo': voucher_no,
                            'salesAccount': sa_disp,
                            'product': prod_disp,
                            'expectedCategory': None,
                            'predictedCategory': predicted,
                            'status': 'valid',
                            'issues': [],
                        },
                    )
                    continue

                breakdown[expected] += 1

                issues: list[str] = []
                if predicted is None:
                    unknown_products += 1
                    issues.append(ISSUE_MSG)
                    ok = False
                elif predicted != expected:
                    issues.append(ISSUE_MSG)
                    ok = False
                else:
                    ok = True

                records.append(
                    {
                        'voucherNo': voucher_no,
                        'salesAccount': sa_disp,
                        'product': prod_disp,
                        'expectedCategory': expected,
                        'predictedCategory': predicted,
                        'status': 'valid' if ok else 'invalid',
                        'issues': issues,
                    },
                )
        finally:
            wb.close()

        parse_ms = (time.perf_counter() - t_parse) * 1000.0

        total = len(records)
        invalid = sum(1 for r in records if r.get('status') == 'invalid')
        valid = total - invalid

        wall_ms = (time.perf_counter() - t_wall) * 1000.0
        rps = (total / (wall_ms / 1000.0)) if wall_ms > 1e-6 else 0.0

        summary = {
            'total': total,
            'valid': valid,
            'invalid': invalid,
            'fuzzyMatches': fuzzy_matches,
            'unknownProducts': unknown_products,
            'categoryBreakdown': breakdown,
        }

        skipped_no_rule = breakdown['unknown']

        row_stats = {
            'blankRowsSkipped': blank_skipped,
            'skippedNoRule': skipped_no_rule,
            'skippedHeaderRows': skipped_header,
            'skippedTotalLines': skipped_total,
            'headerRowExcel': header_row,
            'scanCapTruncated': truncated,
            'fuzzyMatches': fuzzy_matches,
            'unknownProducts': unknown_products,
            'engine': 'openpyxl_read_only',
        }

        performance = {
            'readTimeMs': round(read_ms, 3),
            'parseTimeMs': round(parse_ms, 3),
            'validateTimeMs': 0.0,
            'rowsPerSecond': round(rps, 2),
            'wallTimeMs': round(wall_ms, 3),
        }

        out = build_processing_response(
            file_type='sales_audit',
            total_rows=total,
            error_rows=invalid,
            summary=summary,
            records=records,
            performance=performance,
            row_stats=row_stats,
        )
        out['module'] = 'sales_audit'
        log.info(
            'Sales audit complete total={} valid={} invalid={} fuzzy={} unknownProd={} wallMs={:.1f}',
            total,
            valid,
            invalid,
            fuzzy_matches,
            unknown_products,
            wall_ms,
        )
        return out
