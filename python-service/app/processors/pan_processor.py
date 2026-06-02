import re
from io import BytesIO
from datetime import date, datetime
from pathlib import Path
from time import perf_counter
from typing import Any

import pandas as pd
import polars as pl
from openpyxl import load_workbook

from app.core.issue_engine import messages_for_codes
from app.config.settings import get_settings
from app.engines.vectorized_validation_engine import LoadedValidationSheet
from app.engines.vectorized_validation_engine import VectorizedValidationEngine
from app.processors.base import BaseProcessor
from app.utils.constants import SPREADSHEET_EMPTY_TOKENS, compact_pan_input_for_validation


from app.utils.header_cleaner import normalize_header
from app.utils.logger import get_logger
from app.utils.response_builder import build_processing_response


class PanProcessor(BaseProcessor):
    REQUIRED_BASE_COLUMNS = {'total_value'}
    PAN_COLUMN_OPTIONS = {'pan', 'pan1'}
    ADDRESS_COLUMN_OPTIONS = {'add_proof', 'add_proof_2'}

    def __init__(self) -> None:
        self.engine = VectorizedValidationEngine('pan')
        self._log = get_logger()

    def process(self, file_bytes: bytes) -> dict[str, Any]:
        total_start = perf_counter()
        try:
            loaded = self._load_pan_workbook(file_bytes)
        except Exception as exc:
            raise ValueError('Invalid or unreadable Excel file') from exc

        df = loaded.dataframe
        data_columns = self.engine.user_columns(df)
        self._validate_required_columns(data_columns)
        total_rows = len(df)

        validation_start = perf_counter()
        invalid_df = self._validate_dataframe(df, data_columns)
        validation_ms = (perf_counter() - validation_start) * 1000


        # Debug: print a small preview of the read/validated data in terminal
        if get_settings().debug_exports_enabled():
            try:
                # df is a Polars DataFrame; preview as dict rows
                df_preview = df.head(20).to_dicts()
                print('PAN audit - parsed dataframe preview (first 20 rows):')
                for r in df_preview:
                    print(r)

                invalid_preview = invalid_df.head(20).to_dicts()
                print('PAN audit - invalid rows preview (first 20 rows):')
                for r in invalid_preview:
                    print(r)

            except Exception as exc:
                print(f'PAN audit debug print failed: {exc}')


        extraction_start = perf_counter()

        records: list[dict[str, Any]] = []
        valid_pan_count = int(invalid_df['valid_pan_issue'].sum() or 0)
        invalid_pan_format_count = int(invalid_df['invalid_pan_issue'].sum() or 0)
        no_pan_no_form60_count = int(invalid_df['no_pan_no_form60_issue'].sum() or 0)
        no_pan_form60_available_count = int(invalid_df['no_pan_form60_available_issue'].sum() or 0)
        no_pan_invalid_form60_count = int(invalid_df['no_pan_invalid_form60_issue'].sum() or 0)
        # GST >= 50k address-related counts
        gst50k_address_missing_count = int(invalid_df['gst50k_address_missing_issue'].sum() or 0) if 'gst50k_address_missing_issue' in invalid_df.columns else 0
        incorrect_address_format_count = int(invalid_df['incorrect_address_format_issue'].sum() or 0) if 'incorrect_address_format_issue' in invalid_df.columns else 0
        valid_address_format_count = int(invalid_df['valid_address_format_issue'].sum() or 0) if 'valid_address_format_issue' in invalid_df.columns else 0
        eligible_pan_count = valid_pan_count + invalid_pan_format_count

        for invalid_row in invalid_df.to_dicts():
            row_num = invalid_row.get('row_number')
            
            issues: list[str] = []
            report_type = None
            address_report = None
            
            if invalid_row.get('invalid_pan_issue'):
                issues.append('INVALID_PAN_FORMAT')
                report_type = 'invalidPan'
            elif invalid_row.get('valid_pan_issue'):
                report_type = 'validPan'
            elif invalid_row.get('no_pan_no_form60_issue'):
                issues.append('NO_PAN_NO_FORM60')
                report_type = 'noPanNoForm60'
            elif invalid_row.get('no_pan_form60_available_issue'):
                issues.append('NO_PAN_FORM60_AVAILABLE')
                report_type = 'noPanForm60Available'
            elif invalid_row.get('no_pan_invalid_form60_issue'):
                issues.append('NO_PAN_INVALID_FORM60')
                report_type = 'noPanInvalidForm60'

            # GST >= 50k address checks
            if invalid_row.get('gst50k_address_missing_issue'):
                issues.append('MISSING_ADDRESS_PROOF_ABOVE_50K')
                address_report = 'gst50kAddressMissing'
            elif invalid_row.get('incorrect_address_format_issue'):
                issues.append('INVALID_ADDRESS')
                address_report = 'incorrectAddressFormat'
            elif invalid_row.get('valid_address_format_issue'):
                issues.append('VALID_ADDRESS_FORMAT')
                address_report = 'validAddressFormat'

            # Build record with all original columns
            record = {'rowNumber': self._json_value(row_num)}
            
            # Add all columns from original dataframe
            for col in data_columns:
                col_value = invalid_row.get(col)
                if col == 'total_value':
                    col_value = invalid_row.get('__total_value_amount')
                elif col == 'gross_amount':
                    col_value = invalid_row.get('__gross_amount_amount')
                # Use camelCase for output
                camel_col = self._to_camel_case(col)
                if col in ['date']:
                    record[camel_col] = self._format_cell_value(col_value)
                else:
                    record[camel_col] = self._json_value(col_value) if isinstance(col_value, (int, float)) else self._format_cell_value(col_value)
            
            record['issues'] = issues
            record['messages'] = self._messages_for_issues(issues)
            record['panReport'] = report_type or 'invalidPan'
            if address_report is not None:
                record['addressReport'] = address_report
            records.append(record)

        missing_address_proof_records: list[dict[str, Any]] = []
        invalid_address_records: list[dict[str, Any]] = []

        extraction_ms = (perf_counter() - extraction_start) * 1000
        self.engine.log_benchmark(
            row_count=total_rows,
            header_row_index=loaded.header_row_index,
            header_detection_ms=loaded.header_detection_ms,
            load_ms=loaded.load_ms,
            validation_ms=validation_ms,
            extraction_ms=extraction_ms,
            total_ms=(perf_counter() - total_start) * 1000,
        )

        summary = {
            'eligiblePanCount': eligible_pan_count,
            'validPanCount': valid_pan_count,
            'incorrectPanFormatCount': invalid_pan_format_count,
            'missingPanCount': 0,
            'missingForm60Count': 0,
            'invalidPanFormatCount': invalid_pan_format_count,
            'missingAddressProofCount': 0,
            'invalidAddressCount': 0,
            'missingPanAbove2L': 0,
            'missingForm60': 0,
            'invalidPanFormat': invalid_pan_format_count,
            'noPanNoForm60Count': no_pan_no_form60_count,
            'noPanForm60AvailableCount': no_pan_form60_available_count,
            'noPanInvalidForm60Count': no_pan_invalid_form60_count,
            'gst50kAddressMissingCount': gst50k_address_missing_count,
            'incorrectAddressFormatCount': incorrect_address_format_count,
            'validAddressFormatCount': valid_address_format_count,
            'missingAddressProofAbove50K': 0,
            'invalidAddress': 0,
        }


        if get_settings().debug_exports_enabled():
            self._export_issue_rows_debug(
                records,
                missing_address_proof_records,
                invalid_address_records,
            )

        response = build_processing_response(
            file_type='pan',
            total_rows=total_rows,
            error_rows=invalid_pan_format_count,
            summary=summary,
            records=records,
        )
        response['missingAddressProofRecords'] = missing_address_proof_records
        response['invalidAddressRecords'] = invalid_address_records
        return response

    def _export_issue_rows_debug(
        self,
        records: list[dict[str, Any]],
        missing_address_proof_records: list[dict[str, Any]],
        invalid_address_records: list[dict[str, Any]],
    ) -> None:
        if not records:
            return

        issue_rows = pd.DataFrame(records).head(200)
        missing_address_rows = pd.DataFrame(missing_address_proof_records).head(200)
        invalid_address_rows = pd.DataFrame(invalid_address_records).head(200)

        output_path = (
            Path(__file__).resolve().parents[2]
            / "pan_issue_rows_debug.xlsx"
        )
        with pd.ExcelWriter(output_path) as writer:
            issue_rows.to_excel(writer, index=False, sheet_name='All Issues')
            missing_address_rows.to_excel(
                writer,
                index=False,
                sheet_name='Missing Address Proof',
            )
            invalid_address_rows.to_excel(
                writer,
                index=False,
                sheet_name='Invalid Address',
            )

    @staticmethod
    def _records_with_issue(
        records: list[dict[str, Any]],
        issue: str,
    ) -> list[dict[str, Any]]:
        return [
            record
            for record in records
            if issue in record.get('issues', [])
        ]

    def _load_pan_workbook(self, file_bytes: bytes) -> LoadedValidationSheet:
        header_detection_ms = 0.0
        load_ms = 0.0
        loaded_sheets: list[LoadedValidationSheet] = []

        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                loaded = self._load_matching_worksheet(worksheet)
                header_detection_ms += loaded.header_detection_ms
                load_ms += loaded.load_ms
                if loaded.dataframe.is_empty() and not self.engine.user_columns(loaded.dataframe):
                    continue
                if self.engine.user_columns(loaded.dataframe):
                    loaded_sheets.append(loaded)
        finally:
            workbook.close()

        if not loaded_sheets:
            return LoadedValidationSheet(
                dataframe=pl.DataFrame(
                    schema={
                        'source_excel_row_number': pl.Int64,
                        '__excel_row_number__': pl.Int64,
                    }
                ),
                header_row_index=0,
                header_detection_ms=header_detection_ms,
                load_ms=load_ms,
            )

        combined_sheet = self._first_sheet_matching(
            loaded_sheets,
            lambda columns: self._headers_match_pan_sheet(columns) and self._headers_match_address_sheet(columns),
        )
        if combined_sheet is not None:
            combined_sheet.header_detection_ms = header_detection_ms
            combined_sheet.load_ms = load_ms
            return combined_sheet

        pan_sheet = self._first_sheet_matching(loaded_sheets, self._headers_match_pan_sheet)
        address_sheet = self._first_sheet_matching(loaded_sheets, self._headers_match_address_sheet)

        if pan_sheet is None:
            pan_sheet = loaded_sheets[0]

        if address_sheet is None or address_sheet is pan_sheet:
            pan_sheet.header_detection_ms = header_detection_ms
            pan_sheet.load_ms = load_ms
            return pan_sheet

        dataframe = self._merge_pan_and_address_frames(
            pan_sheet.dataframe,
            address_sheet.dataframe,
        )
        return LoadedValidationSheet(
            dataframe=dataframe,
            header_row_index=pan_sheet.header_row_index,
            header_detection_ms=header_detection_ms,
            load_ms=load_ms,
        )

    def _load_matching_worksheet(self, worksheet: Any) -> LoadedValidationSheet:
        header_scan_start = perf_counter()
        header_row_index: int | None = None
        headers: list[str] = []
        positions: list[int] = []
        columns: dict[str, list[Any]] = {}
        source_excel_row_numbers: list[int] = []

        for physical_row, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            row_tuple = tuple(row)
            if header_row_index is None:
                labels = self._header_labels_from_values(row_tuple)
                if self._headers_match_relevant_sheet(labels):
                    header_row_index = physical_row - 1
                    headers, positions = self._normalize_headers(row_tuple)
                    columns = {header: [] for header in headers}
                continue
            if not headers:
                continue
            self._append_row(columns, row_tuple, headers, positions)
            source_excel_row_numbers.append(physical_row)

        if header_row_index is None:
            header_row_index = 0

        header_detection_ms = (perf_counter() - header_scan_start) * 1000
        load_start = perf_counter()
        if headers and source_excel_row_numbers:
            excel_rows = pl.Series('source_excel_row_number', source_excel_row_numbers)
            dataframe = pl.DataFrame(columns, strict=False).with_columns(
                excel_rows.alias('source_excel_row_number'),
                excel_rows.alias('__excel_row_number__'),
            )
        elif headers:
            dataframe = pl.DataFrame(columns, strict=False).with_columns(
                pl.lit(None).cast(pl.Int64).alias('source_excel_row_number'),
                pl.lit(None).cast(pl.Int64).alias('__excel_row_number__'),
            )
        else:
            dataframe = pl.DataFrame(
                schema={
                    'source_excel_row_number': pl.Int64,
                    '__excel_row_number__': pl.Int64,
                }
            )
        load_ms = (perf_counter() - load_start) * 1000

        return LoadedValidationSheet(
            dataframe=dataframe,
            header_row_index=header_row_index,
            header_detection_ms=header_detection_ms,
            load_ms=load_ms,
        )

    def _merge_pan_and_address_frames(self, pan_df: pl.DataFrame, address_df: pl.DataFrame) -> pl.DataFrame:
        address_columns = [
            column
            for column in self.engine.user_columns(address_df)
            if column not in pan_df.columns
        ]
        if not address_columns:
            return pan_df

        if 'voucher_no' in pan_df.columns and 'voucher_no' in address_df.columns:
            right = address_df.select(['voucher_no', *address_columns])
            return pan_df.join(right, on='voucher_no', how='left')

        row_indexed_pan = pan_df.with_row_index('__pan_row_index__')
        row_indexed_address = address_df.select(address_columns).with_row_index('__pan_row_index__')
        return (
            row_indexed_pan
            .join(row_indexed_address, on='__pan_row_index__', how='left')
            .drop('__pan_row_index__')
        )

    def _validate_dataframe(self, df: pl.DataFrame, data_columns: list[str]) -> pl.DataFrame:
        working = self._ensure_columns(df, ['total_value', 'pan', 'pan1', '__excel_row_number__', 'add_proof', 'add_proof_2', 'address'])
        column_set = set(data_columns)

        validated = working.with_columns(
            pl.col('__excel_row_number__').cast(pl.Int64, strict=False).alias('row_number'),
            pl.col('total_value').map_elements(self._parse_amount_float, return_dtype=pl.Float64).alias('__total_value_amount'),
            pl.col('gross_amount').map_elements(self._parse_amount_float, return_dtype=pl.Float64).alias('__gross_amount_amount'),
            pl.col('pan').map_elements(self.normalize_empty_value, return_dtype=pl.Utf8).alias('__pan_text'),
            pl.col('pan1').map_elements(self.normalize_empty_value, return_dtype=pl.Utf8).alias('__pan1_text'),
            pl.col('add_proof').map_elements(self.normalize_empty_value, return_dtype=pl.Utf8).alias('__add_proof_text'),
            pl.col('add_proof_2').map_elements(self.normalize_empty_value, return_dtype=pl.Utf8).alias('__add_proof_2_text'),
            pl.col('address').map_elements(self.normalize_empty_value, return_dtype=pl.Utf8).alias('__address_text'),
            pl.struct(data_columns).map_elements(
                lambda row: self._should_skip_row_dict(dict(row), column_set),
                return_dtype=pl.Boolean,
            ).alias('__should_skip'),
        ).with_columns(
            pl.col('__pan_text').map_elements(
                lambda value: value is not None and self.is_valid_pan(value),
                return_dtype=pl.Boolean,
            ).alias('__pan_ok'),
            pl.col('__pan1_text').map_elements(
                lambda value: value is not None and self.is_valid_pan(value),
                return_dtype=pl.Boolean,
            ).alias('__pan1_ok'),
            pl.col('__add_proof_text').map_elements(
                lambda value: value is not None and len(value) > 5,
                return_dtype=pl.Boolean,
            ).alias('__add_proof_valid'),
            pl.col('__add_proof_2_text').map_elements(
                lambda value: value is not None and len(value) > 5,
                return_dtype=pl.Boolean,
            ).alias('__add_proof_2_valid'),
            pl.col('__address_text').map_elements(
                lambda value: value is not None and len(value) > 5,
                return_dtype=pl.Boolean,
            ).alias('__address_valid'),
        )

        

        should_check = pl.col('__should_skip').fill_null(False).not_()
        pan_needed = pl.col('__total_value_amount') > 200000
        pan_valid = pl.col('__pan_ok').fill_null(False) | pl.col('__pan1_ok').fill_null(False)
        pan_present = pl.col('__pan_text').is_not_null() | pl.col('__pan1_text').is_not_null()
        pan_invalid_specific = (
            (pl.col('__pan_text').is_not_null() & pl.col('__pan_ok').fill_null(False).not_())
            | (pl.col('__pan1_text').is_not_null() & pl.col('__pan1_ok').fill_null(False).not_())
        )
        
        # Only check add_proof/add_proof_2/address when BOTH pan and pan1 are blank
        pan_both_blank = pl.col('__pan_text').is_null() & pl.col('__pan1_text').is_null()

        # Form 60 validation logic (only when pan_needed AND pan_both_blank)
        add_proof_text = pl.col('__add_proof_text')
        add_proof_2_text = pl.col('__add_proof_2_text')
        address_text = pl.col('__address_text')
        add_proof_valid = pl.col('__add_proof_valid').fill_null(False)
        add_proof_2_valid = pl.col('__add_proof_2_valid').fill_null(False)
        address_valid = pl.col('__address_valid').fill_null(False)
        
        # All three are blank (add_proof, add_proof_2, address)
        all_form60_blank = add_proof_text.is_null() & add_proof_2_text.is_null() & address_text.is_null()
        
        # Any of the three has length > 5 (form 60 available)
        form60_available = add_proof_valid | add_proof_2_valid | address_valid
        
        # Any of the three is not blank but has length <= 5, AND none have length > 5 (form 60 invalid)
        form60_has_short = (
            (add_proof_text.is_not_null() & add_proof_valid.not_())
            | (add_proof_2_text.is_not_null() & add_proof_2_valid.not_())
            | (address_text.is_not_null() & address_valid.not_())
        )
        form60_invalid = form60_has_short & form60_available.not_()

        # Gross amount checks for address validation when gross_amount >= 50000
        gross_needed = pl.col('__gross_amount_amount') >= 50000

        with_issues = validated.with_columns(
            (should_check & pan_needed & pan_present & pan_valid & pan_invalid_specific.not_())
            .fill_null(False)
            .alias('valid_pan_issue'),
            (should_check & pan_needed & pan_present & pan_invalid_specific)
            .fill_null(False)
            .alias('invalid_pan_issue'),
            (should_check & pan_needed & pan_both_blank & all_form60_blank)
            .fill_null(False)
            .alias('no_pan_no_form60_issue'),
            (should_check & pan_needed & pan_both_blank & form60_available & all_form60_blank.not_())
            .fill_null(False)
            .alias('no_pan_form60_available_issue'),
            (should_check & pan_needed & pan_both_blank & form60_invalid)
            .fill_null(False)
            .alias('no_pan_invalid_form60_issue'),
            # GST >= 50k address checks
            (should_check & gross_needed & all_form60_blank)
            .fill_null(False)
            .alias('gst50k_address_missing_issue'),
            (should_check & gross_needed & form60_invalid)
            .fill_null(False)
            .alias('incorrect_address_format_issue'),
            (should_check & gross_needed & form60_available & all_form60_blank.not_())
            .fill_null(False)
            .alias('valid_address_format_issue'),
        )

        return with_issues.filter(
            pl.col('valid_pan_issue') | pl.col('invalid_pan_issue') | pl.col('no_pan_no_form60_issue')
            | pl.col('no_pan_form60_available_issue') | pl.col('no_pan_invalid_form60_issue')
            | pl.col('gst50k_address_missing_issue') | pl.col('incorrect_address_format_issue') | pl.col('valid_address_format_issue')
        ).sort('row_number')

    def _parse_amount_float(self, value: Any) -> float | None:
        parsed = self.parse_amount(value)
        return float(parsed) if parsed is not None else None

    def normalize_empty_value(self, value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, float) and pd.isna(value):
            return None

        text = str(value).strip()
        if not text:
            return None
        if text.lower() in SPREADSHEET_EMPTY_TOKENS:
            return None
        return text

    def parse_amount(self, value: Any) -> float | int | None:
        if value is None:
            return None
        if isinstance(value, (int, float)) and not pd.isna(value):
            return int(value) if float(value).is_integer() else float(value)

        text = str(value).strip()
        if not text:
            return None
        if text.lower() in SPREADSHEET_EMPTY_TOKENS:
            return None

        cleaned = re.sub(r'[^0-9.\-]', '', text.replace(',', ''))
        if cleaned.count('.') > 1:
            parts = cleaned.split('.')
            cleaned = f"{''.join(parts[:-1])}.{parts[-1]}"
        if cleaned in {'', '-', '.', '-.'}:
            return None

        number = float(cleaned)
        return int(number) if number.is_integer() else number

    @staticmethod
    def _messages_for_issues(issues: list[str]) -> list[str]:
        return messages_for_codes(issues)

    def is_valid_pan(self, pan_value: str) -> bool:
        compact = compact_pan_input_for_validation(pan_value)
        return bool(compact and re.fullmatch(r'^[A-Z]{5}[0-9]{4}[A-Z]$', compact))

    def _validate_required_columns(self, columns: Any) -> None:
        column_set = set(columns)
        missing_base = self.REQUIRED_BASE_COLUMNS - column_set
        if missing_base:
            raise KeyError(f"Missing required columns: {', '.join(sorted(missing_base))}")

        missing_pan_columns = self.PAN_COLUMN_OPTIONS - column_set
        if missing_pan_columns:
            raise KeyError(f"Missing required columns: {', '.join(sorted(missing_pan_columns))}")

        if 'gross_amount' not in column_set:
            self._log.warning('gross_amount column not found; invalid address validation will not flag rows')

    def _columns_sufficient_for_pan(self, cols: set[str]) -> bool:
        return 'pan' in cols or 'pan1' in cols

    def _headers_match_pan_sheet(self, headers: set[str]) -> bool:
        if 'total_value' not in headers or not ('pan' in headers or 'pan1' in headers):
            return False
        return True

    def _headers_match_address_sheet(self, headers: set[str]) -> bool:
        return bool(
            headers
            & {
                'gross_amount',
                'address',
                'add_proof',
                'add_proof_2',
            }
        )

    def _headers_match_relevant_sheet(self, headers: set[str]) -> bool:
        return self._headers_match_pan_sheet(headers) or self._headers_match_address_sheet(headers)

    def _first_sheet_matching(
        self,
        loaded_sheets: list[LoadedValidationSheet],
        predicate: Any,
    ) -> LoadedValidationSheet | None:
        for loaded in loaded_sheets:
            if predicate(set(self.engine.user_columns(loaded.dataframe))):
                return loaded
        return None

    def _should_skip_row_dict(self, row: dict[str, Any], columns: set[str]) -> bool:
        if all(self.normalize_empty_value(value) is None for value in row.values()):
            return True

        total_value = row.get('total_value')
        if self.normalize_empty_value(total_value) is not None:
            if normalize_header(total_value) == 'total_value':
                return True

        pan = row.get('pan')
        pan1 = row.get('pan1')
        if isinstance(pan, str) and pan.strip().upper() == 'PAN':
            if isinstance(pan1, str) and pan1.strip().upper().replace(' ', '') == 'PAN1':
                return True

        voucher = row.get('voucher_no')
        if isinstance(voucher, str):
            voucher_text = voucher.strip()
            if normalize_header(voucher_text) == 'voucher_no':
                return True
            if re.search(r'(^\s*sub\s*total\b|\bgrand\s*total\b|^\s*total\s*$)', voucher_text, re.IGNORECASE):
                return True

        for key in ('party', 'narration', 'description'):
            value = row.get(key)
            if isinstance(value, str) and re.search(
                r'(^\s*sub\s*total\b|\bgrand\s*total\b|^\s*total\s*$)',
                value.strip(),
                re.IGNORECASE,
            ):
                return True

        if 'voucher_no' in columns and self.normalize_empty_value(row.get('voucher_no')) is None:
            return True

        return False

    def _ensure_columns(self, df: pl.DataFrame, columns: list[str]) -> pl.DataFrame:
        missing = [column for column in columns if column not in df.columns]
        if not missing:
            return df
        return df.with_columns([pl.lit(None).alias(column) for column in missing])

    @staticmethod
    def _header_labels_from_values(row: Any) -> set[str]:
        labels: set[str] = set()
        for value in row:
            label = normalize_header(value)
            if label:
                labels.add(label)
        return labels

    @staticmethod
    def _normalize_headers(row: Any) -> tuple[list[str], list[int]]:
        headers: list[str] = []
        positions: list[int] = []
        seen: dict[str, int] = {}
        for idx, raw_header in enumerate(row):
            label = normalize_header(raw_header)
            if not label:
                continue
            seen[label] = seen.get(label, 0) + 1
            unique_label = label if seen[label] == 1 else f'{label}_{seen[label]}'
            headers.append(unique_label)
            positions.append(idx)
        return headers, positions

    @staticmethod
    def _append_row(
        columns: dict[str, list[Any]],
        row: Any,
        headers: list[str],
        positions: list[int],
    ) -> None:
        for header, position in zip(headers, positions, strict=True):
            value = row[position] if position < len(row) else None
            columns[header].append(value)

    def _format_cell_value(self, value: Any) -> str:
        if value is None:
            return ''
        if pd.isna(value):
            return ''
        if isinstance(value, pd.Timestamp):
            return value.strftime('%d-%m-%Y')
        if isinstance(value, datetime):
            return value.strftime('%d-%m-%Y')
        if isinstance(value, date):
            return value.strftime('%d-%m-%Y')
        return str(value).strip()

    def _json_value(self, value: Any) -> Any:
        if value is None:
            return None
        if pd.isna(value):
            return None
        if hasattr(value, 'item'):
            return value.item()
        return value

    @staticmethod
    def _to_camel_case(snake_str: str) -> str:
        """Convert snake_case to camelCase."""
        components = snake_str.split('_')
        return components[0] + ''.join(x.title() for x in components[1:])
