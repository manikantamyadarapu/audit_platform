from __future__ import annotations

import math
import re
import time
from collections import deque
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config.settings import get_settings
from app.processors.base import BaseProcessor
from app.validators.gross_weight_validator import display_float_two_dp, validate_triplet
from app.utils.excel_reader import effective_excel_max_row, tuple_cell_1
from app.utils.header_cleaner import normalize_header
from app.utils.logger import get_logger
from app.utils.response_builder import build_processing_response

log = get_logger('gross-weight')

_VOUCHER_MARKER = re.compile(r'voucher\s*no', re.I)
_VOUCHER_TAIL = re.compile(r'voucher\s*no\s*[.:\s\-]*(.+)$', re.I | re.S)

COL_MANUAL = 2
COL_AUTO = 3
COL_DIFF = 4
MAX_COL_SCAN = 24
DATA_ROW_LOOKAHEAD = 12
# Workbooks often have cover rows; headers may sit well below row 50.
TABULAR_HEADER_MAX_PROBE_ROW = 500


def _sniff_extension(file_bytes: bytes, filename: str) -> str:
    ext = Path(filename or '').suffix.lower()
    if ext in {'.xlsx', '.xlsm', '.xls'}:
        return ext
    if len(file_bytes) >= 2 and file_bytes[:2] == b'PK':
        return '.xlsx'
    if len(file_bytes) >= 8 and file_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return '.xls'
    return '.xlsx'


def _is_blank_cell(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return True
    s = str(value).strip().replace('\u00a0', ' ')
    if not s:
        return True
    if s.lower() in {'', 'na', 'n/a', 'null', 'none', '-', '--', 'n.a.'}:
        return True
    return False


def _normalize_spaces(text: str) -> str:
    return ' '.join(text.replace('\u00a0', ' ').split())


def _parse_number(value: Any) -> float | None:
    """Parse numeric cell; supports comma thousands/decimals and stray spaces."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    s = str(value).strip().replace('\u00a0', ' ')
    s = s.replace(',', '')
    if not s or s.lower() in {'na', 'n/a', 'null', 'none', '-', '--'}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _normalize_header_cell(value: Any) -> str:
    return normalize_header(str(value) if value is not None else '')


def _is_manual_gross_header(norm: str) -> bool:
    if not norm or 'difference' in norm:
        return False
    if 'auto' in norm and 'manual' not in norm:
        return False
    return 'manual' in norm and ('gross' in norm or 'wt' in norm or 'weight' in norm)


def _is_auto_gross_header(norm: str) -> bool:
    if not norm or 'difference' in norm or 'manual' in norm:
        return False
    return 'auto' in norm and ('gross' in norm or 'wt' in norm or 'weight' in norm)


def _is_difference_gross_header(norm: str) -> bool:
    if not norm:
        return False
    if 'difference' in norm or norm.startswith('diff'):
        return 'gross' in norm or 'wt' in norm or 'weight' in norm
    return False


def _find_tabular_header_layout_from_probe(
    probe: list[tuple[int, tuple[Any, ...]]],
    max_col: int = 40,
) -> tuple[int, int, int, int, int] | None:
    """
    Locate header row and column indices from pre-read probe rows.

    Each probe entry is ``(excel_row_1based, row_tuple)`` with column A at index 0.
    """
    for hr, row_tuple in probe:
        norms: dict[int, str] = {}
        width = min(max_col, len(row_tuple))
        for c in range(1, width + 1):
            raw = row_tuple[c - 1] if c - 1 < len(row_tuple) else None
            norms[c] = _normalize_header_cell(raw)

        manual_col = None
        auto_col = None
        diff_col = None
        for c, n in norms.items():
            if not n:
                continue
            if manual_col is None and _is_manual_gross_header(n):
                manual_col = c
            elif auto_col is None and _is_auto_gross_header(n):
                auto_col = c
            elif diff_col is None and _is_difference_gross_header(n):
                diff_col = c

        if manual_col is not None and auto_col is not None and diff_col is not None:
            if len({manual_col, auto_col, diff_col}) < 3:
                continue
            voucher_col = manual_col - 1 if manual_col > 1 else 1
            return hr, voucher_col, manual_col, auto_col, diff_col

    return None


def _tabular_row_looks_like_repeated_header(voucher_line: str, manual_raw: Any) -> bool:
    """Row where column A repeats 'SNo' and the weight column again shows a header label."""
    v = voucher_line.strip().lower()
    if v not in {'sno', 'sl no', 'sl_no', 'serial', 'serial no', '#'}:
        return False
    if manual_raw is None or _is_blank_cell(manual_raw):
        return True
    return _is_manual_gross_header(_normalize_header_cell(manual_raw))


def _tabular_row_should_skip_non_data_row(
    voucher_line: str,
    manual_raw: Any,
    auto_raw: Any,
    diff_raw: Any,
) -> bool:
    """
    Skip rows that are not real voucher data (PAN-style: do not emit as valid or invalid).

    Covers repeated header rows, header text pasted into weight columns, SNo-only labels,
    and common guard / summary lines (grand total, subtotal, filters) when they are not
    a `Voucher No:` line.
    """
    if _tabular_row_looks_like_repeated_header(voucher_line, manual_raw):
        return True

    v_st = voucher_line.strip()
    if v_st:
        low = v_st.lower()
        nv = _normalize_header_cell(v_st)
        if nv in {'sno', 'serial_no', 'sl_no'} or low in {'sno', 'sl no', 'sl_no', 'serial', 'serial no', '#'}:
            return True
        if _is_manual_gross_header(nv) or _is_auto_gross_header(nv) or _is_difference_gross_header(nv):
            return True
        if not _VOUCHER_MARKER.search(v_st):
            if any(
                t in low
                for t in (
                    'grand total',
                    'sub total',
                    'subtotal',
                    'page total',
                    'filter applied',
                    'total gross',
                )
            ):
                return True

    for raw in (manual_raw, auto_raw, diff_raw):
        if raw is None or _is_blank_cell(raw):
            continue
        if isinstance(raw, (int, float)) and not isinstance(raw, bool):
            continue
        n = _normalize_header_cell(str(raw))
        if _is_manual_gross_header(n) or _is_auto_gross_header(n) or _is_difference_gross_header(n):
            return True
        if n in {'sno', 'serial_no', 'sl_no', 'voucher_no'}:
            return True

    return False


def _build_records_from_tabular_openpyxl(
    ws: Worksheet,
    header_row: int,
    voucher_col: int,
    manual_col: int,
    auto_col: int,
    diff_col: int,
    effective_max_row: int,
) -> tuple[list[dict[str, Any]], dict[str, int], dict[str, int]]:
    """
    Tabular body rows via ``iter_rows(values_only=True)`` — avoids per-cell ``ws.cell()`` calls.

    ``effective_max_row`` is the last physical row to scan (may reflect ``EXCEL_MAX_ROWS`` cap).
    """
    records: list[dict[str, Any]] = []
    mismatch_manual_auto = 0
    difference_violations = 0
    diff_only_violations = 0
    max_c = max(voucher_col, manual_col, auto_col, diff_col)
    stats = {
        'tabularRawRowsScanned': 0,
        'tabularBlankRowsSkipped': 0,
        'tabularSkippedNonDataRows': 0,
        'tabularParsedRows': 0,
    }

    if effective_max_row <= header_row:
        return (
            records,
            {'mismatchManualAuto': 0, 'differenceViolations': 0, 'diffOnlyViolations': 0},
            stats,
        )

    for r, row in enumerate(
        ws.iter_rows(
            min_row=header_row + 1,
            max_row=effective_max_row,
            min_col=1,
            max_col=max_c,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        stats['tabularRawRowsScanned'] += 1
        manual_raw = tuple_cell_1(row, manual_col)
        auto_raw = tuple_cell_1(row, auto_col)
        diff_raw = tuple_cell_1(row, diff_col)
        vraw = tuple_cell_1(row, voucher_col)

        if (
            _is_blank_cell(manual_raw)
            and _is_blank_cell(auto_raw)
            and _is_blank_cell(diff_raw)
            and _is_blank_cell(vraw)
        ):
            stats['tabularBlankRowsSkipped'] += 1
            continue

        voucher_line = str(vraw).strip() if not _is_blank_cell(vraw) else ''
        if _tabular_row_should_skip_non_data_row(voucher_line, manual_raw, auto_raw, diff_raw):
            stats['tabularSkippedNonDataRows'] += 1
            continue

        voucher_no = _extract_voucher_no(voucher_line) if voucher_line else ''
        if not voucher_no and voucher_line:
            voucher_no = _normalize_spaces(voucher_line)
        if not voucher_no:
            voucher_no = f'Row {r}'

        manual = _parse_number(manual_raw)
        auto = _parse_number(auto_raw)
        diff = _parse_number(diff_raw)

        issues, mm, dv, dov = validate_triplet(manual, auto, diff)
        if mm:
            mismatch_manual_auto += 1
        if dv:
            difference_violations += 1
        if dov:
            diff_only_violations += 1

        manual_disp = display_float_two_dp(manual) if manual is not None else None
        auto_disp = display_float_two_dp(auto) if auto is not None else None
        diff_disp = display_float_two_dp(diff) if diff is not None else None

        status = 'valid' if not issues else 'invalid'
        stats['tabularParsedRows'] += 1
        records.append(
            {
                'voucherNo': voucher_no,
                'manualGross': manual_disp,
                'autoGross': auto_disp,
                'difference': diff_disp,
                'status': status,
                'issues': issues,
                'voucherRow': r,
                'dataRow': r,
            }
        )

    return (
        records,
        {
            'mismatchManualAuto': mismatch_manual_auto,
            'differenceViolations': difference_violations,
            'diffOnlyViolations': diff_only_violations,
        },
        stats,
    )


def _extract_voucher_no(line: str) -> str | None:
    if not line or not _VOUCHER_MARKER.search(line):
        return None
    m = _VOUCHER_TAIL.search(line.strip())
    if not m:
        return None
    tail = _normalize_spaces(m.group(1).strip())
    return tail or None


def _first_non_empty_sheet_openpyxl(wb: Any) -> Worksheet:
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row and ws.max_row > 0:
            return ws
    return wb.active


def _sheet_quick_has_voucher_openpyxl(ws: Worksheet, max_probe_rows: int = 800) -> bool:
    """Fast probe so we pick a data sheet, not an empty cover sheet before a populated tab."""
    mr = min(ws.max_row or 0, max_probe_rows)
    mc = min(ws.max_column or MAX_COL_SCAN, MAX_COL_SCAN)
    if mr == 0:
        return False
    for row in ws.iter_rows(min_row=1, max_row=mr, min_col=1, max_col=mc, values_only=True):
        for raw in row:
            if raw is not None and not _is_blank_cell(raw) and _VOUCHER_MARKER.search(str(raw)):
                return True
    return False


def _pick_worksheet_openpyxl(wb: Any) -> Worksheet:
    for name in wb.sheetnames:
        ws = wb[name]
        if not (ws.max_row or 0):
            continue
        if _sheet_quick_has_voucher_openpyxl(ws):
            return ws
    return _first_non_empty_sheet_openpyxl(wb)


def _find_data_row(
    get_cell: Callable[[int, int], Any],
    start_row: int,
    max_row: int,
) -> int | None:
    """First row after voucher with any non-blank value in B/C/D (skips blank spacer rows)."""
    end = min(start_row + DATA_ROW_LOOKAHEAD - 1, max_row)
    for r in range(start_row, end + 1):
        b = get_cell(r, COL_MANUAL)
        c = get_cell(r, COL_AUTO)
        d = get_cell(r, COL_DIFF)
        if not (_is_blank_cell(b) and _is_blank_cell(c) and _is_blank_cell(d)):
            return r
    return None


def _row_get_cell_openpyxl(ws: Worksheet, r: int, c: int) -> Any:
    """Single-row read for semi-structured layout (voucher blocks, sparse workbooks)."""
    rows = list(
        ws.iter_rows(min_row=r, max_row=r, min_col=1, max_col=MAX_COL_SCAN, values_only=True)
    )
    if not rows:
        return None
    return tuple_cell_1(rows[0], c)


def _scan_openpyxl_sheet(ws: Worksheet, max_row: int) -> list[dict[str, Any]]:
    """Forward scan with bounded lookahead — compatible with ``read_only`` worksheets."""
    if max_row == 0:
        return []

    pending: deque[tuple[int, tuple[Any, ...]]] = deque()
    row_iter = enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=MAX_COL_SCAN,
            values_only=True,
        ),
        start=1,
    )

    def take_row() -> tuple[int, tuple[Any, ...]] | None:
        if pending:
            return pending.popleft()
        try:
            return next(row_iter)
        except StopIteration:
            return None

    blocks: list[dict[str, Any]] = []
    seen_voucher_keys: set[str] = set()

    while True:
        item = take_row()
        if item is None:
            break
        r, row_tuple = item
        voucher_line: str | None = None
        voucher_col = 1
        for c, raw in enumerate(row_tuple, start=1):
            if raw is None or _is_blank_cell(raw):
                continue
            text = str(raw).strip()
            if _VOUCHER_MARKER.search(text):
                voucher_line = text
                voucher_col = c
                break
        if not voucher_line:
            continue

        buf: list[tuple[int, tuple[Any, ...]]] = []
        for _ in range(DATA_ROW_LOOKAHEAD):
            nxt = take_row()
            if nxt is None:
                break
            buf.append(nxt)

        voucher_no = _extract_voucher_no(voucher_line) or _normalize_spaces(
            voucher_line.replace('Voucher No', '').replace(':', '').strip()
        )

        data_row: int | None = None
        for br, bt in buf:
            b = tuple_cell_1(bt, COL_MANUAL)
            c = tuple_cell_1(bt, COL_AUTO)
            d = tuple_cell_1(bt, COL_DIFF)
            if not (_is_blank_cell(b) and _is_blank_cell(c) and _is_blank_cell(d)):
                data_row = br
                break

        key = f'{r}:{voucher_no}'
        if key not in seen_voucher_keys:
            seen_voucher_keys.add(key)
            blocks.append(
                {
                    'voucher_row': r,
                    'voucher_col': voucher_col,
                    'voucher_line': voucher_line,
                    'voucher_no': voucher_no,
                    'data_row': data_row,
                }
            )

        if data_row is not None:
            found_at = next((i for i, (br, _) in enumerate(buf) if br == data_row), None)
            if found_at is not None:
                for u in reversed(buf[found_at + 1 :]):
                    pending.appendleft(u)
        else:
            for u in reversed(buf):
                pending.appendleft(u)

    return blocks


def _xlrd_cell_value(sheet: Any, row_1: int, col_1: int) -> Any:
    r0, c0 = row_1 - 1, col_1 - 1
    if r0 < 0 or c0 < 0 or r0 >= sheet.nrows or c0 >= sheet.ncols:
        return None
    merged = getattr(sheet, 'merged_cells', []) or []
    for rlo, rhi, clo, chi in merged:
        if rlo <= r0 < rhi and clo <= c0 < chi:
            return sheet.cell_value(rlo, clo)
    return sheet.cell_value(r0, c0)


def _find_tabular_header_layout_xlrd(sheet: Any) -> tuple[int, int, int, int, int] | None:
    max_probe_row = min(sheet.nrows, TABULAR_HEADER_MAX_PROBE_ROW)
    max_col = min(sheet.ncols, 40)
    if max_probe_row == 0:
        return None
    probe: list[tuple[int, tuple[Any, ...]]] = []
    for hr in range(1, max_probe_row + 1):
        vals = tuple(_xlrd_cell_value(sheet, hr, c) for c in range(1, max_col + 1))
        probe.append((hr, vals))
    return _find_tabular_header_layout_from_probe(probe, max_col)


def _build_records_from_tabular_xlrd(
    sheet: Any,
    header_row: int,
    voucher_col: int,
    manual_col: int,
    auto_col: int,
    diff_col: int,
    effective_max_row: int,
) -> tuple[list[dict[str, Any]], dict[str, int], dict[str, int]]:
    records: list[dict[str, Any]] = []
    mismatch_manual_auto = 0
    difference_violations = 0
    diff_only_violations = 0
    stats = {
        'tabularRawRowsScanned': 0,
        'tabularBlankRowsSkipped': 0,
        'tabularSkippedNonDataRows': 0,
        'tabularParsedRows': 0,
    }
    for r in range(header_row + 1, effective_max_row + 1):
        stats['tabularRawRowsScanned'] += 1
        manual_raw = _xlrd_cell_value(sheet, r, manual_col)
        auto_raw = _xlrd_cell_value(sheet, r, auto_col)
        diff_raw = _xlrd_cell_value(sheet, r, diff_col)
        vraw = _xlrd_cell_value(sheet, r, voucher_col)
        if (
            _is_blank_cell(manual_raw)
            and _is_blank_cell(auto_raw)
            and _is_blank_cell(diff_raw)
            and _is_blank_cell(vraw)
        ):
            stats['tabularBlankRowsSkipped'] += 1
            continue
        voucher_line = str(vraw).strip() if not _is_blank_cell(vraw) else ''
        if _tabular_row_should_skip_non_data_row(voucher_line, manual_raw, auto_raw, diff_raw):
            stats['tabularSkippedNonDataRows'] += 1
            continue
        voucher_no = _extract_voucher_no(voucher_line) if voucher_line else ''
        if not voucher_no and voucher_line:
            voucher_no = _normalize_spaces(voucher_line)
        if not voucher_no:
            voucher_no = f'Row {r}'
        manual = _parse_number(manual_raw)
        auto = _parse_number(auto_raw)
        diff = _parse_number(diff_raw)
        issues, mm, dv, dov = validate_triplet(manual, auto, diff)
        if mm:
            mismatch_manual_auto += 1
        if dv:
            difference_violations += 1
        if dov:
            diff_only_violations += 1
        manual_disp = display_float_two_dp(manual) if manual is not None else None
        auto_disp = display_float_two_dp(auto) if auto is not None else None
        diff_disp = display_float_two_dp(diff) if diff is not None else None
        status = 'valid' if not issues else 'invalid'
        stats['tabularParsedRows'] += 1
        records.append(
            {
                'voucherNo': voucher_no,
                'manualGross': manual_disp,
                'autoGross': auto_disp,
                'difference': diff_disp,
                'status': status,
                'issues': issues,
                'voucherRow': r,
                'dataRow': r,
            }
        )
    return (
        records,
        {
            'mismatchManualAuto': mismatch_manual_auto,
            'differenceViolations': difference_violations,
            'diffOnlyViolations': diff_only_violations,
        },
        stats,
    )


def _sheet_quick_has_voucher_xlrd(sheet: Any, max_probe_rows: int = 800) -> bool:
    nrows = min(sheet.nrows, max_probe_rows)
    last_col = min(sheet.ncols, MAX_COL_SCAN)
    for r in range(1, nrows + 1):
        for c in range(1, last_col + 1):
            raw = _xlrd_cell_value(sheet, r, c)
            if raw is not None and not _is_blank_cell(raw) and _VOUCHER_MARKER.search(str(raw)):
                return True
    return False


def _scan_xlrd_sheet(sheet: Any, max_row: int) -> list[dict[str, Any]]:
    """xlrd uses 0-based indices; expose 1-based get_cell to shared logic."""
    nrows = sheet.nrows
    scan_upto = min(nrows, max_row)

    def get_cell_1based(row: int, col: int) -> Any:
        return _xlrd_cell_value(sheet, row, col)

    blocks: list[dict[str, Any]] = []
    seen_voucher_keys: set[str] = set()
    r = 1

    while r <= scan_upto:
        voucher_line: str | None = None
        voucher_col = 1
        last_col = min(sheet.ncols, MAX_COL_SCAN)
        for c in range(1, last_col + 1):
            raw = get_cell_1based(r, c)
            if raw is None or _is_blank_cell(raw):
                continue
            text = str(raw).strip()
            if _VOUCHER_MARKER.search(text):
                voucher_line = text
                voucher_col = c
                break
        if voucher_line:
            voucher_no = _extract_voucher_no(voucher_line) or _normalize_spaces(
                voucher_line.replace('Voucher No', '').replace(':', '').strip()
            )
            data_row = _find_data_row(get_cell_1based, r + 1, scan_upto)
            key = f'{r}:{voucher_no}'
            if key not in seen_voucher_keys:
                seen_voucher_keys.add(key)
                blocks.append(
                    {
                        'voucher_row': r,
                        'voucher_col': voucher_col,
                        'voucher_line': voucher_line,
                        'voucher_no': voucher_no,
                        'data_row': data_row,
                    }
                )
            if data_row is not None:
                r = data_row + 1
            else:
                r += 1
            continue
        r += 1

    return blocks


def _build_records_from_blocks(
    blocks: list[dict[str, Any]],
    get_cell: Callable[[int, int], Any],
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    records: list[dict[str, Any]] = []
    mismatch_manual_auto = 0
    difference_violations = 0
    diff_only_violations = 0

    for block in blocks:
        vline = block.get('voucher_line') or ''
        voucher_no = block.get('voucher_no') or ''
        data_row = block.get('data_row')
        issues: list[str] = []

        if data_row is None:
            if _tabular_row_should_skip_non_data_row(vline, None, None, None):
                continue
            issues.append('Could not locate values row after voucher label')
            records.append(
                {
                    'voucherNo': voucher_no,
                    'manualGross': None,
                    'autoGross': None,
                    'difference': None,
                    'status': 'invalid',
                    'issues': issues,
                    'voucherRow': block['voucher_row'],
                    'dataRow': None,
                }
            )
            continue

        manual_raw = get_cell(data_row, COL_MANUAL)
        auto_raw = get_cell(data_row, COL_AUTO)
        diff_raw = get_cell(data_row, COL_DIFF)

        if _tabular_row_should_skip_non_data_row(vline, manual_raw, auto_raw, diff_raw):
            continue

        manual = _parse_number(manual_raw)
        auto = _parse_number(auto_raw)
        diff = _parse_number(diff_raw)

        issues, mm, dv, dov = validate_triplet(manual, auto, diff)
        if mm:
            mismatch_manual_auto += 1
        if dv:
            difference_violations += 1
        if dov:
            diff_only_violations += 1

        manual_disp = display_float_two_dp(manual) if manual is not None else None
        auto_disp = display_float_two_dp(auto) if auto is not None else None
        diff_disp = display_float_two_dp(diff) if diff is not None else None

        status = 'valid' if not issues else 'invalid'
        records.append(
            {
                'voucherNo': voucher_no,
                'manualGross': manual_disp,
                'autoGross': auto_disp,
                'difference': diff_disp,
                'status': status,
                'issues': issues,
                'voucherRow': block['voucher_row'],
                'dataRow': data_row,
            }
        )

    return records, {
        'mismatchManualAuto': mismatch_manual_auto,
        'differenceViolations': difference_violations,
        'diffOnlyViolations': diff_only_violations,
    }


class GrossWeightProcessor(BaseProcessor):
    """Parses tabular or semi-structured gross-weight workbooks; validates manual, auto, and difference."""

    def process(self, file_bytes: bytes, **kwargs: Any) -> dict[str, Any]:
        filename = str(kwargs.get('original_filename') or kwargs.get('filename') or '')
        log.info('Gross weight processing started filename={}', filename or '(none)')

        if not file_bytes:
            raise ValueError('Uploaded file is empty')

        ext = _sniff_extension(file_bytes, filename)

        t_wall = time.perf_counter()
        try:
            if ext == '.xls':
                records, meta, aux = self._process_xls(file_bytes)
            else:
                records, meta, aux = self._process_openxml(file_bytes)
        except ValueError:
            raise
        except Exception as exc:
            log.exception('Gross weight workbook read failed')
            raise ValueError('Invalid or unreadable Excel file') from exc

        wall_ms = (time.perf_counter() - t_wall) * 1000
        total = len(records)
        invalid = sum(1 for r in records if r.get('status') == 'invalid')
        valid = total - invalid
        mismatch_count = meta.get('mismatchManualAuto', 0)
        diff_violations = meta.get('differenceViolations', 0)
        diff_only = meta.get('diffOnlyViolations', 0)

        read_ms = float(aux.get('readTimeMs', 0.0))
        parse_ms = float(aux.get('parseTimeMs', wall_ms))
        validate_ms = float(aux.get('validateTimeMs', 0.0))
        rps = (total / (wall_ms / 1000.0)) if wall_ms > 1e-6 else 0.0

        summary = {
            'total': total,
            'valid': valid,
            'invalid': invalid,
            'mismatchCount': mismatch_count,
            'diffOnlyViolations': diff_only,
            'differenceViolations': diff_violations,
            'layoutMode': aux.get('layoutMode'),
        }

        row_stats = {
            'scanCapTruncated': aux.get('scanCapTruncated', False),
            'tabularRawRowsScanned': aux.get('tabularRawRowsScanned', 0),
            'tabularBlankRowsSkipped': aux.get('tabularBlankRowsSkipped', 0),
            'tabularSkippedNonDataRows': aux.get('tabularSkippedNonDataRows', 0),
            'tabularParsedRows': aux.get('tabularParsedRows', 0),
            'semiVoucherBlocks': aux.get('semiVoucherBlocks', 0),
        }

        performance = {
            'readTimeMs': round(read_ms, 3),
            'parseTimeMs': round(parse_ms, 3),
            'validateTimeMs': round(validate_ms, 3),
            'rowsPerSecond': round(rps, 2),
            'wallTimeMs': round(wall_ms, 3),
        }

        out = build_processing_response(
            file_type='gross_weight',
            total_rows=total,
            error_rows=invalid,
            summary=summary,
            records=records,
            performance=performance,
            row_stats=row_stats,
        )
        out['module'] = 'gross_weight'
        # Lets clients confirm the new parser is live (old builds raised KeyError for missing columns).
        out['layoutEngine'] = 'gross-weight-v2'
        log.info(
            'Gross weight complete total={} valid={} invalid={} wallMs={:.1f} rps={:.1f}',
            total,
            valid,
            invalid,
            wall_ms,
            rps,
        )
        return out

    def _process_openxml(self, file_bytes: bytes) -> tuple[list[dict[str, Any]], dict[str, int], dict[str, Any]]:
        settings = get_settings()
        cap = settings.excel_max_rows
        aux: dict[str, Any] = {
            'scanCapTruncated': False,
            'tabularRawRowsScanned': 0,
            'tabularBlankRowsSkipped': 0,
            'tabularSkippedNonDataRows': 0,
            'tabularParsedRows': 0,
            'semiVoucherBlocks': 0,
            'readTimeMs': 0.0,
            'parseTimeMs': 0.0,
            'validateTimeMs': 0.0,
            'layoutMode': None,
        }

        t_open = time.perf_counter()
        bio = BytesIO(file_bytes)
        wb = load_workbook(bio, read_only=True, data_only=True, keep_links=False)
        open_ms = (time.perf_counter() - t_open) * 1000.0
        aux['readTimeMs'] = open_ms

        try:
            tabular_fallback: tuple[list[dict[str, Any]], dict[str, int], dict[str, int]] | None = None
            for name in wb.sheetnames:
                ws = wb[name]
                ws_max = ws.max_row or 0
                if not ws_max:
                    continue
                eff_max, truncated = effective_excel_max_row(ws_max, cap)
                aux['scanCapTruncated'] = aux['scanCapTruncated'] or truncated

                probe_n = min(ws_max, TABULAR_HEADER_MAX_PROBE_ROW)
                probe_rows = list(
                    ws.iter_rows(
                        min_row=1,
                        max_row=probe_n,
                        min_col=1,
                        max_col=40,
                        values_only=True,
                    )
                )
                probe = [(i + 1, t) for i, t in enumerate(probe_rows)]
                layout = _find_tabular_header_layout_from_probe(probe, 40)
                if layout is None:
                    continue
                hr, vcol, mcol, acol, dcol = layout
                t_body = time.perf_counter()
                records, meta, tstat = _build_records_from_tabular_openpyxl(
                    ws, hr, vcol, mcol, acol, dcol, eff_max
                )
                body_ms = (time.perf_counter() - t_body) * 1000.0
                aux['parseTimeMs'] += body_ms
                for k in tstat:
                    aux[k] = aux.get(k, 0) + tstat[k]

                log.info(
                    'Gross weight tabular sheet={} header_row={} voucher_col={} manual={} auto={} diff={} rows={}',
                    name,
                    hr,
                    vcol,
                    mcol,
                    acol,
                    dcol,
                    len(records),
                )
                if records:
                    aux['layoutMode'] = 'tabular'
                    return records, meta, aux
                if tabular_fallback is None:
                    tabular_fallback = (records, meta, tstat)

            if tabular_fallback is not None:
                rec, meta, tstat = tabular_fallback
                for k in tstat:
                    aux[k] = aux.get(k, 0) + tstat[k]
                aux['layoutMode'] = 'tabular_empty'
                return rec, meta, aux

            ws = _pick_worksheet_openpyxl(wb)
            ws_max = ws.max_row or 0
            eff_max, truncated = effective_excel_max_row(ws_max, cap)
            aux['scanCapTruncated'] = aux['scanCapTruncated'] or truncated
            t_semi = time.perf_counter()
            blocks = _scan_openpyxl_sheet(ws, eff_max)
            aux['semiVoucherBlocks'] = len(blocks)

            def get_cell(r: int, c: int) -> Any:
                return _row_get_cell_openpyxl(ws, r, c)

            records, meta = _build_records_from_blocks(blocks, get_cell)
            aux['parseTimeMs'] += (time.perf_counter() - t_semi) * 1000.0
            aux['layoutMode'] = 'semi_structured'
            return records, meta, aux
        finally:
            wb.close()

    def _process_xls(self, file_bytes: bytes) -> tuple[list[dict[str, Any]], dict[str, int], dict[str, Any]]:
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError('Reading .xls requires the xlrd package') from exc

        settings = get_settings()
        cap = settings.excel_max_rows
        aux: dict[str, Any] = {
            'scanCapTruncated': False,
            'tabularRawRowsScanned': 0,
            'tabularBlankRowsSkipped': 0,
            'tabularSkippedNonDataRows': 0,
            'tabularParsedRows': 0,
            'semiVoucherBlocks': 0,
            'readTimeMs': 0.0,
            'parseTimeMs': 0.0,
            'validateTimeMs': 0.0,
            'layoutMode': None,
        }

        t_open = time.perf_counter()
        book = xlrd.open_workbook(file_contents=file_bytes, formatting_info=False)
        aux['readTimeMs'] = (time.perf_counter() - t_open) * 1000.0

        tabular_fallback: tuple[list[dict[str, Any]], dict[str, int], dict[str, int]] | None = None
        for i in range(book.nsheets):
            sh = book.sheet_by_index(i)
            if sh.nrows <= 0:
                continue
            layout = _find_tabular_header_layout_xlrd(sh)
            if layout is None:
                continue
            hr, vcol, mcol, acol, dcol = layout
            eff_max, truncated = effective_excel_max_row(sh.nrows, cap)
            aux['scanCapTruncated'] = aux['scanCapTruncated'] or truncated
            t_body = time.perf_counter()
            records, meta, tstat = _build_records_from_tabular_xlrd(
                sh, hr, vcol, mcol, acol, dcol, eff_max
            )
            aux['parseTimeMs'] += (time.perf_counter() - t_body) * 1000.0
            for k in tstat:
                aux[k] = aux.get(k, 0) + tstat[k]
            log.info('Gross weight tabular .xls sheet_index={} rows={}', i, len(records))
            if records:
                aux['layoutMode'] = 'tabular'
                return records, meta, aux
            if tabular_fallback is None:
                tabular_fallback = (records, meta, tstat)

        if tabular_fallback is not None:
            rec, meta, tstat = tabular_fallback
            for k in tstat:
                aux[k] = aux.get(k, 0) + tstat[k]
            aux['layoutMode'] = 'tabular_empty'
            return rec, meta, aux

        sheet = None
        for i in range(book.nsheets):
            sh = book.sheet_by_index(i)
            if sh.nrows <= 0:
                continue
            if _sheet_quick_has_voucher_xlrd(sh):
                sheet = sh
                break
        if sheet is None:
            for i in range(book.nsheets):
                sh = book.sheet_by_index(i)
                if sh.nrows > 0:
                    sheet = sh
                    break
        if sheet is None:
            aux['layoutMode'] = 'empty'
            return [], {'mismatchManualAuto': 0, 'differenceViolations': 0, 'diffOnlyViolations': 0}, aux

        eff_max, truncated = effective_excel_max_row(sheet.nrows, cap)
        aux['scanCapTruncated'] = aux['scanCapTruncated'] or truncated
        t_semi = time.perf_counter()
        blocks = _scan_xlrd_sheet(sheet, eff_max)
        aux['semiVoucherBlocks'] = len(blocks)

        def get_cell(r: int, c: int) -> Any:
            return _xlrd_cell_value(sheet, r, c)

        records, meta = _build_records_from_blocks(blocks, get_cell)
        aux['parseTimeMs'] += (time.perf_counter() - t_semi) * 1000.0
        aux['layoutMode'] = 'semi_structured'
        return records, meta, aux
