from __future__ import annotations

from functools import lru_cache
from pathlib import Path
import re
from typing import Any

import polars as pl
from openpyxl import load_workbook

from app.utils.header_cleaner import normalize_header
from app.utils.logger import get_logger
from app.utils.normalization_engine import normalize_blankable_text, normalize_blankable_text_expr

MASTER_RULE_WORKBOOK = Path(__file__).resolve().parents[1] / 'data' / 'master_sales_rules.xlsx'
MASTER_RULE_DEBUG_OUTPUT = (
    Path(__file__).resolve().parents[1] / 'debug' / 'flattened_master_rules.csv'
)
_MASTER_REQUIRED_HEADERS = frozenset({'sales_account_type', 'product', 'category', 'status'})
_INACTIVE_STATUSES = frozenset({'INACTIVE', 'DISABLED', 'ARCHIVED'})
_IGNORE_ROW_PATTERNS = (
    'SALES RATE VERIFICATION',
    'ENTERPRISE MASTER SALES VERIFICATION',
    'MASTER SALES RULE',
    'MASTER RULE',
    'NOTE',
    'NOTES',
    'REMARK',
    'REMARKS',
    'META',
    'UPDATED ON',
    'DATE:',
)
_DECORATIVE_ROW_RE = re.compile(r'^[\-\_=*#~. ]+$')


class MasterRuleService:
    def __init__(
        self,
        workbook_path: Path | None = None,
        debug_output_path: Path | None = None,
    ) -> None:
        self.workbook_path = workbook_path or MASTER_RULE_WORKBOOK
        self.debug_output_path = debug_output_path or MASTER_RULE_DEBUG_OUTPUT

    def load_master_rules(self) -> pl.DataFrame:
        stat = self.workbook_path.stat()
        return _load_master_rules_cached(
            str(self.workbook_path),
            str(self.debug_output_path),
            stat.st_mtime_ns,
        )


@lru_cache(maxsize=4)
def _load_master_rules_cached(
    workbook_path: str, debug_output_path: str, workbook_mtime_ns: int
) -> pl.DataFrame:
    del workbook_mtime_ns
    log = get_logger()
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f'Master sales rules workbook not found: {path}')

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
        header_row_index = _find_header_row_index(rows)
        if header_row_index is None:
            raise ValueError(
                'Master sales rules workbook is missing required headers: '
                + ', '.join(sorted(_MASTER_REQUIRED_HEADERS))
            )

        header_map = _header_map(rows[header_row_index])
        log.info(
            f"[master_rules] detected header row={header_row_index + 1} columns={sorted(header_map)}"
        )
        raw_frame = _raw_master_frame(rows, header_row_index, header_map)
        _log_parent_rows(log, raw_frame)
        dataframe = _flatten_rows_with_header(raw_frame, log)
        _log_master_integrity(log, raw_frame, dataframe)
        _validate_master_integrity(dataframe)
        debug_path = Path(debug_output_path)
        debug_path.parent.mkdir(parents=True, exist_ok=True)
        dataframe.write_csv(debug_path)
        log.info(
            f"[master_rules] total valid flattened rules={len(dataframe)} "
            f"debug_output={debug_path.as_posix()}"
        )
        return dataframe
    finally:
        workbook.close()


def _header_matches(row: tuple[Any, ...]) -> bool:
    labels = {normalize_header(value) for value in row if normalize_header(value)}
    return _MASTER_REQUIRED_HEADERS <= labels


def _header_map(row: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        header = normalize_header(value)
        if header and header not in mapping:
            mapping[header] = index
    return mapping


def _find_header_row_index(rows: list[tuple[Any, ...]], scan_limit: int = 80) -> int | None:
    for index, row in enumerate(rows[:scan_limit]):
        if _header_matches(row):
            return index
    return None


def _raw_master_frame(
    rows: list[tuple[Any, ...]], header_row_index: int, header_map: dict[str, int]
) -> pl.DataFrame:
    records: list[dict[str, Any]] = []
    for absolute_index, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        texts = _normalized_texts(row)
        if _should_ignore_row(texts):
            continue
        records.append(
            {
                'row_number': absolute_index,
                'sales_account': _cell_value(row, header_map, 'sales_account_type'),
                'product': _cell_value(row, header_map, 'product'),
                'category': _cell_value(row, header_map, 'category'),
                'status': _cell_value(row, header_map, 'status'),
            }
        )
    if not records:
        return pl.DataFrame(
            schema={
                'row_number': pl.Int64,
                'sales_account': pl.Utf8,
                'product': pl.Utf8,
                'category': pl.Utf8,
                'status': pl.Utf8,
            }
        )
    return pl.DataFrame(records, strict=False)


def _flatten_rows_with_header(raw_frame: pl.DataFrame, log: Any) -> pl.DataFrame:
    if raw_frame.is_empty():
        return _empty_master_rule_frame()

    normalized = raw_frame.with_columns(
        [
            normalize_blankable_text_expr('sales_account').alias('sales_account'),
            normalize_blankable_text_expr('product').alias('product'),
            normalize_blankable_text_expr('category').alias('category'),
            normalize_blankable_text_expr('status').alias('status'),
        ]
    )
    forward_filled = normalized.with_columns(
        [
            pl.col('sales_account').fill_null(strategy='forward').alias('sales_account'),
            pl.col('category').fill_null(strategy='forward').alias('category'),
            pl.col('status').fill_null(strategy='forward').alias('status'),
        ]
    ).with_columns(pl.col('status').fill_null('ACTIVE').alias('status'))

    skipped_rows = forward_filled.filter(
        pl.col('sales_account').is_null()
        | pl.col('product').is_null()
        | pl.col('status').is_in(sorted(_INACTIVE_STATUSES))
    )
    for row in skipped_rows.select('row_number', 'sales_account', 'product', 'status').to_dicts():
        reason = (
            f"skipped inactive row status={row['status']}"
            if row.get('status') in _INACTIVE_STATUSES
            else 'missing sales account or product after forward-fill'
        )
        _log_skipped_row(log, int(row['row_number']), reason, tuple(v for v in row.values()))

    valid = (
        forward_filled.filter(
            pl.col('sales_account').is_not_null()
            & pl.col('product').is_not_null()
            & (~pl.col('status').is_in(sorted(_INACTIVE_STATUSES)))
        )
        .with_columns(
            [
                pl.col('sales_account').alias('normalized_sales_account'),
                pl.col('product').alias('normalized_product'),
                pl.col('category').fill_null('').alias('category'),
            ]
        )
        .unique(subset=['sales_account', 'product'], maintain_order=True)
    )
    for row in valid.select('row_number', 'sales_account', 'product', 'category', 'status').head(25).to_dicts():
        _log_child_row(
            log,
            row_number=int(row['row_number']),
            sales_account=row['sales_account'],
            product=row['product'],
            category=row.get('category'),
            status=row.get('status') or 'ACTIVE',
            inherited=True,
        )
    return valid.select(
        [
            'sales_account',
            'product',
            'category',
            'status',
            'normalized_sales_account',
            'normalized_product',
        ]
    )


def _cell_value(row: tuple[Any, ...], header_map: dict[str, int], header: str) -> Any:
    index = header_map.get(header)
    if index is None or index >= len(row):
        return None
    return row[index]


def _master_rule_schema() -> dict[str, pl.DataType]:
    return {
        'sales_account': pl.Utf8,
        'product': pl.Utf8,
        'category': pl.Utf8,
        'status': pl.Utf8,
        'normalized_sales_account': pl.Utf8,
        'normalized_product': pl.Utf8,
    }


def _empty_master_rule_frame() -> pl.DataFrame:
    return pl.DataFrame(schema=_master_rule_schema())


def _normalized_texts(row: tuple[Any, ...]) -> list[str]:
    return [text for value in row if (text := normalize_blankable_text(value)) is not None]


def _should_ignore_row(texts: list[str]) -> bool:
    if not texts:
        return True
    combined = ' '.join(texts)
    if _DECORATIVE_ROW_RE.fullmatch(combined):
        return True
    return any(token in combined for token in _IGNORE_ROW_PATTERNS)


def _log_parent_detected(
    log: Any,
    row_number: int,
    sales_account: str,
    category: str | None,
    status: str,
) -> None:
    log.info(
        f"[master_rules] detected parent account row={row_number} sales_account={sales_account!r} "
        f"category={category!r} status={status!r}"
    )


def _log_parent_rows(log: Any, raw_frame: pl.DataFrame) -> None:
    if raw_frame.is_empty():
        return
    parent_rows = raw_frame.with_columns(
        [
            normalize_blankable_text_expr('sales_account').alias('sales_account'),
            normalize_blankable_text_expr('product').alias('product'),
            normalize_blankable_text_expr('category').alias('category'),
            normalize_blankable_text_expr('status').alias('status'),
        ]
    ).filter(pl.col('sales_account').is_not_null())
    for row in parent_rows.select('row_number', 'sales_account', 'category', 'status').to_dicts():
        _log_parent_detected(
            log,
            int(row['row_number']),
            row['sales_account'],
            row.get('category'),
            row.get('status') or 'ACTIVE',
        )


def _log_child_row(
    log: Any,
    row_number: int,
    sales_account: str,
    product: str,
    category: str | None,
    status: str,
    *,
    inherited: bool,
) -> None:
    action = 'inherited child row' if inherited else 'flattened direct row'
    log.info(
        f"[master_rules] {action} row={row_number} sales_account={sales_account!r} "
        f"product={product!r} category={category!r} status={status!r}"
    )


def _log_skipped_row(log: Any, row_number: int, reason: str, row: tuple[Any, ...]) -> None:
    log.info(
        f"[master_rules] skipped invalid row row={row_number} reason={reason} "
        f"values={tuple(value for value in row if value not in (None, ''))!r}"
    )


def _log_master_integrity(log: Any, raw_frame: pl.DataFrame, flattened_frame: pl.DataFrame) -> None:
    sales_account_nulls = int(raw_frame['sales_account'].null_count()) if 'sales_account' in raw_frame.columns else 0
    product_nulls = int(raw_frame['product'].null_count()) if 'product' in raw_frame.columns else 0
    log.info(
        '[master_rules] integrity total_master_rows={total_rows} distinct_sales_accounts={accounts} '
        'distinct_products={products} sales_account_nulls={sales_account_nulls} '
        'product_nulls={product_nulls}'.format(
            total_rows=raw_frame.height,
            accounts=flattened_frame['sales_account'].n_unique() if not flattened_frame.is_empty() else 0,
            products=flattened_frame['product'].n_unique() if not flattened_frame.is_empty() else 0,
            sales_account_nulls=sales_account_nulls,
            product_nulls=product_nulls,
        )
    )


def _validate_master_integrity(flattened_frame: pl.DataFrame) -> None:
    if flattened_frame.is_empty():
        raise ValueError('Master sales rules workbook produced zero valid flattened rules.')
    if flattened_frame['sales_account'].n_unique() == 0:
        raise ValueError('Master sales rules workbook has zero distinct sales accounts.')
    if flattened_frame['product'].n_unique() == 0:
        raise ValueError('Master sales rules workbook has zero distinct products.')
