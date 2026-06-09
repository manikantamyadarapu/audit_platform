from __future__ import annotations

from functools import lru_cache
from pathlib import Path
import re
from typing import Any

import polars as pl
from openpyxl import load_workbook

from app.utils.header_cleaner import normalize_header
from app.utils.logger import get_logger
from app.utils.normalization_engine import normalize_blankable_text, normalize_blankable_text_expr

MASTER_RATE_WORKBOOK = Path(__file__).resolve().parents[1] / 'data' / 'master_sales_rate_rules.xlsx'
MASTER_RATE_DEBUG_OUTPUT = Path(__file__).resolve().parents[1] / 'debug' / 'flattened_sales_rate_rules.csv'
_RATE_REQUIRED_HEADERS = frozenset(
    {
        'sales_account_type',
        'product',
        'standard_rate',
        'allowed_deviation_percent',
        'minimum_allowed_rate',
        'maximum_allowed_rate',
        'status',
    }
)
_INACTIVE_STATUSES = frozenset({'INACTIVE', 'DISABLED', 'ARCHIVED'})
_IGNORE_ROW_PATTERNS = (
    'SALES RATE VERIFICATION',
    'ENTERPRISE MASTER SALES RATE',
    'MASTER SALES RATE',
    'MASTER RATE',
    'NOTE',
    'NOTES',
    'REMARK',
    'REMARKS',
    'META',
    'UPDATED ON',
    'DATE:',
)
_DECORATIVE_ROW_RE = re.compile(r'^[\-\_=*#~. ]+$')


class MasterSalesRateRuleService:
    def __init__(
        self,
        workbook_path: Path | None = None,
        debug_output_path: Path | None = None,
    ) -> None:
        self.workbook_path = workbook_path or MASTER_RATE_WORKBOOK
        self.debug_output_path = debug_output_path or MASTER_RATE_DEBUG_OUTPUT

    def load_rate_rules(self) -> pl.DataFrame:
        stat = self.workbook_path.stat()
        return _load_sales_rate_rules_cached(
            str(self.workbook_path),
            str(self.debug_output_path),
            stat.st_mtime_ns,
        )


@lru_cache(maxsize=4)
def _load_sales_rate_rules_cached(
    workbook_path: str, debug_output_path: str, workbook_mtime_ns: int
) -> pl.DataFrame:
    del workbook_mtime_ns
    log = get_logger()
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f'Master sales rate rules workbook not found: {path}')

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
        header_row_index = _find_header_row_index(rows)
        if header_row_index is None:
            raise ValueError(
                'Master sales rate rules workbook is missing required headers: '
                + ', '.join(sorted(_RATE_REQUIRED_HEADERS))
            )
        header_map = _header_map(rows[header_row_index])
        log.info(
            f'[sales_rate_rules] detected header row={header_row_index + 1} '
            f'columns={sorted(header_map)}'
        )
        raw_frame = _raw_rate_frame(rows, header_row_index, header_map)
        dataframe = _flatten_rate_rows(raw_frame, log)
        debug_path = Path(debug_output_path)
        debug_path.parent.mkdir(parents=True, exist_ok=True)
        dataframe.write_csv(debug_path)
        log.info(
            f'[sales_rate_rules] total valid rate rules={len(dataframe)} '
            f'debug_output={debug_path.as_posix()}'
        )
        return dataframe
    finally:
        workbook.close()


def _header_matches(row: tuple[Any, ...]) -> bool:
    labels = {normalize_header(value) for value in row if normalize_header(value)}
    return _RATE_REQUIRED_HEADERS <= labels


def _header_map(row: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        header = normalize_header(value)
        if header and header not in mapping:
            mapping[header] = index
    return mapping


def _find_header_row_index(rows: list[tuple[Any, ...]], scan_limit: int = 80) -> int | None:
    for index, row in enumerate(rows[:scan_limit]):
        if _header_matches(row):
            return index
    return None


def _raw_rate_frame(
    rows: list[tuple[Any, ...]], header_row_index: int, header_map: dict[str, int]
) -> pl.DataFrame:
    records: list[dict[str, Any]] = []
    for absolute_index, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        texts = _normalized_texts(row)
        if _should_ignore_row(texts):
            continue
        records.append(
            {
                'row_number': absolute_index,
                'sales_account': _cell_value(row, header_map, 'sales_account_type'),
                'product': _cell_value(row, header_map, 'product'),
                'standard_rate': _cell_value(row, header_map, 'standard_rate'),
                'allowed_deviation_percent': _cell_value(row, header_map, 'allowed_deviation_percent'),
                'minimum_allowed_rate': _cell_value(row, header_map, 'minimum_allowed_rate'),
                'maximum_allowed_rate': _cell_value(row, header_map, 'maximum_allowed_rate'),
                'status': _cell_value(row, header_map, 'status'),
            }
        )
    if not records:
        return pl.DataFrame(
            schema={
                'row_number': pl.Int64,
                'sales_account': pl.Utf8,
                'product': pl.Utf8,
                'standard_rate': pl.Float64,
                'allowed_deviation_percent': pl.Float64,
                'minimum_allowed_rate': pl.Float64,
                'maximum_allowed_rate': pl.Float64,
                'status': pl.Utf8,
            }
        )
    return pl.DataFrame(records, strict=False)


def _cell_value(row: tuple[Any, ...], header_map: dict[str, int], header: str) -> Any:
    index = header_map.get(header)
    if index is None or index >= len(row):
        return None
    return row[index]


def _flatten_rate_rows(raw_frame: pl.DataFrame, log: Any) -> pl.DataFrame:
    if raw_frame.is_empty():
        raise ValueError('Master sales rate rules workbook produced zero raw rows.')

    normalized = raw_frame.with_columns(
        [
            normalize_blankable_text_expr('sales_account').alias('sales_account'),
            normalize_blankable_text_expr('product').alias('product'),
            normalize_blankable_text_expr('status').alias('status'),
            pl.col('standard_rate').cast(pl.Float64, strict=False),
            pl.col('allowed_deviation_percent').cast(pl.Float64, strict=False),
            pl.col('minimum_allowed_rate').cast(pl.Float64, strict=False),
            pl.col('maximum_allowed_rate').cast(pl.Float64, strict=False),
        ]
    )
    forward_filled = normalized.with_columns(
        [
            pl.col('sales_account').fill_null(strategy='forward').alias('sales_account'),
            pl.col('status').fill_null(strategy='forward').alias('status'),
        ]
    ).with_columns(pl.col('status').fill_null('ACTIVE').alias('status'))

    skipped = forward_filled.filter(
        pl.col('sales_account').is_null()
        | pl.col('product').is_null()
        | pl.col('standard_rate').is_null()
        | pl.col('status').is_in(sorted(_INACTIVE_STATUSES))
    )
    for row in skipped.select('row_number', 'sales_account', 'product', 'status').to_dicts():
        reason = (
            f'skipped inactive row status={row["status"]}'
            if row.get('status') in _INACTIVE_STATUSES
            else 'missing sales account, product, or standard rate after forward-fill'
        )
        log.info(
            f'[sales_rate_rules] skipped row row={row["row_number"]} reason={reason} '
            f'values={tuple(v for v in row.values() if v not in (None, ""))!r}'
        )

    valid = forward_filled.filter(
        pl.col('sales_account').is_not_null()
        & pl.col('product').is_not_null()
        & pl.col('standard_rate').is_not_null()
        & (~pl.col('status').is_in(sorted(_INACTIVE_STATUSES)))
    ).with_columns(
        [
            pl.col('allowed_deviation_percent').fill_null(30.0).alias('allowed_deviation_percent'),
        ]
    )
    # Enterprise band: standard rate ±30% from master standard only (not upload-derived).
    valid = valid.with_columns(
        [
            (pl.col('standard_rate') * 0.70).alias('min_allowed_rate'),
            (pl.col('standard_rate') * 1.30).alias('max_allowed_rate'),
        ]
    )
    out = (
        valid.with_columns(
            [
                pl.col('sales_account').alias('normalized_sales_account'),
                pl.col('product').alias('normalized_product'),
            ]
        )
        .unique(subset=['sales_account', 'product'], maintain_order=True)
        .select(
            [
                'normalized_sales_account',
                'normalized_product',
                'standard_rate',
                'allowed_deviation_percent',
                'min_allowed_rate',
                'max_allowed_rate',
            ]
        )
    )
    if out.is_empty():
        raise ValueError('Master sales rate rules workbook produced zero valid rate rules.')
    return out


def _normalized_texts(row: tuple[Any, ...]) -> list[str]:
    return [text for value in row if (text := normalize_blankable_text(value)) is not None]


def _should_ignore_row(texts: list[str]) -> bool:
    if not texts:
        return True
    combined = ' '.join(texts)
    if _DECORATIVE_ROW_RE.fullmatch(combined):
        return True
    return any(token in combined for token in _IGNORE_ROW_PATTERNS)
