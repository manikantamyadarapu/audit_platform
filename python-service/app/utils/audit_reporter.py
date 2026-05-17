from __future__ import annotations

import json
from collections import Counter
from io import BytesIO
from time import perf_counter
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from app.core.issue_engine import build_issue, issue_message

SUMMARY_SHEET = 'Summary'
ISSUE_BREAKDOWN_SHEET = 'Issue Breakdown'
ISSUE_GROUPING_SHEET = 'Issue Grouping'
PROCESSING_STATS_SHEET = 'Processing Statistics'
EXECUTION_TIMING_SHEET = 'Execution Timing'

_HEADER_FILL = PatternFill(fill_type='solid', fgColor='1F4E78')
_HEADER_FONT = Font(color='FFFFFF', bold=True)


def build_audit_excel_report(
    *,
    report_title: str,
    invalid_sheet_name: str,
    source_processor: str,
    records: list[dict[str, Any]],
    export_columns: list[str],
    summary: dict[str, Any] | None = None,
    processing_statistics: dict[str, Any] | None = None,
    execution_timing: dict[str, Any] | None = None,
) -> bytes:
    if not records:
        raise ValueError('No invalid records found to export')

    report_build_start = perf_counter()
    invalid_rows_df = _build_invalid_rows_frame(records, export_columns)
    issue_rows = _build_issue_rows(records, source_processor=source_processor)
    issue_breakdown_df = _build_issue_breakdown_frame(issue_rows)
    issue_grouping_df = _build_issue_grouping_frame(records)
    processing_stats_df = _build_key_value_frame(
        _derive_processing_statistics(records, summary, processing_statistics)
    )
    summary_df = _build_key_value_frame(
        _derive_summary_rows(
            report_title=report_title,
            source_processor=source_processor,
            records=records,
            summary=summary,
        )
    )
    execution_timing_payload = dict(execution_timing or {})
    execution_timing_payload.setdefault(
        'reportGenerationMs', round((perf_counter() - report_build_start) * 1000, 2)
    )
    execution_timing_df = _build_key_value_frame(execution_timing_payload)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        summary_df.to_excel(writer, index=False, sheet_name=SUMMARY_SHEET)
        issue_breakdown_df.to_excel(writer, index=False, sheet_name=ISSUE_BREAKDOWN_SHEET)
        issue_grouping_df.to_excel(writer, index=False, sheet_name=ISSUE_GROUPING_SHEET)
        processing_stats_df.to_excel(writer, index=False, sheet_name=PROCESSING_STATS_SHEET)
        execution_timing_df.to_excel(writer, index=False, sheet_name=EXECUTION_TIMING_SHEET)
        invalid_rows_df.to_excel(writer, index=False, sheet_name=invalid_sheet_name)

        workbook = writer.book
        header_format = workbook.add_format(
            {
                'bold': True,
                'font_color': 'white',
                'bg_color': '#1F4E78',
                'border': 1,
            }
        )
        for sheet_name, dataframe in (
            (SUMMARY_SHEET, summary_df),
            (ISSUE_BREAKDOWN_SHEET, issue_breakdown_df),
            (ISSUE_GROUPING_SHEET, issue_grouping_df),
            (PROCESSING_STATS_SHEET, processing_stats_df),
            (EXECUTION_TIMING_SHEET, execution_timing_df),
            (invalid_sheet_name, invalid_rows_df),
        ):
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes(1, 0)
            worksheet.autofilter(0, 0, max(len(dataframe), 1), max(len(dataframe.columns) - 1, 0))
            for idx, column in enumerate(dataframe.columns):
                width = max(len(str(column)), _max_value_length(dataframe[column])) + 2
                worksheet.set_column(idx, idx, min(width, 60))
                worksheet.write(0, idx, column, header_format)

    output.seek(0)
    return _finalize_workbook(output.getvalue())


def _build_invalid_rows_frame(
    records: list[dict[str, Any]], export_columns: list[str]
) -> pd.DataFrame:
    dataframe = pd.DataFrame(records).copy()
    for column in export_columns:
        if column not in dataframe.columns:
            dataframe[column] = ''
    if 'issues' in dataframe.columns:
        dataframe['issues'] = dataframe['issues'].apply(_stringify_listish)
    if 'messages' in dataframe.columns:
        dataframe['messages'] = dataframe['messages'].apply(_stringify_listish)
    return dataframe[export_columns]


def _build_issue_rows(
    records: list[dict[str, Any]], *, source_processor: str
) -> list[dict[str, Any]]:
    issue_rows: list[dict[str, Any]] = []
    for record in records:
        row_number = _safe_int(record.get('rowNumber'))
        issue_codes = _as_list(record.get('issues'))
        messages = _as_list(record.get('messages'))
        for index, issue_code in enumerate(issue_codes):
            message = messages[index] if index < len(messages) and messages[index] else issue_message(issue_code)
            issue = build_issue(
                issue_code,
                row_number=row_number,
                source_processor=source_processor,
                message=message,
                metadata={'record': _json_safe(record)},
                processor_stage='export',
            )
            issue_rows.append(issue.to_dict())
    return issue_rows


def _build_issue_breakdown_frame(issue_rows: list[dict[str, Any]]) -> pd.DataFrame:
    if not issue_rows:
        return pd.DataFrame(
            columns=[
                'issue_code',
                'severity',
                'category',
                'message',
                'occurrence_count',
                'affected_rows',
                'source_processor',
            ]
        )

    grouped: dict[tuple[str, str, str, str, str], set[int]] = {}
    counts: Counter[tuple[str, str, str, str, str]] = Counter()
    for row in issue_rows:
        key = (
            str(row['issue_code']),
            str(row['severity']),
            str(row['category']),
            str(row['message']),
            str(row['source_processor']),
        )
        counts[key] += 1
        grouped.setdefault(key, set())
        if row.get('row_number') is not None:
            grouped[key].add(int(row['row_number']))

    breakdown_rows = [
        {
            'issue_code': key[0],
            'severity': key[1],
            'category': key[2],
            'message': key[3],
            'occurrence_count': count,
            'affected_rows': len(grouped[key]),
            'source_processor': key[4],
        }
        for key, count in counts.items()
    ]
    breakdown_rows.sort(key=lambda row: (-int(row['occurrence_count']), str(row['issue_code'])))
    return pd.DataFrame(breakdown_rows)


def _build_issue_grouping_frame(records: list[dict[str, Any]]) -> pd.DataFrame:
    groups: dict[str, dict[str, Any]] = {}
    for record in records:
        issues = _as_list(record.get('issues'))
        group_key = ' | '.join(issues) if issues else 'NO_ISSUES'
        row_number = _safe_int(record.get('rowNumber'))
        entry = groups.setdefault(
            group_key,
            {
                'issue_group': group_key,
                'row_count': 0,
                'sample_rows': [],
                'sample_messages': [],
            },
        )
        entry['row_count'] += 1
        if row_number is not None and len(entry['sample_rows']) < 5:
            entry['sample_rows'].append(row_number)
        messages = _as_list(record.get('messages'))
        for message in messages:
            if message and message not in entry['sample_messages'] and len(entry['sample_messages']) < 3:
                entry['sample_messages'].append(message)

    grouping_rows = [
        {
            'issue_group': value['issue_group'],
            'row_count': value['row_count'],
            'sample_rows': ', '.join(str(item) for item in value['sample_rows']),
            'sample_messages': ' | '.join(value['sample_messages']),
        }
        for value in groups.values()
    ]
    grouping_rows.sort(key=lambda row: (-int(row['row_count']), str(row['issue_group'])))
    return pd.DataFrame(grouping_rows)


def _derive_processing_statistics(
    records: list[dict[str, Any]],
    summary: dict[str, Any] | None,
    processing_statistics: dict[str, Any] | None,
) -> dict[str, Any]:
    supplied = dict(processing_statistics or {})
    issue_codes = [code for record in records for code in _as_list(record.get('issues'))]
    supplied.setdefault('invalidRowCount', len(records))
    supplied.setdefault('totalIssueCount', len(issue_codes))
    supplied.setdefault('uniqueIssueCodeCount', len(set(issue_codes)))
    supplied.setdefault(
        'rowsWithMultipleIssues',
        sum(1 for record in records if len(_as_list(record.get('issues'))) > 1),
    )
    if summary:
        supplied.setdefault('summaryFieldCount', len(summary))
    return supplied


def _derive_summary_rows(
    *,
    report_title: str,
    source_processor: str,
    records: list[dict[str, Any]],
    summary: dict[str, Any] | None,
) -> dict[str, Any]:
    issue_codes = [code for record in records for code in _as_list(record.get('issues'))]
    summary_rows: dict[str, Any] = {
        'reportTitle': report_title,
        'sourceProcessor': source_processor,
        'generatedAt': pd.Timestamp.now(tz='UTC').isoformat(),
        'invalidRowCount': len(records),
        'totalIssueCount': len(issue_codes),
        'uniqueIssueCodes': ', '.join(sorted(set(issue_codes))),
    }
    if summary:
        for key, value in summary.items():
            summary_rows[f'summary.{key}'] = _json_safe(value)
    return summary_rows


def _build_key_value_frame(payload: dict[str, Any]) -> pd.DataFrame:
    rows = [{'metric': key, 'value': _json_safe(value)} for key, value in payload.items()]
    return pd.DataFrame(rows or [{'metric': '', 'value': ''}])


def _finalize_workbook(workbook_bytes: bytes) -> bytes:
    workbook = load_workbook(BytesIO(workbook_bytes))
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = 'A2'
        for cell in worksheet[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        _autosize_columns(worksheet)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


def _autosize_columns(worksheet: Any) -> None:
    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _stringify_listish(value: Any) -> str:
    if isinstance(value, list):
        return ', '.join(str(item) for item in value)
    if value is None:
        return ''
    return str(value)


def _as_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value if item is not None and str(item)]
    if value is None:
        return []
    text = str(value).strip()
    return [text] if text else []


def _safe_int(value: Any) -> int | None:
    if value is None or value == '':
        return None
    return int(value)


def _json_safe(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    try:
        return json.dumps(value, default=str)
    except TypeError:
        return str(value)


def _max_value_length(series: pd.Series) -> int:
    if series.empty:
        return 0
    return max(len(str(value)) for value in series.fillna('').tolist())
